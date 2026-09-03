import math
import zipfile
import base64
import csv
import json
import requests as http_requests
from django.db.models import Sum
from django.db.models.functions import Coalesce, TruncMonth
from decimal import Decimal
from django.core.signing import TimestampSigner, BadSignature, SignatureExpired
from datetime import date,  timedelta, time as datetime_time
from io import BytesIO
from django.contrib.auth import logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import UserCreationForm
from django.contrib.sessions.models import Session
from django.contrib.sites import requests
from django.core.paginator import Paginator
from django.db import IntegrityError
from django.db.models import  Sum, F
from django.template.loader import get_template
from django.utils.dateparse import parse_date, parse_datetime
from django.views.decorators.cache import never_cache
from django.views.decorators.clickjacking import xframe_options_sameorigin
from django.views.decorators.http import require_POST
from psycopg import rows
from xhtml2pdf import pisa
from rbac.templatetags.menu_tags import build_menu_tree
from rbac.models import Menu
from config import settings
from .forms import VehicleForm, InsuranceCompanyForm, CustomerForm, SurveyorForm, EmployeeForm, JobCardForm, DriverMasterForm
from .models import (
    InsuranceCompany, VehicleModel, Customer, ColumnPreference, Surveyor, JobCardPart,
    JobCardLabour, JobCardAssessmentPart, JobCardAssessmentLabour, JobCardTyreInventory,
    CommunicationLog, UserNotification, ClaimStageCode, WorkProgress, WorkAllocation,
    AnnouncementRead, Announcement,  PartOrderHeader, WorkAllocationPart,
    WorkAllocationLabour, JobCardReInspectionPhoto, JobCardVehicleConditionPhoto,
    ClaimDocument, WorkProgressPhoto, JobCardAdditionalApprovalPhoto, Branch, GateInEntry,
    UserLoginActivity, PartStockTransaction, PartRequisition, PartRequisitionLine,
    PartRequisitionFulfillment,
    CustomerApprovalEvidence,
    CustomerApprovalAttachment,
    JobCardInventory, JobCardVehicleConditionPhoto, CompanySetup,
    CustomerApprovalPhotoAnnotation,
    JobCardDamageAISuggestion,
    ClaimTimeline, DriverMaster,
)
from rbac.models import (
    Menu,
    RoleMenuPermission,
    UserMenuPermission,
)
from .whatsapp import send_whatsapp_template_message
from .services.damage_ai import get_damage_ai_provider, DamageAIConfigurationError
from .numbering import branch_for_claim, branch_for_user, next_claim_no, next_jobcard_no
from .validators import VEHICLE_NUMBER_ERROR, is_valid_vehicle_number, normalize_vehicle_number
from apps.claims.services.repair_workflow_service import (
    RepairWorkflowBlocked,
    RepairWorkflowService,
)
from apps.claims.services.dashboard_financial_service import DashboardFinancialService
from apps.claims.repositories.dashboard_repository import DashboardLookupRepository
from apps.claims.services.dashboard_metrics_service import DashboardMetricsService
from apps.claims.services.advisor_dashboard_service import AdvisorDashboardReadService
from apps.notifications.services.notification_service import notify_jobcard_advisor


REINSPECTION_MAX_PHOTOS_PER_JOBCARD = getattr(settings, "REINSPECTION_MAX_PHOTOS_PER_JOBCARD", 25)
REINSPECTION_MAX_IMAGE_SIZE_MB = getattr(settings, "REINSPECTION_MAX_IMAGE_SIZE_MB", 8)
REINSPECTION_MAX_TOTAL_SIZE_MB = getattr(settings, "REINSPECTION_MAX_TOTAL_SIZE_MB", 50)
REINSPECTION_MAX_IMAGE_SIZE_BYTES = REINSPECTION_MAX_IMAGE_SIZE_MB * 1024 * 1024
REINSPECTION_MAX_TOTAL_SIZE_BYTES = REINSPECTION_MAX_TOTAL_SIZE_MB * 1024 * 1024


@login_required
@require_POST
def send_paid_jobcard_approval(request, pk):
    job = get_object_or_404(JobCard.objects.select_related("claim__vehicle__customer", "vehicle"), pk=pk)
    if job.claim_id and job.claim and int(job.claim.claim_stage or 0) < 6:
        current_stage = job.claim.get_claim_stage_display()
        messages.error(
            request,
            f"Customer approval link cannot be sent before Insurance Approval. Current claim stage: {current_stage}.",
        )
        return redirect("jobcard_edit", pk=job.id)
    if not JobCardVehicleConditionPhoto.objects.filter(job=job).exists():
        messages.error(request, "Upload at least one Inspection photo before sending the customer approval link.")
        return redirect("jobcard_edit", pk=job.id)
    vehicle = job.claim.vehicle if job.claim_id and job.claim and job.claim.vehicle_id else job.vehicle
    customer = vehicle.customer if vehicle and vehicle.customer_id else None
    mobile = (customer.mobile_no or customer.whatsapp_no) if customer else ""
    if not customer or not mobile:
        messages.error(request, "Customer WhatsApp/mobile number is required before sending approval.")
        return redirect("jobcard_edit", pk=job.id)
    evidence = CustomerApprovalEvidence.objects.create(
        jobcard=job, communication_type="WhatsApp", customer_name=customer.name,
        mobile_no=mobile, remarks="Paid Job Card estimate approval requested.",
    )
    token = TimestampSigner().sign(f"{job.id}:{evidence.id}")
    approval_url = request.build_absolute_uri(reverse("customer_jobcard_approval", args=[token]))
    result = send_whatsapp_template_message(
        mobile,
        getattr(settings, "WHATSAPP_PAID_APPROVAL_TEMPLATE", "jobcard_estimate_approval"),
        parameters=[customer.name or "Customer", job.job_no, approval_url],
    )
    evidence.message_reference = approval_url
    evidence.save(update_fields=["message_reference"])
    # Keep the generated URL available on the redirected edit page so the
    # advisor can copy or print it even when Meta/WhatsApp delivery is pending.
    request.session[f"paid_approval_link_{job.id}"] = approval_url
    if result.get("success"):
        messages.success(request, "Estimate approval link sent to the customer on WhatsApp.")
    else:
        error_text = result.get("response", "Unknown error")
        if '"code":190' in error_text or "Authentication Error" in error_text:
            messages.error(request, "Approval link created, but WhatsApp authentication failed. Replace the Meta WhatsApp access token in .env, restart Django, and resend.")
        else:
            messages.warning(request, f"Approval link created, but WhatsApp sending failed: {error_text}")
        messages.info(request, f"Customer approval link: {approval_url}")
    return redirect("jobcard_edit", pk=job.id)


@login_required
@require_POST
def update_customer_consent_line_decision(request, pk):
    job = get_object_or_404(JobCard, pk=pk)
    if job.claim_id:
        messages.error(request, "Claim job cards must be updated from Claim Entry Assessment.")
        return redirect("jobcard_edit", pk=job.id)
    line_type = request.POST.get("line_type")
    try:
        line_id = int(request.POST.get("line_id"))
    except (TypeError, ValueError):
        messages.error(request, "Invalid consent line.")
        return redirect("jobcard_edit", pk=job.id)
    status = request.POST.get("status")
    if status not in {"New", "Repair", "Reject", "KO", "Approved", "Rejected", "Pending"} or line_type not in {"part", "labour"}:
        messages.error(request, "Invalid consent decision.")
        return redirect("jobcard_edit", pk=job.id)
    if line_type == "part":
        line = get_object_or_404(JobCardPart, pk=line_id, job=job)
        JobCardAssessmentPart.objects.update_or_create(
            job=job, part=line,
            defaults={"decision": status if status in {"New", "Repair", "Reject", "KO"} else "New" if status == "Approved" else "Reject" if status == "Rejected" else "KO",
                      "approval_date": timezone.now(), "updated_by": request.user},
        )
    else:
        line = get_object_or_404(JobCardLabour, pk=line_id, job=job)
        JobCardAssessmentLabour.objects.update_or_create(
            job=job, labour=line,
            defaults={"decision": status if status in {"Approved", "Reject"} else "Approved" if status == "Pending" else "Reject",
                      "approval_date": timezone.now(), "updated_by": request.user},
        )
    decision_scope = "customer approval" if not job.claim_id else "insurance/customer approval"
    messages.success(request, f"{line_type.title()} {decision_scope} decision updated to {status}.")
    return redirect("jobcard_edit", pk=job.id)


@login_required
@require_POST
def upload_customer_approval_evidence(request, pk):
    job = get_object_or_404(JobCard, pk=pk)
    evidence = (
        CustomerApprovalEvidence.objects.filter(jobcard=job, id=request.POST.get("approval_id"))
        .first()
        if request.POST.get("approval_id") else
        CustomerApprovalEvidence.objects.filter(jobcard=job).order_by("-created_at").first()
    )
    upload = request.FILES.get("approval_evidence_file")
    if not evidence:
        messages.error(request, "Send an approval link before uploading evidence.")
    elif not upload:
        messages.error(request, "Choose a WhatsApp, email, SMS, or other evidence file first.")
    elif upload.size > 10 * 1024 * 1024:
        messages.error(request, "Evidence file must be 10 MB or smaller.")
    else:
        content_type = (upload.content_type or "").lower()
        allowed = content_type.startswith("image/") or content_type in {
            "application/pdf", "text/plain", "message/rfc822"
        }
        if not allowed:
            messages.error(request, "Upload an image, PDF, email file, or text evidence document.")
        else:
            CustomerApprovalAttachment.objects.create(
                approval=evidence,
                evidence_type=request.POST.get("evidence_type") or "Other",
                file=upload,
                caption=(request.POST.get("evidence_caption") or "").strip(),
                uploaded_by=request.user,
            )
            messages.success(request, "Approval evidence uploaded successfully.")
    return redirect("jobcard_edit", pk=job.id)


@require_POST
def save_customer_approval_annotations(request, pk):
    job = get_object_or_404(JobCard, pk=pk)
    if not request.user.is_authenticated:
        token = request.headers.get("X-Approval-Token") or request.POST.get("approval_token")
        try:
            raw_job_id, evidence_id = TimestampSigner().unsign(token, max_age=60 * 60 * 24 * 30).split(":", 1)
            if str(raw_job_id) != str(job.id) or not CustomerApprovalEvidence.objects.filter(id=evidence_id, jobcard=job).exists():
                raise ValueError
        except (BadSignature, SignatureExpired, ValueError, AttributeError):
            return JsonResponse({"ok": False, "error": "Approval link is invalid or expired."}, status=403)
    try:
        payload = json.loads(request.body.decode("utf-8"))
        photo_id = int(payload.get("photo_id"))
        annotations = payload.get("annotations", [])
    except (ValueError, TypeError, json.JSONDecodeError):
        return JsonResponse({"ok": False, "error": "Invalid annotation payload."}, status=400)
    photo = get_object_or_404(JobCardVehicleConditionPhoto, pk=photo_id, job=job)
    if not isinstance(annotations, list) or len(annotations) > 100:
        return JsonResponse({"ok": False, "error": "Invalid annotation list."}, status=400)
    allowed_types = {"circle", "arrow", "rectangle", "text", "freehand"}
    clean = []
    for item in annotations:
        if not isinstance(item, dict) or item.get("type") not in allowed_types:
            continue
        clean.append({key: item[key] for key in ("type", "x", "y", "x2", "y2", "text", "color", "points") if key in item})
    record, _ = CustomerApprovalPhotoAnnotation.objects.update_or_create(
        jobcard=job, photo=photo,
        defaults={"annotations": clean, "updated_by": request.user if request.user.is_authenticated else None},
    )
    return JsonResponse({"ok": True, "id": record.id, "count": len(clean)})

@login_required
@login_required
def react_menu_api(request):

    menu_tree = build_react_menu_tree(
        request.user
    )


    return JsonResponse({

        "success": True,

        "menu": menu_tree,

    })
@login_required
@require_POST
def review_damage_ai_suggestion(request, pk, suggestion_id):
    suggestion = get_object_or_404(JobCardDamageAISuggestion, pk=suggestion_id, jobcard_id=pk)
    decision = request.POST.get("decision")
    if decision not in {"Accepted", "Rejected", "Pending"}:
        return JsonResponse({"ok": False, "error": "Invalid review decision."}, status=400)
    suggestion.status = decision
    suggestion.reviewed_by = request.user if decision != "Pending" else None
    suggestion.reviewed_at = timezone.now() if decision != "Pending" else None
    suggestion.save(update_fields=["status", "reviewed_by", "reviewed_at", "updated_at"])
    if decision == "Accepted":
        annotation_record, _ = CustomerApprovalPhotoAnnotation.objects.get_or_create(
            jobcard=suggestion.jobcard,
            photo=suggestion.photo,
            defaults={"annotations": [], "updated_by": request.user},
        )
        annotations = [item for item in (annotation_record.annotations or [])
                       if item.get("ai_suggestion_id") != suggestion.id]
        annotations.append({
            "type": "rectangle",
            "x": float(suggestion.x),
            "y": float(suggestion.y),
            "x2": float(suggestion.x + suggestion.width),
            "y2": float(suggestion.y + suggestion.height),
            "text": f"{suggestion.get_category_display()} {suggestion.confidence}%",
            "color": "#f97316",
            "ai_suggestion_id": suggestion.id,
        })
        annotation_record.annotations = annotations
        annotation_record.updated_by = request.user
        annotation_record.save(update_fields=["annotations", "updated_by", "updated_at"])
    return JsonResponse({"ok": True, "status": suggestion.status})


@login_required
@require_POST
def run_damage_ai_analysis(request, pk, photo_id):
    job = get_object_or_404(JobCard, pk=pk)
    photo = get_object_or_404(JobCardVehicleConditionPhoto, pk=photo_id, job=job)
    try:
        provider = get_damage_ai_provider()
        suggestions = provider.analyze(photo.image.file)
    except DamageAIConfigurationError as exc:
        return JsonResponse({"ok": False, "error": str(exc)}, status=503)
    except (http_requests.RequestException, ValueError, json.JSONDecodeError) as exc:
        return JsonResponse({"ok": False, "error": f"AI analysis failed: {exc}"}, status=502)
    JobCardDamageAISuggestion.objects.filter(jobcard=job, photo=photo, status="Pending").delete()
    created = []
    allowed = {choice[0] for choice in JobCardDamageAISuggestion.CATEGORY_CHOICES}
    for item in suggestions[:20]:
        if not isinstance(item, dict) or item.get("category") not in allowed:
            continue
        def number(key):
            try:
                return max(0, min(100, float(item.get(key) or 0)))
            except (TypeError, ValueError):
                return 0
        created.append(JobCardDamageAISuggestion.objects.create(
            jobcard=job, photo=photo, category=item["category"], confidence=number("confidence"),
            x=number("x"), y=number("y"), width=number("width"), height=number("height"),
            note=str(item.get("note") or "")[:255], provider=provider.name,
        ))
    return JsonResponse({"ok": True, "count": len(created), "provider": provider.name})


def customer_jobcard_tracking(request, token):
    try:
        raw_job_id = TimestampSigner().unsign(token, max_age=60 * 60 * 24 * 365)
    except (BadSignature, SignatureExpired, ValueError):
        return HttpResponse("This tracking link is invalid or expired.", status=400)
    job = get_object_or_404(
        JobCard.objects.select_related(
            "claim", "claim__vehicle__customer", "vehicle", "vehicle__customer", "branch", "advisor"
        ),
        pk=raw_job_id,
    )
    vehicle = job.vehicle or (job.claim.vehicle if job.claim_id and job.claim and job.claim.vehicle_id else None)
    inventory = JobCardInventory.objects.filter(job=job).first()
    tyre_inventory = list(job.tyres.all().order_by("position"))
    inspection_photos = JobCardVehicleConditionPhoto.objects.filter(job=job).exclude(image="").order_by("id")
    inspection_photo_data = [
        {"url": photo.image.url, "caption": photo.caption}
        for photo in inspection_photos
    ]
    # Keep the customer-facing tracker focused on the job's workshop work,
    # regardless of whether the job is insurance or cashless/paid.
    has_inspection = bool(inspection_photos or inventory)
    gate_in_done = bool(getattr(job, "gate_in_datetime", None))
    repair_done = job.repair_status in {"Completed", "Closed"}
    delivery_done = bool(getattr(job, "actual_delivery", None) or getattr(job, "ready_for_delivery", False))
    delivery_datetime = job.expected_delivery_datetime or job.estimated_delivery
    work_progress_qs = WorkProgress.objects.filter(
        allocation__job=job
    ).select_related("employee").prefetch_related("photos").order_by("id")
    work_progress_by_stage = {row.stage: row for row in work_progress_qs}
    workshop_stage_choices = dict(WorkProgress.STAGES)
    workshop_progress_rows = []
    for stage_key, stage_label in WorkProgress.STAGES:
        row = work_progress_by_stage.get(stage_key)
        if row and row.finish_time:
            status = "Completed"
        elif row and row.start_time:
            status = "In Progress"
        elif stage_key == "Dismantling" and gate_in_done and not repair_done:
            status = "In Progress"
        else:
            status = "Pending"
        workshop_progress_rows.append({
            "label": stage_label,
            "status": status,
            "start_time": row.start_time if row else None,
            "finish_time": row.finish_time if row else None,
            "start_timestamp": row.start_time.isoformat() if row and row.start_time else "",
            "finish_timestamp": row.finish_time.isoformat() if row and row.finish_time else "",
            "employee": str(row.employee) if row and row.employee else "",
            "remarks": row.remarks if row else "",
            "photos": list(row.photos.all()) if row else [],
        })
    job_progress_rows = [
        {"label": "Vehicle received", "status": "Completed" if gate_in_done else "In Progress"},
        {"label": "Inspection", "status": "Completed" if has_inspection else ("In Progress" if gate_in_done else "Pending")},
        # An open job that has been received is already in the repair queue;
        # do not leave Repair work marked Pending just because inspection
        # photos/inventory have not been uploaded yet.
        {"label": "Repair work", "status": "Completed" if repair_done else ("In Progress" if gate_in_done else "Pending")},
        {"label": "Quality check", "status": "Completed" if job.qc_done else ("In Progress" if repair_done else "Pending")},
        {"label": "Delivery", "status": "Completed" if delivery_done else ("In Progress" if getattr(job, "ready_for_delivery", False) else "Pending")},
    ]
    repair_progress_rows = job_progress_rows
    job_progress_rows = workshop_progress_rows
    claim_progress = []
    if job.claim_id and job.claim:
        current_stage = int(job.claim.claim_stage or 1)
        claim_progress = [
            {"number": number, "label": label, "done": number < current_stage, "current": number == current_stage}
            for number, label in Claim.CLAIM_STAGES
        ]
    return render(request, "jobcard/customerTracking.html", {
        "job": job,
        "vehicle": vehicle,
        "customer": vehicle.customer if vehicle and vehicle.customer_id else None,
        "claim": job.claim,
        "company": CompanySetup.objects.first(),
        "inventory": inventory,
        "inspection_photos": inspection_photos,
        "inspection_photo_data": inspection_photo_data,
        "parts_total": job.parts_total or 0,
        "labour_total": job.labour_total or 0,
        "grand_total": job.grand_total or 0,
        "delivery_datetime": delivery_datetime,
        "job_progress_rows": job_progress_rows,
        "workshop_progress_rows": workshop_progress_rows,
        "claim_progress": claim_progress,
        "tracking_react_data": {
            "company": (CompanySetup.objects.first().company_name if CompanySetup.objects.first() else "MasterApp Bodyshop"),
            "job_no": job.job_no,
            "status": job.get_repair_status_display(),
            "vehicle": vehicle.registration_no if vehicle else "-",
            "model": str(vehicle.model) if vehicle and vehicle.model else "-",
            "customer": (vehicle.customer.name if vehicle and vehicle.customer_id else "Customer"),
            "advisor": {
                "name": job.advisor.name if job.advisor_id else "Service Advisor",
                "mobile": job.advisor.mobile_no if job.advisor_id else "",
                "email": job.advisor.email if job.advisor_id else "",
                "photo": job.advisor.profile_photo.url if job.advisor_id and job.advisor.profile_photo else "",
            },
            "claim_no": job.claim.claim_no if job.claim_id and job.claim else "Service Job",
            "claim_details": {
                "intimation_date": job.claim.intimation_date.isoformat() if job.claim_id and job.claim and job.claim.intimation_date else "",
                "survey_date": job.claim.survey_date.isoformat() if job.claim_id and job.claim and job.claim.survey_date else "",
                "surveyor": str(job.claim.surveyor) if job.claim_id and job.claim and job.claim.surveyor_id else "Not assigned",
                "survey_status": job.claim.survey_status if job.claim_id and job.claim else "Pending",
                "claim_type": job.claim.get_claim_type_display() if job.claim_id and job.claim else "Service Job",
                "insurance_company": str(job.claim.insurance_company) if job.claim_id and job.claim and job.claim.insurance_company_id else "Not provided",
                "policy_no": job.claim.policy_no if job.claim_id and job.claim else "Not provided",
                "ic_claim_no": job.claim.ic_claim_no if job.claim_id and job.claim else "Not provided",
                "approval_date": job.claim.insurance_approval_date.isoformat() if job.claim_id and job.claim and job.claim.insurance_approval_date else "",
                "approval_amount": float(job.claim.approved_amount or 0) if job.claim_id and job.claim else 0,
                "approval_note": job.claim.insurance_note if job.claim_id and job.claim else "",
            },
            "km": str(job.km or "-"),
            "fuel": float(inventory.fuel_percent or 0) if inventory else 0,
            "cng": float(inventory.cng_percent or 0) if inventory else 0,
            "inventory": {
                "mud_flap_count": getattr(inventory, "mud_flap_count", 0) if inventory else 0,
                "floor_mat_count": getattr(inventory, "floor_mat_count", 0) if inventory else 0,
                "lh_mirror": bool(getattr(inventory, "lh_mirror", False)) if inventory else False,
                "rh_mirror": bool(getattr(inventory, "rh_mirror", False)) if inventory else False,
                "center_mirror": bool(getattr(inventory, "center_mirror", False)) if inventory else False,
                "frt_wiper": bool(getattr(inventory, "frt_wiper", False)) if inventory else False,
                "rr_wiper": bool(getattr(inventory, "rr_wiper", False)) if inventory else False,
                "accessories": bool(getattr(inventory, "accessories", False)) if inventory else False,
                "spare_wheel": bool(getattr(inventory, "spare_wheel", False)) if inventory else False,
                "jack": bool(getattr(inventory, "jack", False)) if inventory else False,
                "tool_kit": bool(getattr(inventory, "tool_kit", False)) if inventory else False,
                "stereo": bool(getattr(inventory, "stereo", False)) if inventory else False,
                "battery": bool(getattr(inventory, "battery", False)) if inventory else False,
                "number_plate": bool(getattr(inventory, "number_plate", False)) if inventory else False,
            },
            "tyre_inventory": [{
                "position": t.get_position_display(),
                "make": t.make or "-",
                "size": t.size or "-",
                "depth": str(t.depth) if t.depth is not None else "-",
                "wheel_cap": t.get_wheel_cap_display() if t.wheel_cap else "-",
            } for t in tyre_inventory],
            "parts_total": float(job.parts_total or 0),
            "labour_total": float(job.labour_total or 0),
            "grand_total": float(job.grand_total or 0),
            "delivery": delivery_datetime.isoformat() if delivery_datetime else "",
            "photos": inspection_photo_data,
            "progress": [{"label": row["label"], "status": row["status"]} for row in job_progress_rows],
            "repair_progress": [{
                "label": row["label"],
                "status": row["status"],
                "photos": [{"url": photo.image.url, "caption": row["label"]} for photo in row.get("photos", [])],
            } for row in workshop_progress_rows],
            "claim_progress": [{"number": row["number"], "label": row["label"], "done": row["done"], "current": row["current"]} for row in claim_progress],
        },
    })


def customer_jobcard_approval(request, token):
    try:
        raw = TimestampSigner().unsign(token, max_age=60 * 60 * 24 * 30)
        job_id, evidence_id = raw.split(":", 1)
    except (BadSignature, SignatureExpired, ValueError):
        return HttpResponse("This approval link is invalid or expired.", status=400)
    evidence = get_object_or_404(CustomerApprovalEvidence.objects.select_related("jobcard", "jobcard__claim__vehicle__customer"), id=evidence_id, jobcard_id=job_id)
    if request.method == "POST" and evidence.status == "Pending":
        decision = request.POST.get("decision")
        if decision in {"Approved", "Rejected", "Need Clarifications"}:
            evidence.status = decision
            evidence.approval_date = timezone.now()
            evidence.remarks = (request.POST.get("remarks") or "").strip()
            evidence.save(update_fields=["status", "approval_date", "remarks"])
            if evidence.jobcard.claim_id:
                ClaimTimeline.objects.create(
                    claim=evidence.jobcard.claim,
                    stage=f"Customer Consent {decision}",
                    remarks=(
                        f"Customer responded {decision.lower()} to the approval link."
                        + (f" Remarks: {evidence.remarks}" if evidence.remarks else "")
                    ),
                )
            notify_jobcard_advisor(
                evidence.jobcard,
                f"Customer {decision}",
                f"Customer responded {decision.lower()} to the job card approval link for {evidence.jobcard.job_no}."
                + (f" Remarks: {evidence.remarks}" if evidence.remarks else ""),
            )
    job = evidence.jobcard
    vehicle = (job.claim.vehicle if job.claim_id and job.claim and job.claim.vehicle_id else job.vehicle)
    inventory = JobCardInventory.objects.filter(job=job).first()
    photos = job.vehicle_condition_photos.order_by("id")
    approval_attachments = evidence.attachments.order_by("-uploaded_at", "-id")
    approval_photo_annotations = {
        row.photo_id: row.annotations
        for row in CustomerApprovalPhotoAnnotation.objects.filter(jobcard=job)
    }
    approval_photo_rows = [
        {"id": photo.id, "url": photo.image.url, "caption": photo.caption,
         "annotations": approval_photo_annotations.get(photo.id, [])}
        for photo in photos if photo.image
    ]
    customer_damage_marks = list(inventory.damage_marks if inventory else [])
    if approval_photo_rows:
        for annotation in approval_photo_rows[0]["annotations"]:
            if annotation.get("x") is not None and annotation.get("y") is not None:
                customer_damage_marks.append({
                    "x": annotation.get("x"), "y": annotation.get("y"),
                    "type": annotation.get("type", "annotation"),
                    "label": annotation.get("text", "Saved photo annotation"),
                })
    return render(request, "jobcard/customerApproval.html", {
        "evidence": evidence,
        "job": job,
        "vehicle": vehicle,
        "inventory": inventory,
        "damage_marks": customer_damage_marks,
        "customer_damage_marks_json": json.dumps(customer_damage_marks),
        "inspection_photos": photos,
        "rotation_photos": photos,
        "approval_attachments": approval_attachments,
        "approval_photo_annotations": approval_photo_annotations,
        "approval_photo_rows": approval_photo_rows,
        "approval_photo_json": json.dumps(approval_photo_rows),
        "approval_token": token,
        "company": CompanySetup.objects.first(),
    })

VEHICLE_CONDITION_PHOTO_CAPTIONS = [
    "Front View",
    "Front Right Corner View",
    "Full Right View",
    "Rear Right Corner View",
    "Rear View",
    "Rear Left Corner View",
    "Full Left View",
    "Front Left Corner View",
    "Engine Compartment",
    "Windshield Glass",
    "Instrument Cluster Photo",
    "Full Dashboard Photo",
    "Interior View Photo",
    "Chassis No View",
    "Rear Glass View",
    "Dicky View",
    "Jack & Tools View",
    "Stepney View",
    "Other View 1",
    "Other View 2",
]

CLAIM_DOCUMENT_TYPES = [
    "RC Book",
    "Driving License",
    "Insurance Policy",
    "Aadhar Card",
    "PAN Card",
    "Claim Form",
    "Other 1",
    "Other 2",
]


def is_parts_manager(employee):
    if not employee:
        return False
    role_text = " ".join([
        employee.employee_type or "",
        employee.designation or "",
        employee.department or "",
    ]).upper()
    return any(keyword in role_text for keyword in [
        "PARTS MANAGER",
        "PART MANAGER",
        "SPARES MANAGER",
        "PARTS DEPARTMENT",
        "SPARE PARTS",
    ])


def get_reinspection_photo_storage_size(job):
    total_size = 0

    for photo in job.reinspection_photos.all():
        if not photo.image:
            continue

        try:
            total_size += photo.image.size
        except (OSError, ValueError):
            continue

    return total_size


def vehicle_condition_photo_input_name(index):
    return f"vehicle_condition_photo_{index}"


def get_vehicle_condition_photo_slots(job):
    existing = {}

    if job:
        existing = {
            photo.caption: photo
            for photo in job.vehicle_condition_photos.all()
        }

    slots = [
        {
            "index": index,
            "caption": caption,
            "input_name": vehicle_condition_photo_input_name(index),
            "photo": existing.get(caption),
        }
        for index, caption in enumerate(
            VEHICLE_CONDITION_PHOTO_CAPTIONS,
            start=1
        )
    ]

    rotation_photos = [
        slot
        for slot in slots
        if slot["index"] <= 8 and slot["photo"]
    ]

    return {
        "slots": slots,
        "rotation_photos": rotation_photos,
    }

def save_vehicle_condition_photos(request, job):
    for index, caption in enumerate(VEHICLE_CONDITION_PHOTO_CAPTIONS, start=1):
        image = request.FILES.get(vehicle_condition_photo_input_name(index))
        if not image:
            continue

        photo = JobCardVehicleConditionPhoto.objects.filter(
            job=job,
            caption=caption,
        ).first()

        if photo and photo.image:
            photo.image.delete(save=False)

        if photo:
            photo.image = image
            photo.save()
        else:
            JobCardVehicleConditionPhoto.objects.create(
                job=job,
                caption=caption,
                image=image,
            )


def save_signature_data(job, field_name, data_url):
    data_url = (data_url or "").strip()
    field = getattr(job, field_name)

    if data_url == "__clear__":
        if field:
            field.delete(save=False)
        setattr(job, field_name, None)
        return True

    if not data_url.startswith("data:image/png;base64,"):
        return False

    try:
        image_data = base64.b64decode(data_url.split(",", 1)[1])
    except (ValueError, IndexError):
        return False

    if field:
        field.delete(save=False)

    filename = f"{job.job_no}_{field_name}.png".replace("/", "_")
    setattr(job, field_name, ContentFile(image_data, name=filename))
    return True


def save_jobcard_signatures(request, job):
    changed = False
    changed = save_signature_data(
        job,
        "advisor_signature",
        request.POST.get("advisor_signature_data")
    ) or changed
    changed = save_signature_data(
        job,
        "customer_signature",
        request.POST.get("customer_signature_data")
    ) or changed

    if changed:
        job.save(update_fields=["advisor_signature", "customer_signature"])


def claim_document_input_name(index):
    return f"claim_document_{index}"


def get_claim_document_slots(claim):
    existing = {}
    if claim:
        for document in ClaimDocument.objects.filter(claim=claim).order_by("uploaded_at", "id"):
            existing.setdefault(document.document_type, []).append(document)

    return [
        {
            "index": index,
            "document_type": document_type,
            "input_name": claim_document_input_name(index),
            "documents": existing.get(document_type, []),
            "max_files": 8,
        }
        for index, document_type in enumerate(CLAIM_DOCUMENT_TYPES, start=1)
    ]


def save_claim_documents(request, claim):
    if not claim:
        return

    for index, document_type in enumerate(CLAIM_DOCUMENT_TYPES, start=1):
        input_name = claim_document_input_name(index)
        uploaded_files = request.FILES.getlist(input_name)
        if not uploaded_files:
            continue

        existing_documents = list(ClaimDocument.objects.filter(
            claim=claim, document_type=document_type
        ).order_by("uploaded_at", "id"))

        available_slots = max(0, 8 - len(existing_documents))
        if available_slots:
            for uploaded_file in uploaded_files[:available_slots]:
                ClaimDocument.objects.create(
                    claim=claim, document_type=document_type, file=uploaded_file
                )
            continue

        uploaded_file = uploaded_files[0]
        document = existing_documents[0] if existing_documents else None

        if document and document.file:
            document.file.delete(save=False)

        if document:
            document.file = uploaded_file
            document.save()
        else:
            ClaimDocument.objects.create(
                claim=claim,
                document_type=document_type,
                file=uploaded_file,
            )


def progress_photo_input_name(stage):
    safe_stage = "".join(
        char if char.isalnum() or char in ["-", "_"] else "_"
        for char in str(stage or "")
    )
    return f"progress_photo_{safe_stage}"


def is_repair_resource(employee):
    if not employee:
        return False

    role_text = f"{employee.employee_type or ''} {employee.designation or ''}".upper()
    return any(
        keyword in role_text
        for keyword in ["TECHNICIAN", "DENTER", "PAINTER"]
    )


def is_floor_supervisor(employee):
    if not employee:
        return False

    role_text = f"{employee.employee_type or ''} {employee.designation or ''}".upper()
    return any(
        keyword in role_text
        for keyword in ["FLOOR SUPERVISOR", "FLOOR INCHARGE", "FLOOR IN-CHARGE"]
    )


def workflow_date_value(value):
    if not value:
        return None

    if isinstance(value, datetime):
        if timezone.is_aware(value):
            return timezone.localtime(value)
        return timezone.make_aware(value)

    if hasattr(value, "date"):
        return timezone.make_aware(datetime.combine(value, datetime_time.min))

    return timezone.make_aware(datetime.combine(value, datetime_time.min))


def parse_workflow_datetime(value):
    value = (value or "").strip()
    if not value:
        return None

    parsed = parse_datetime(value)
    if parsed:
        return parsed if timezone.is_aware(parsed) else timezone.make_aware(parsed)

    parsed_date = parse_date(value)
    if parsed_date:
        return timezone.make_aware(datetime.combine(parsed_date, datetime_time.min))

    return None


def datetime_local_value(value):
    if not value:
        return ""

    if isinstance(value, datetime):
        if timezone.is_aware(value):
            value = timezone.localtime(value)
        return value.strftime("%Y-%m-%dT%H:%M")

    return datetime.combine(value, datetime_time.min).strftime("%Y-%m-%dT%H:%M")


def validate_no_future_workflow_dates(points):
    now = timezone.now()

    for label, value in points:
        current_value = control_board_datetime_value(value)
        if current_value and current_value > now:
            return f"{label} cannot be future date/time."

    return ""


def validate_workflow_dates(points, validate_labels=None):
    validate_all = validate_labels is None
    validate_labels = set(validate_labels or [])
    previous_label = None
    previous_date = None

    for label, value in points:
        current_date = workflow_date_value(value)
        if not current_date:
            continue

        should_validate = validate_all or label in validate_labels

        if should_validate and previous_date and current_date < previous_date:
            return (
                f"{label} cannot be before {previous_label}. "
                f"{previous_label}: {previous_date.strftime('%d-%m-%Y %H:%M')}, "
                f"{label}: {current_date.strftime('%d-%m-%Y %H:%M')}"
            )

        previous_label = label
        previous_date = current_date

    return ""


def workflow_date_changed(old_value, new_value):
    old_date = workflow_date_value(old_value)
    new_date = workflow_date_value(new_value)
    return bool(new_date and old_date != new_date)


def validate_claim_job_workflow_dates(
        claim,
        job=None,
        allocation=None,
        claim_created_date=None,
        job_created_date=None,
        validate_labels=None
):
    claim_created_value = claim_created_date or (claim.created_at if claim else None)
    job_created_value = job_created_date or (job.job_date if job else None) or (job.created_at if job else None)
    first_progress = None
    work_completed = None

    if allocation:
        first_progress = (
            allocation.progress
            .filter(start_time__isnull=False)
            .order_by("start_time")
            .first()
        )
        work_completed = (
            allocation.progress
            .filter(finish_time__isnull=False)
            .order_by("-finish_time")
            .first()
        )

    return validate_workflow_dates([
        ("Gate In Date", job.gate_in_datetime if job else None),
        ("Claim Created Date", claim_created_value),
        ("Jobcard Created Date", job_created_value),
        ("Claim Intimation Date", claim.intimation_date if claim else None),
        ("Survey Date", claim.survey_date if claim else None),
        ("Insurance Approval Date", claim.insurance_approval_date if claim else None),
        ("Work Allocation Date", allocation.allotment_date if allocation else None),
        ("Repair Start Date", first_progress.start_time if first_progress else None),
        ("Work Completed Date", work_completed.finish_time if work_completed else None),
        ("Re-Inspection Date", job.reinspection_date if job else None),
        ("Liability Received Date", claim.liability_received_at if claim else None),
        ("Invoice Date", claim.invoice_datetime if claim else None),
        ("Delivery Date", claim.delivery_datetime if claim else None),
    ], validate_labels=validate_labels)


def my_work_base_queryset(employee, from_date=None, to_date=None):
    progress = (
        WorkProgress.objects
        .select_related(
            "allocation",
            "allocation__job",
            "allocation__job__claim",
            "allocation__job__claim__vehicle",
            "allocation__job__claim__vehicle__customer",
            "allocation__job__claim__vehicle__model",
            "employee",
        )
        .prefetch_related("photos")
        .filter(employee=employee)
    )

    if from_date:
        progress = progress.filter(allocation__job__created_at__date__gte=from_date)

    if to_date:
        progress = progress.filter(allocation__job__created_at__date__lte=to_date)

    return progress


def apply_my_work_status_filter(progress, status_filter):
    if status_filter == "wip":
        return progress.filter(start_time__isnull=False, finish_time__isnull=True)

    if status_filter == "completed":
        return progress.filter(finish_time__isnull=False)

    return progress.filter(start_time__isnull=True)


def my_work_row_payload(progress):
    job = progress.allocation.job if progress.allocation_id else None
    claim = job.claim if job and job.claim_id else None
    vehicle = claim.vehicle if claim and claim.vehicle_id else None

    return {
        "id": progress.id,
        "stage": progress.stage,
        "stage_label": progress.get_stage_display(),
        "start_time": progress.start_time,
        "finish_time": progress.finish_time,
        "remarks": progress.remarks or "",
        "photo_count": progress.photos.count(),
        "job": job,
        "claim": claim,
        "vehicle": vehicle,
        "additional_approval_required": bool(
            job and job.additional_approval_required
        ),
        "second_approval_status": (
            job.second_approval_status
            if job
            else ""
        ),
        "additional_approval_reason": (
            job.additional_approval_reason
            if job
            else ""
        ),
    }


def start_work_progress(progress):
    if not progress.start_time:
        progress.start_time = timezone.now()
        progress.save(update_fields=["start_time"])
        return True
    return False


def finish_work_progress(progress):
    changed = False
    if not progress.start_time:
        progress.start_time = timezone.now()
        changed = True
    if not progress.finish_time:
        progress.finish_time = timezone.now()
        changed = True
    progress.save(update_fields=["start_time", "finish_time"])
    return changed


def save_work_progress_uploaded_photos(request, progress):
    for image in request.FILES.getlist("progress_photos"):
        WorkProgressPhoto.objects.create(
            progress=progress,
            image=image
        )


def is_admin_or_manager_user(user, employee=None):
    if not user or not user.is_authenticated:
        return False

    role_text = f"{employee.employee_type if employee else ''} {employee.designation if employee else ''}".upper()
    return (
        user.is_superuser
        or "MANAGER" in role_text
        or "ADMIN" in role_text
        or user.groups.filter(name__iexact="Manager").exists()
    )


def is_admin_user(user, employee=None):
    if not user or not user.is_authenticated:
        return False

    role_text = f"{employee.employee_type if employee else ''} {employee.designation if employee else ''}".upper()
    return (
        user.is_superuser
        or user.groups.filter(name__iexact="Admin").exists()
        or "ADMIN" in role_text
    )


def dashboard_stage_rows(stage_counts):
    stage_labels = dict(Claim.CLAIM_STAGES)
    stage_order = [value for value, _label in Claim.CLAIM_STAGES]

    rows = []
    for item in stage_counts:
        current_stage = item["claim_stage"]
        try:
            stage_index = stage_order.index(current_stage)
        except ValueError:
            stage_index = -1

        next_stage = (
            stage_order[stage_index + 1]
            if stage_index >= 0 and stage_index + 1 < len(stage_order)
            else current_stage
        )

        rows.append({
            "claim_stage": current_stage,
            "current_stage_label": stage_labels.get(current_stage, "Unknown"),
            "pending_stage": next_stage,
            "pending_stage_label": stage_labels.get(next_stage, "Unknown"),
            "total": item["total"],
        })

    return rows


def active_session_user_ids():
    active_ids = set()
    for session in Session.objects.filter(expire_date__gt=timezone.now()):
        data = session.get_decoded()
        user_id = data.get("_auth_user_id")
        if user_id:
            try:
                active_ids.add(int(user_id))
            except (TypeError, ValueError):
                continue
    return active_ids


@login_required
def login_activity_page(request):
    logged_emp = Employee.objects.filter(user=request.user).select_related("branch").first()
    if not is_admin_user(request.user, logged_emp):
        messages.error(request, "Only admin users can view login activity.")
        return redirect("dashboard")

    today = timezone.localdate()
    from_date_value = request.GET.get("from_date") or today.isoformat()
    to_date_value = request.GET.get("to_date") or today.isoformat()
    search_text = (request.GET.get("search") or "").strip()

    from_date = parse_date(from_date_value) or today
    to_date = parse_date(to_date_value) or from_date
    if to_date < from_date:
        from_date, to_date = to_date, from_date

    activities = (
        UserLoginActivity.objects
        .select_related("user", "user__employee", "user__employee__branch")
        .filter(login_at__date__gte=from_date, login_at__date__lte=to_date)
    )

    if search_text:
        activities = activities.filter(
            Q(user__username__icontains=search_text)
            | Q(user__first_name__icontains=search_text)
            | Q(user__last_name__icontains=search_text)
            | Q(user__employee__name__icontains=search_text)
            | Q(user__employee__mobile_no__icontains=search_text)
        )

    active_user_ids = active_session_user_ids()
    activity_rows = []
    for activity in activities[:500]:
        employee = getattr(activity.user, "employee", None)
        activity_rows.append({
            "activity": activity,
            "employee": employee,
            "branch": employee.branch if employee and employee.branch_id else None,
            "is_active_now": activity.user_id in active_user_ids,
        })

    summary = {
        "total_logins": activities.count(),
        "unique_users": activities.values("user_id").distinct().count(),
        "active_now": len(active_user_ids),
    }

    return render(request, "admin/login_activity.html", {
        "activity_rows": activity_rows,
        "summary": summary,
        "from_date": from_date.isoformat(),
        "to_date": to_date.isoformat(),
        "search_text": search_text,
        "max_rows": 500,
    })


def branch_scoped_queryset_for_user(queryset, user, branch_field="branch"):
    employee = Employee.objects.filter(user=user).select_related("branch").first()

    if is_admin_user(user, employee):
        return queryset

    if employee and employee.branch_id:
        return queryset.filter(**{branch_field: employee.branch})

    return queryset.none()


def report_branch_context(request):
    employee = Employee.objects.filter(user=request.user).select_related("branch").first()
    admin_user = is_admin_user(request.user, employee)
    branches = Branch.objects.filter(is_active=True).order_by("name")
    selected_branch_id = request.GET.get("branch_id") or ""
    selected_branch = None

    if admin_user and selected_branch_id:
        selected_branch = branches.filter(pk=selected_branch_id).first()
    elif not admin_user:
        selected_branch = employee.branch if employee and employee.branch_id else None
        selected_branch_id = str(selected_branch.id) if selected_branch else ""

    return {
        "report_is_admin": admin_user,
        "report_branches": branches,
        "report_branch": selected_branch,
        "report_branch_id": selected_branch_id,
        "report_branch_name": selected_branch.name if selected_branch else "All Branches",
    }


def apply_report_branch_scope(queryset, branch_context, branch_field="branch"):
    branch = branch_context.get("report_branch")
    if branch:
        return queryset.filter(**{branch_field: branch})
    if branch_context.get("report_is_admin"):
        return queryset
    return queryset.none()


def create_user_notification(user, title, message, url=""):
    if not user:
        return

    UserNotification.objects.create(
        user=user,
        title=title,
        message=message,
        url=url or "",
    )


def notify_jobcard_advisor(job, title, message):
    if not job:
        return

    advisor_user = None
    if job.advisor_id and job.advisor and job.advisor.user_id:
        advisor_user = job.advisor.user
    elif job.claim_id and job.claim and job.claim.employee_id and job.claim.employee.user_id:
        advisor_user = job.claim.employee.user

    create_user_notification(
        advisor_user,
        title,
        message,
        f"/jobCard/{job.id}/edit/",
    )


def notify_parts_team(title, message, url=""):
    for employee in Employee.objects.select_related("user").filter(is_active=True, user__isnull=False):
        if is_parts_manager(employee):
            create_user_notification(employee.user, title, message, url)


def floor_incharge_users():
    employees = Employee.objects.filter(
        is_active=True,
        user__isnull=False,
    )
    users = []
    seen_user_ids = set()

    for employee in employees:
        role_text = f"{employee.employee_type or ''} {employee.designation or ''}".upper()

        if not any(
            keyword in role_text
            for keyword in ["FLOOR INCHARGE", "FLOOR IN-CHARGE", "FLOOR SUPERVISOR"]
        ):
            continue

        if employee.user_id in seen_user_ids:
            continue

        users.append(employee.user)
        seen_user_ids.add(employee.user_id)

    return users


def reception_users(branch=None):
    employees = Employee.objects.filter(
        is_active=True,
        user__isnull=False,
    )
    if branch:
        employees = employees.filter(Q(branch=branch) | Q(branch__isnull=True))

    users = []
    seen_user_ids = set()

    for employee in employees:
        role_text = " ".join([
            employee.employee_type or "",
            employee.designation or "",
            employee.department or "",
        ]).upper()

        if not any(keyword in role_text for keyword in ["RECEPTION", "FRONT OFFICE"]):
            continue

        if employee.user_id in seen_user_ids:
            continue

        users.append(employee.user)
        seen_user_ids.add(employee.user_id)

    return users


def notify_reception_gate_in(entry):
    if not entry or entry.service_type != "Bodyshop":
        return

    gate_time = timezone.localtime(entry.gate_in_datetime).strftime("%d/%m/%Y %H:%M")
    message = (
        f"Vehicle {entry.registration_no} entered Bodyshop gate at {gate_time}. "
        f"Current KM: {entry.current_km}. Create claim entry."
    )

    for user in reception_users(entry.branch):
        create_user_notification(
            user,
            "Gate In - Claim Creation",
            message,
            "/claim/",
        )


def notify_reception_gate_in_changed(entry, cancelled=False):
    if not entry:
        return

    if not cancelled and entry.service_type != "Bodyshop":
        return

    gate_time = timezone.localtime(entry.gate_in_datetime).strftime("%d/%m/%Y %H:%M")
    if cancelled:
        title = "Gate In Cancelled"
        message = (
            f"Gate In for vehicle {entry.registration_no} was cancelled. "
            f"Remark: {entry.cancellation_remark or '-'}"
        )
        url = "/"
    else:
        title = "Gate In Updated - Claim Creation"
        message = (
            f"Gate In updated for vehicle {entry.registration_no}. "
            f"Gate time: {gate_time}. Current KM: {entry.current_km}. "
            f"Create claim entry."
        )
        url = "/claim/"

    for user in reception_users(entry.branch):
        create_user_notification(
            user,
            title,
            message,
            url,
        )


def is_reception_employee(employee):
    if not employee:
        return False

    role_text = " ".join([
        employee.employee_type or "",
        employee.designation or "",
        employee.department or "",
    ]).upper()

    return any(keyword in role_text for keyword in ["RECEPTION", "FRONT OFFICE"])


def compact_wait_time(value):
    if not value:
        return "-"

    current = timezone.now()
    if timezone.is_naive(value):
        value = timezone.make_aware(value)

    total_minutes = max(0, int((current - value).total_seconds() // 60))
    if total_minutes < 60:
        return f"{total_minutes}m"

    hours = total_minutes // 60
    minutes = total_minutes % 60
    if hours < 24:
        return f"{hours}h {minutes}m"

    days = hours // 24
    rem_hours = hours % 24
    return f"{days}d {rem_hours}h"


def notify_floor_incharge_work_allocated(job):
    if not job:
        return

    claim = job.claim if job.claim_id else None
    vehicle = claim.vehicle if claim and claim.vehicle_id else None
    registration_no = vehicle.registration_no if vehicle else "-"
    model_name = vehicle.model.name if vehicle and vehicle.model_id else "-"
    message = (
        f"Work allocated for Jobcard {job.job_no} "
        f"({registration_no} - {model_name})"
    )

    for user in floor_incharge_users():
        create_user_notification(
            user,
            "Work Allocated",
            message,
            f"/work-allocation/{job.id}/",
        )


def notify_floor_incharge_work_allocation_pending(claim):
    if not claim:
        return

    job = JobCard.objects.filter(claim=claim).select_related(
        "claim",
        "claim__vehicle",
        "claim__vehicle__model",
    ).first()
    vehicle = claim.vehicle if claim.vehicle_id else None
    registration_no = vehicle.registration_no if vehicle else "-"
    model_name = vehicle.model.name if vehicle and vehicle.model_id else "-"
    job_no = job.job_no if job else "-"
    url = f"/work-allocation/{job.id}/" if job else f"/claim/{claim.id}/edit/"
    message = (
        f"Claim {claim.claim_no} Assign For Work Allocation. "
        f"Jobcard {job_no} ({registration_no} - {model_name})"
    )

    for user in floor_incharge_users():
        create_user_notification(
            user,
            "Work Allocation Pending",
            message,
            url,
        )

    if not job or not hasattr(job, "allocation"):
        return

    notified_user_ids = set()
    for progress in job.allocation.progress.select_related("employee__user"):
        employee = progress.employee
        if not employee or not employee.user_id or employee.user_id in notified_user_ids:
            continue
        if not is_repair_resource(employee):
            continue
        notified_user_ids.add(employee.user_id)
        message = (
            f"Insurance approval is complete. You can start "
            f"{progress.get_stage_display()} for Jobcard {job.job_no} "
            f"({registration_no})."
        )
        notification, _ = UserNotification.objects.get_or_create(
            user=employee.user,
            title="Repair Work Ready to Start",
            message=message,
            is_read=False,
            defaults={"url": "/my-work/"},
        )


def notify_progress_employee_assigned(progress):
    if not progress or not progress.employee_id or not progress.employee.user_id:
        return

    if not is_repair_resource(progress.employee):
        return

    job = progress.allocation.job if progress.allocation_id else None
    if not job:
        return

    claim = job.claim if job.claim_id else None
    vehicle = claim.vehicle if claim and claim.vehicle_id else None
    registration_no = vehicle.registration_no if vehicle else "-"
    model_name = vehicle.model.name if vehicle and vehicle.model_id else "-"
    message = (
        f"{progress.get_stage_display()} assigned for "
        f"Jobcard {job.job_no} ({registration_no} - {model_name})"
    )

    create_user_notification(
        progress.employee.user,
        "New Work Assigned",
        message,
        "/my-work/",
    )


def notify_floor_incharge_work_progress(progress, action_label):
    job = progress.allocation.job if progress and progress.allocation_id else None
    if not job:
        return

    claim = job.claim if job.claim_id else None
    vehicle = claim.vehicle if claim and claim.vehicle_id else None
    registration_no = vehicle.registration_no if vehicle else "-"
    employee_name = progress.employee.name if progress.employee_id else "-"
    message = (
        f"{employee_name} {action_label} "
        f"{progress.get_stage_display()} for Jobcard {job.job_no} "
        f"({registration_no})"
    )

    for user in floor_incharge_users():
        create_user_notification(
            user,
            "Work Progress Updated",
            message,
            f"/work-allocation/{job.id}/",
        )


def notify_work_progress_change(progress, action_label):
    job = progress.allocation.job if progress and progress.allocation_id else None
    if not job:
        return

    vehicle = job.claim.vehicle if job.claim_id and job.claim and job.claim.vehicle_id else None
    registration_no = vehicle.registration_no if vehicle else "-"
    message = (
        f"Jobcard {job.job_no} {progress.get_stage_display()} "
        f"{action_label} for {registration_no}"
    )
    notify_jobcard_advisor(
        job,
        "Repair Work Progress Updated",
        message,
    )

    if action_label in ["started", "finished"]:
        notify_floor_incharge_work_progress(progress, action_label)


def can_update_second_approval(user, employee, job):
    if is_admin_or_manager_user(user, employee):
        return True

    if not employee or not job:
        return False

    return (
        (job.advisor_id and job.advisor_id == employee.id)
        or (
            job.claim_id
            and job.claim
            and job.claim.employee_id == employee.id
        )
    )


def latest_pending_gate_entry_for_claim(claim):
    if not claim or not claim.vehicle_id:
        return None

    branch = branch_for_claim(claim)
    return latest_pending_gate_entry_for_vehicle(claim.vehicle, branch=branch)


def latest_pending_gate_entry_for_vehicle(vehicle, branch=None, strict_branch=False):
    if not vehicle:
        return None

    entries = GateInEntry.objects.filter(
        Q(vehicle=vehicle) | Q(registration_no__iexact=vehicle.registration_no),
        status="Pending",
        jobcard__isnull=True,
    )
    if branch:
        entries = entries.filter(
            Q(branch=branch) if strict_branch else Q(branch=branch) | Q(branch__isnull=True)
        )
    elif strict_branch:
        entries = entries.filter(branch__isnull=True)

    return entries.order_by("-gate_in_datetime").first()


@login_required
def vehicle_gate_in_status(request):
    vehicle_id = request.GET.get("vehicle_id") or request.GET.get("id")
    if not vehicle_id or not str(vehicle_id).isdigit():
        return JsonResponse({
            "status": "error",
            "message": "Select valid vehicle first.",
        }, status=400)

    vehicle = Vehicle.objects.filter(pk=vehicle_id).first()
    if not vehicle:
        return JsonResponse({
            "status": "error",
            "message": "Vehicle not found.",
        }, status=404)

    branch = None if is_admin_user(request.user) else branch_for_user(request.user)
    entry = latest_pending_gate_entry_for_vehicle(vehicle, branch=branch, strict_branch=True)
    if not entry:
        return JsonResponse({
            "status": "missing",
            "message": "First Gate In Entry then continue.",
        })

    return JsonResponse({
        "status": "success",
        "gate_entry_id": entry.id,
        "message": "Gate In Entry found.",
        "gate_in_display": timezone.localtime(entry.gate_in_datetime).strftime("%d/%m/%Y %I:%M %p"),
        "current_km": entry.current_km,
        "service_type": entry.service_type,
    })


@never_cache
@login_required
def gate_in_entry(request):
    if request.method == "POST":
        registration_no = (request.POST.get("registration_no") or "").strip().upper()
        current_km_raw = (request.POST.get("current_km") or "").strip()
        service_type = (request.POST.get("service_type") or "").strip()
        remarks = (request.POST.get("remarks") or "").strip()

        if not registration_no:
            messages.error(request, "Vehicle No is required.")
            return redirect("gate_in_entry")

        try:
            current_km = int(current_km_raw)
        except (TypeError, ValueError):
            current_km = 0

        if current_km <= 0:
            messages.error(request, "Enter valid Current KM.")
            return redirect("gate_in_entry")

        valid_service_types = dict(GateInEntry.SERVICE_TYPE_CHOICES)
        if service_type not in valid_service_types:
            messages.error(request, "Select valid Service Type.")
            return redirect("gate_in_entry")

        duplicate = GateInEntry.objects.filter(
            registration_no__iexact=registration_no,
            status="Pending",
        ).first()
        if duplicate:
            messages.error(
                request,
                f"{registration_no} already has a pending Gate In entry from "
                f"{timezone.localtime(duplicate.gate_in_datetime):%d/%m/%Y %H:%M}.",
            )
            return redirect("gate_in_entry")

        vehicle = Vehicle.objects.filter(registration_no__iexact=registration_no).first()
        entry = GateInEntry.objects.create(
            registration_no=registration_no,
            current_km=current_km,
            service_type=service_type,
            gate_in_datetime=timezone.now(),
            branch=branch_for_user(request.user),
            vehicle=vehicle,
            entered_by=request.user,
            remarks=remarks,
        )
        notify_reception_gate_in(entry)

        messages.success(request, f"Gate In saved for {registration_no}.")
        return redirect("gate_in_entry")

    return render(
        request,
        "gate/gate_in_entry.html",
        {
            "service_types": GateInEntry.SERVICE_TYPE_CHOICES,
        },
    )


@never_cache
@login_required
def gate_in_entry_data(request):
    status_filter = (request.GET.get("status") or "all").strip()
    service_type = (request.GET.get("service_type") or "").strip()
    entries = GateInEntry.objects.select_related(
        "vehicle",
        "vehicle__customer",
        "branch",
        "entered_by",
        "jobcard",
    ).order_by("-gate_in_datetime")

    if status_filter and status_filter.lower() != "all":
        entries = entries.filter(status=status_filter)

    if service_type:
        entries = entries.filter(service_type=service_type)

    # Closed jobs that bypassed Gate-In still need a controlled Gate-Out record.
    # Create a Converted register entry for them so security can capture the
    # physical gate pass; the display intentionally leaves Gate-In date blank.
    closed_jobs = JobCard.objects.select_related("claim", "claim__vehicle", "vehicle").filter(
        repair_status="Closed",
        claim__isnull=False,
    ).filter(Q(claim__status="Closed") | Q(claim__claim_stage=ClaimStageCode.CLOSED))
    for job in closed_jobs[:200]:
        vehicle = job.vehicle or (job.claim.vehicle if job.claim and job.claim.vehicle_id else None)
        if not vehicle:
            continue
        if not GateInEntry.objects.filter(jobcard=job).exists():
            GateInEntry.objects.create(
                registration_no=vehicle.registration_no,
                current_km=job.km or 0,
                service_type="Bodyshop",
                gate_in_datetime=job.gate_in_datetime or timezone.now(),
                branch=job.branch,
                vehicle=vehicle,
                status="Converted",
                jobcard=job,
                remarks="Created for Gate-Out; original Gate-In entry was not recorded.",
            )

    data = []
    for entry in entries[:200]:
        # Repair legacy/mobile-created records that have a Job Card for the
        # same vehicle but were not linked when the Job Card was saved.
        if entry.status == "Pending" and not entry.jobcard_id:

            matching_query = Q(claim__vehicle__registration_no__iexact=entry.registration_no)

            if entry.vehicle_id:
                matching_query |= (
                        Q(vehicle_id=entry.vehicle_id) |
                        Q(claim__vehicle_id=entry.vehicle_id)
                )

            matching_job = JobCard.objects.filter(
                matching_query
            ).order_by("-id").first()

            if matching_job:

                existing_entry = GateInEntry.objects.filter(
                    jobcard=matching_job
                ).exclude(
                    id=entry.id
                ).first()

                if existing_entry:
                    # JobCard already converted
                    entry.status = "Pending"
                    entry.save(update_fields=["status", "updated_at"])

                else:
                    entry.jobcard = matching_job
                    entry.status = "Converted"
                    entry.save(update_fields=[
                        "jobcard",
                        "status",
                        "updated_at"
                    ])
        linked_claim = entry.jobcard.claim if entry.jobcard_id and entry.jobcard.claim_id else None
        can_gate_out = bool(
            entry.status == "Converted"
            and entry.jobcard_id
            and linked_claim
            and entry.jobcard.repair_status == "Closed"
            and (linked_claim.status == "Closed" or linked_claim.claim_stage == ClaimStageCode.CLOSED)
        )
        data.append({
            "id": entry.id,
            "registration_no": entry.registration_no,
            "customer": entry.vehicle.customer.name if entry.vehicle_id and entry.vehicle.customer_id else "",
            "current_km": entry.current_km,
            "service_type": entry.service_type,
            "gate_in_datetime": "" if "original Gate-In entry" in (entry.remarks or "") else timezone.localtime(entry.gate_in_datetime).strftime("%d/%m/%Y %H:%M"),
            "branch": entry.branch.name if entry.branch_id else "",
            "entered_by": entry.entered_by.username if entry.entered_by_id else "",
            "status": entry.status,
            "job_no": entry.jobcard.job_no if entry.jobcard_id else "",
            "can_gate_out": can_gate_out,
            "remarks": entry.remarks,
            "cancellation_remark": entry.cancellation_remark,
        })

    return JsonResponse({"data": data})


@login_required
def gate_in_entry_export(request):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    today = timezone.localdate()
    default_start = today.replace(day=1)
    try:
        start_date = datetime.strptime(request.GET.get("date_from") or "", "%Y-%m-%d").date()
    except ValueError:
        start_date = default_start
    try:
        end_date = datetime.strptime(request.GET.get("date_to") or "", "%Y-%m-%d").date()
    except ValueError:
        end_date = today
    if end_date < start_date:
        start_date, end_date = end_date, start_date

    closed_jobs = JobCard.objects.select_related("claim", "claim__vehicle", "vehicle").filter(
        repair_status="Closed", claim__isnull=False,
    ).filter(Q(claim__status="Closed") | Q(claim__claim_stage=ClaimStageCode.CLOSED))
    for job in closed_jobs[:200]:
        vehicle = job.vehicle or (job.claim.vehicle if job.claim and job.claim.vehicle_id else None)
        if vehicle and not GateInEntry.objects.filter(jobcard=job).exists():
            GateInEntry.objects.create(
                registration_no=vehicle.registration_no,
                current_km=job.km or 0,
                service_type="Bodyshop",
                gate_in_datetime=job.gate_in_datetime or timezone.now(),
                branch=job.branch,
                vehicle=vehicle,
                status="Converted",
                jobcard=job,
                remarks="Created for Gate-Out; original Gate-In entry was not recorded.",
            )

    entries = GateInEntry.objects.select_related(
        "vehicle", "jobcard", "jobcard__claim", "gate_out_by"
    ).filter(
        Q(gate_in_datetime__date__range=(start_date, end_date))
        | Q(gate_out_datetime__date__range=(start_date, end_date))
        | Q(created_at__date__range=(start_date, end_date))
    ).distinct().order_by("gate_in_datetime")
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Gate In Out Register"
    sheet.append(["Gate In / Out Register", f"{start_date:%d-%m-%Y} to {end_date:%d-%m-%Y}"])
    sheet.append(["Vehicle No", "Job Card No", "Gate In Date", "Gate Out Date", "Status", "Gate Pass No", "Gate-Out Authorized By", "Service Type", "In KM", "Out KM", "Remarks"])
    for cell in sheet[1]:
        cell.font = Font(bold=True, size=14)
    for cell in sheet[2]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    for entry in entries:
        sheet.append([
            entry.registration_no,
            entry.jobcard.job_no if entry.jobcard_id else "",
            entry.gate_in_datetime.strftime("%d-%m-%Y %H:%M") if entry.gate_in_datetime else "",
            entry.gate_out_datetime.strftime("%d-%m-%Y %H:%M") if entry.gate_out_datetime else "",
            entry.status,
            entry.gate_pass_no,
            entry.gate_out_by.get_full_name() or entry.gate_out_by.username if entry.gate_out_by_id else "",
            entry.service_type,
            entry.current_km,
            entry.out_km or "",
            entry.remarks,
        ])
    for column_cells in sheet.columns:
        width = min(max(len(str(cell.value or "")) for cell in column_cells) + 2, 30)
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = width
    sheet.freeze_panes = "A3"
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="gate_register_{start_date:%Y%m%d}_{end_date:%Y%m%d}.xlsx"'
    workbook.save(response)
    return response


@require_POST
@login_required
def gate_in_entry_update(request, pk):
    entry = get_object_or_404(GateInEntry, pk=pk)
    if entry.status == "Converted":
        messages.error(request, "Converted Gate In entry cannot be edited.")
        return redirect("gate_in_entry")

    registration_no = (request.POST.get("registration_no") or "").strip().upper()
    current_km_raw = (request.POST.get("current_km") or "").strip()
    service_type = (request.POST.get("service_type") or "").strip()
    remarks = (request.POST.get("remarks") or "").strip()
    cancel_entry = request.POST.get("cancel_entry") == "on"
    cancellation_remark = (request.POST.get("cancellation_remark") or "").strip()

    if not registration_no:
        messages.error(request, "Vehicle No is required.")
        return redirect("gate_in_entry")

    try:
        current_km = int(current_km_raw)
    except (TypeError, ValueError):
        current_km = 0

    if current_km <= 0:
        messages.error(request, "Enter valid Current KM.")
        return redirect("gate_in_entry")

    if service_type not in dict(GateInEntry.SERVICE_TYPE_CHOICES):
        messages.error(request, "Select valid Service Type.")
        return redirect("gate_in_entry")

    if cancel_entry and not cancellation_remark:
        messages.error(request, "Cancel remark is required when marking Gate In as cancelled.")
        return redirect("gate_in_entry")

    duplicate = GateInEntry.objects.filter(
        registration_no__iexact=registration_no,
        status="Pending",
    ).exclude(pk=entry.pk).first()
    if duplicate and not cancel_entry:
        messages.error(
            request,
            f"{registration_no} already has another pending Gate In entry from "
            f"{timezone.localtime(duplicate.gate_in_datetime):%d/%m/%Y %H:%M}.",
        )
        return redirect("gate_in_entry")

    entry.registration_no = registration_no
    entry.current_km = current_km
    entry.service_type = service_type
    entry.vehicle = Vehicle.objects.filter(registration_no__iexact=registration_no).first()
    entry.remarks = remarks

    update_fields = [
        "registration_no",
        "current_km",
        "service_type",
        "vehicle",
        "remarks",
        "updated_at",
    ]

    if cancel_entry:
        entry.status = "Cancelled"
        entry.cancellation_remark = cancellation_remark
        entry.cancelled_at = timezone.now()
        entry.cancelled_by = request.user
        update_fields.extend(["status", "cancellation_remark", "cancelled_at", "cancelled_by"])

    entry.save(update_fields=update_fields)
    notify_reception_gate_in_changed(entry, cancelled=cancel_entry)
    messages.success(
        request,
        f"Gate In {'cancelled' if cancel_entry else 'updated'} for {entry.registration_no}.",
    )
    return redirect("gate_in_entry")


@require_POST
@login_required
def gate_in_entry_gate_out(request, pk):
    entry = get_object_or_404(GateInEntry.objects.select_related("jobcard", "jobcard__claim"), pk=pk)
    if entry.status != "Converted" or not entry.jobcard_id:
        messages.error(request, "Gate Out is available only for Converted entries with a Job Card.")
        return redirect("gate_in_entry")
    claim = entry.jobcard.claim if entry.jobcard.claim_id else None
    if not claim or entry.jobcard.repair_status != "Closed" or not (
        claim.status == "Closed" or claim.claim_stage == ClaimStageCode.CLOSED
    ):
        messages.error(request, "Claim and Job Card must be closed before Gate Out.")
        return redirect("gate_in_entry")
    gate_pass_no = (request.POST.get("gate_pass_no") or "").strip()
    try:
        out_km = int((request.POST.get("out_km") or "").strip())
    except (TypeError, ValueError):
        out_km = 0
    evidence = request.FILES.get("gate_pass_evidence")
    if not gate_pass_no or not evidence or out_km <= 0:
        messages.error(request, "Gate Pass No., Out KM, and physical pass evidence are required.")
        return redirect("gate_in_entry")
    entry.gate_pass_no = gate_pass_no
    entry.out_km = out_km
    entry.gate_pass_evidence = evidence
    entry.status = "Gate Out"
    entry.gate_out_datetime = timezone.now()
    entry.gate_out_by = request.user
    entry.save(update_fields=["gate_pass_no", "out_km", "gate_pass_evidence", "status", "gate_out_datetime", "gate_out_by", "updated_at"])
    messages.success(request, f"Gate Out completed for {entry.registration_no}.")
    return redirect("gate_in_entry")


@require_POST
@login_required
def gate_in_entry_cancel(request, pk):
    entry = get_object_or_404(GateInEntry, pk=pk)
    if entry.status == "Converted":
        messages.error(request, "Converted Gate In entry cannot be cancelled.")
        return redirect("gate_in_entry")

    cancellation_remark = (request.POST.get("cancellation_remark") or "").strip()
    if not cancellation_remark:
        messages.error(request, "Cancel remark is required.")
        return redirect("gate_in_entry")

    entry.status = "Cancelled"
    entry.cancellation_remark = cancellation_remark
    entry.cancelled_at = timezone.now()
    entry.cancelled_by = request.user
    entry.save(update_fields=["status", "cancellation_remark", "cancelled_at", "cancelled_by", "updated_at"])
    notify_reception_gate_in_changed(entry, cancelled=True)
    messages.success(request, f"Gate In cancelled for {entry.registration_no}.")
    return redirect("gate_in_entry")


# Create your views here.
@login_required
def dashboard(request):
    from datetime import date
    from django.utils.dateparse import parse_date

    logged_emp = Employee.objects.filter(
        user=request.user
    ).first()

    from apps.quality_check.views import is_quality_inspector

    if is_quality_inspector(logged_emp):
        return redirect("quality_inspector_dashboard")

    if is_parts_manager(logged_emp):
        return redirect("parts_manager_dashboard")

    if (
        is_reception_employee(logged_emp)
        and (logged_emp.employee_type or "").upper() != "ADVISOR"
    ):
        branch = logged_emp.branch if logged_emp and logged_emp.branch_id else None

        gate_entries = GateInEntry.objects.select_related(
            "vehicle",
            "vehicle__customer",
            "branch",
            "entered_by",
        ).filter(
            service_type="Bodyshop",
            status="Pending",
            jobcard__isnull=True,
        )
        pending_claims = Claim.objects.select_related(
            "vehicle",
            "vehicle__customer",
            "vehicle__model",
            "branch",
        ).filter(
            claim_stage=ClaimStageCode.CLAIM_CREATED,
            employee__isnull=True,
        ).exclude(
            status="Closed",
        )

        if branch:
            gate_entries = gate_entries.filter(Q(branch=branch) | Q(branch__isnull=True))
            pending_claims = pending_claims.filter(Q(branch=branch) | Q(branch__isnull=True))

        gate_rows = []
        for entry in gate_entries.order_by("gate_in_datetime")[:30]:
            gate_rows.append({
                "id": entry.id,
                "registration_no": entry.registration_no,
                "customer": entry.vehicle.customer.name if entry.vehicle_id and entry.vehicle.customer_id else "",
                "current_km": entry.current_km,
                "gate_in_datetime": entry.gate_in_datetime,
                "gate_in_display": timezone.localtime(entry.gate_in_datetime).strftime("%d/%m/%Y %H:%M"),
                "waiting": compact_wait_time(entry.gate_in_datetime),
                "branch": entry.branch.name if entry.branch_id else "",
                "remarks": entry.remarks,
            })

        claim_rows = []
        for claim in pending_claims.order_by("created_at")[:30]:
            vehicle = claim.vehicle if claim.vehicle_id else None
            claim_rows.append({
                "id": claim.id,
                "claim_no": claim.claim_no,
                "registration_no": vehicle.registration_no if vehicle else "",
                "customer": vehicle.customer.name if vehicle and vehicle.customer_id else "",
                "model": vehicle.model.name if vehicle and vehicle.model_id else "",
                "created_at": claim.created_at,
                "created_display": timezone.localtime(claim.created_at).strftime("%d/%m/%Y %H:%M"),
                "waiting": compact_wait_time(claim.created_at),
                "branch": claim.branch.name if claim.branch_id else "",
            })

        return render(request, "dashboard/reception_home.html", {
            "logged_emp": logged_emp,
            "gate_rows": gate_rows,
            "claim_rows": claim_rows,
            "gate_pending_count": gate_entries.count(),
            "advisor_pending_count": pending_claims.count(),
            "branch": branch,
        })

    if is_repair_resource(logged_emp):
        return redirect("my_work_list")

    if is_floor_supervisor(logged_emp):
        return redirect("work_allocation_list")

    claims = Claim.objects.none()
    jobcards = JobCard.objects.none()

    show_manager_dashboard = False

    # ADMIN
    if request.user.is_superuser:

        claims = Claim.objects.all()
        jobcards = JobCard.objects.all()
        show_manager_dashboard = True

    # MANAGER
    elif logged_emp and logged_emp.employee_type == "MANAGER":

        claims = branch_scoped_queryset_for_user(
            Claim.objects.all(),
            request.user,
        )
        jobcards = branch_scoped_queryset_for_user(
            JobCard.objects.all(),
            request.user,
            "claim__branch",
        )
        show_manager_dashboard = True

    # ADVISOR
    elif logged_emp and logged_emp.employee_type == "Advisor":

        claims = Claim.objects.filter(
            employee=logged_emp
        )

        jobcards = JobCard.objects.filter(
            advisor=logged_emp
        )

    # STAFF / RECEPTION
    elif logged_emp and logged_emp.employee_type in [
        "STAFF",
        "RECEPTION",
        "ADMIN",
    ]:

        claims = Claim.objects.filter(
            employee__isnull=True
        )

    today = date.today()
    default_from_date = today.replace(day=1)
    dashboard_is_admin = is_admin_user(request.user, logged_emp)
    dashboard_branches = DashboardLookupRepository.active_branches()
    dashboard_branch_id = request.GET.get("branch_id") or ""
    dashboard_selected_branch = (
        dashboard_branches.filter(pk=dashboard_branch_id).first()
        if dashboard_is_admin and dashboard_branch_id
        else None
    )
    from_date = parse_date(request.GET.get("from_date") or "") or default_from_date
    to_date = parse_date(request.GET.get("to_date") or "") or today
    status_scope = request.GET.get("status_scope") or ""
    main_status = request.GET.get("main_status") or ""
    advisor_id = request.GET.get("advisor") or ""

    if dashboard_selected_branch:
        claims = claims.filter(branch=dashboard_selected_branch)
        jobcards = jobcards.filter(claim__branch=dashboard_selected_branch)

    if from_date:
        jobcards = jobcards.filter(created_at__date__gte=from_date)

    if to_date:
        jobcards = jobcards.filter(created_at__date__lte=to_date)

    if advisor_id:
        claims = claims.filter(employee_id=advisor_id)
        jobcards = jobcards.filter(advisor_id=advisor_id)

    if main_status:
        if status_scope == "claim":
            claims = claims.filter(status=main_status)
        elif status_scope == "jobcard":
            jobcards = jobcards.filter(repair_status=main_status)
        else:
            claims = claims.filter(status=main_status)
            jobcards = jobcards.filter(repair_status=main_status)

    claims_for_metrics = claims
    claims = claims.filter(
        created_at__date__gte=from_date,
        created_at__date__lte=to_date,
    )

    advisor_options = DashboardLookupRepository.advisor_options(
        employee=logged_emp,
        selected_branch=dashboard_selected_branch,
        is_admin=is_admin_user(request.user, logged_emp),
    )

    # MANAGER REPORT DEFAULTS
    total_claims = 0
    pending_claims = 0
    closed_claims = 0
    work_allocation_pending = 0
    repair_in_progress = 0
    total_estimate_value = 0
    dashboard_financial = DashboardFinancialService(
        claims=claims_for_metrics,
        start_date=from_date,
        end_date=to_date,
    ).get()

    stage_counts = []
    advisor_counts = []
    recent_jobs = []
    advisor_dashboard = None

    if show_manager_dashboard:
        metrics = DashboardMetricsService(
            claims=claims_for_metrics,
            period_claims=claims,
            jobcards=jobcards,
            start_date=from_date,
            end_date=to_date,
        ).get()
        total_claims = metrics["total_claims"]
        pending_claims = metrics["pending_claims"]
        closed_claims = metrics["closed_claims"]
        work_allocation_pending = metrics["work_allocation_pending"]
        repair_in_progress = metrics["repair_in_progress"]
        stage_counts = dashboard_stage_rows(metrics["stage_counts"])
        advisor_counts = metrics["advisor_counts"]
        total_estimate_value = metrics["total_estimate_value"]
        recent_jobs = DashboardLookupRepository.recent_jobs(jobcards)

    if logged_emp and (logged_emp.employee_type or "").upper() == "ADVISOR":
        advisor_dashboard = AdvisorDashboardReadService(
            claims=claims,
            jobcards=jobcards,
        ).get()

    return render(request, "index.html", {
        "logged_emp": logged_emp,

        "claims": claims,
        "jobcards": jobcards,

        "show_manager_dashboard": show_manager_dashboard,

        "total_claims": total_claims,
        "pending_claims": pending_claims,
        "closed_claims": closed_claims,
        "work_allocation_pending": work_allocation_pending,
        "repair_in_progress": repair_in_progress,
        "stage_counts": stage_counts,
        "advisor_counts": advisor_counts,
        "recent_jobs": recent_jobs,
        "total_estimate_value": total_estimate_value,
        "dashboard_financial": dashboard_financial,
        "advisor_dashboard": advisor_dashboard,
        "dashboard_greeting": (
            "Good Morning" if timezone.localtime().hour < 12
            else "Good Afternoon" if timezone.localtime().hour < 17
            else "Good Evening"
        ),
        "advisor_options": advisor_options,
        "filter_from_date": from_date.strftime("%Y-%m-%d"),
        "filter_to_date": to_date.strftime("%Y-%m-%d"),
        "filter_status_scope": status_scope,
        "filter_main_status": main_status,
        "filter_advisor": advisor_id,
        "dashboard_is_admin": dashboard_is_admin,
        "dashboard_branches": dashboard_branches,
        "dashboard_branch_id": dashboard_branch_id,
        "dashboard_branch": logged_emp.branch if logged_emp and logged_emp.branch_id and not is_admin_user(request.user, logged_emp) else None,
        "claim_status_choices": Claim.STATUS_CHOICES,
        "jobcard_status_choices": [
            ("Open", "Open"),
            ("Completed", "Completed"),
            ("Closed", "Closed"),
            ("Cancellation", "Cancellation"),
        ],
    })


@login_required
def register_view(request):
    if not request.user.is_superuser:
        messages.error(request, "Only Admin can create users.")
        return redirect("dashboard")

    if request.method == 'POST':  # ✅ use uppercase
        form = UserCreationForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('login')  # ✅ redirect after success
        else:
            # Invalid form → re-render with errors
            return render(request, "registration/register.html", {"form": form})
    else:
        # GET request → show blank form
        form = UserCreationForm()
        return render(request, "registration/register.html", {"form": form})


def logout_view(request):
    # ✅ Remove cart data if it exists
    # request.session.pop('cart', {})

    request.session.flush()

    # ✅ Log out the user
    logout(request)

    # ✅ Redirect to homepage
    return redirect('login')  # 'home' should be the name of your homepage URLlogin_required


@login_required
def insurance_list(request):
    context = {

        "breadcrumbs": [

            {
                "title": "Master",
                "icon": "fa fa-database"
            },

            {
                "title": "Insurance List",
                "icon": "fa fa-users"
            }
        ]
    }
    return render(request, 'insurance/list.html', context)


@login_required
def insurance_data(request):
    data = list(InsuranceCompany.objects.values())
    return JsonResponse({'data': data})


@login_required
def insurance_get(request, pk):
    obj = get_object_or_404(InsuranceCompany, pk=pk)
    return JsonResponse({
        'id': obj.id,
        'ins_co_name': obj.ins_co_name,
        'city': obj.city,
        'mobile_no': obj.mobile_no,
    })


@login_required
def insurance_save(request):
    if request.method == 'POST':
        pk = request.POST.get('id')

        if pk:
            obj = get_object_or_404(InsuranceCompany, pk=pk)
            form = InsuranceCompanyForm(request.POST, instance=obj)
        else:
            form = InsuranceCompanyForm(request.POST)

        if form.is_valid():
            form.save()
            return JsonResponse({'success': True})

        return JsonResponse({'success': False, 'errors': form.errors})


@login_required
def insurance_edit(request, pk):
    obj = get_object_or_404(InsuranceCompany, pk=pk)
    form = InsuranceCompanyForm(request.POST or None, instance=obj)

    if form.is_valid():
        form.save()
        return redirect('insurance')

    return render(request, 'insurance/edit.html', {'form': form})


@login_required
def vehicle_list(request):
    return render(request, 'master/vehicle_list.html')


from .models import VehicleVariant


@login_required
def load_variants(request):
    model_id = request.GET.get('model_id')
    variants = VehicleVariant.objects.filter(model_id=model_id).values('id', 'name')
    return JsonResponse(list(variants), safe=False)


@never_cache
@login_required
def vehicle_list_api(request):
    data = list(
        Vehicle.objects.select_related('model', 'variant', 'customer', 'insurance_company')
        .values(
            'id',
            'registration_no',
            'chassis_no',
            'engine_no',
            'model__name',
            'variant__name',
            'color',
            'sale_date',
            'insurance_company',
            'insurance_company__ins_co_name',
            'policy_no',
            'policy_start_date',
            'policy_end_date',
            'last_service_km',
            'last_service_type',
            'last_service_date',
            'vehicle_type',
            'customer',
            'customer__name'
        )
    )
    return JsonResponse(data, safe=False)


from .models import Vehicle

from django.views.decorators.http import require_POST


@require_POST
@login_required
def vehicle_update_api(request, pk):
    try:
        vehicle = get_object_or_404(Vehicle, pk=pk)
        registration_no = normalize_vehicle_number(
            request.POST.get("registration_no", vehicle.registration_no)
        )

        if not is_valid_vehicle_number(registration_no):
            return JsonResponse({
                "status": "error",
                "errors": {
                    "registration_no": [VEHICLE_NUMBER_ERROR]
                }
            }, status=400)

        if Vehicle.objects.filter(registration_no__iexact=registration_no).exclude(pk=vehicle.pk).exists():
            return JsonResponse({
                "status": "error",
                "errors": {
                    "registration_no": ["Registration number already exists"]
                }
            }, status=400)

        customer_id = request.POST.get("customer")
        model_id = request.POST.get("model")
        variant_id = request.POST.get("variant")
        insurance_company_id = request.POST.get("insurance_company")

        if customer_id:
            vehicle.customer_id = customer_id

        if model_id:
            vehicle.model_id = model_id

        if variant_id:
            vehicle.variant_id = variant_id

        vehicle.insurance_company_id = insurance_company_id or None

        vehicle.registration_no = registration_no

        vehicle.chassis_no = request.POST.get(
            "chassis_no",
            vehicle.chassis_no
        )

        vehicle.engine_no = request.POST.get(
            "engine_no",
            vehicle.engine_no
        )

        vehicle.color = request.POST.get(
            "color",
            vehicle.color
        )

        vehicle.vehicle_type = request.POST.get(
            "vehicle_type",
            vehicle.vehicle_type
        )

        vehicle.sale_date = request.POST.get(
            "sale_date",
            vehicle.sale_date
        )

        vehicle.policy_no = request.POST.get(
            "policy_no",
            vehicle.policy_no
        )

        vehicle.policy_start_date = request.POST.get("policy_start_date") or None
        vehicle.policy_end_date = request.POST.get("policy_end_date") or None
        vehicle.primary_driver_id = request.POST.get("primary_driver") or None
        assigned_driver_ids = request.POST.getlist("assigned_drivers")
        if len(assigned_driver_ids) > 5:
            return JsonResponse({"status": "error", "errors": {"assigned_drivers": ["A vehicle can have a maximum of 5 drivers."]}}, status=400)
        if request.FILES.get("rc_document"):
            vehicle.rc_document = request.FILES["rc_document"]
        if request.FILES.get("insurance_policy_document"):
            vehicle.insurance_policy_document = request.FILES["insurance_policy_document"]
        vehicle.last_service_km = request.POST.get("last_service_km") or None
        vehicle.last_service_type = request.POST.get(
            "last_service_type",
            vehicle.last_service_type
        )
        vehicle.last_service_date = request.POST.get("last_service_date") or None

        vehicle.save()
        DriverMaster.objects.filter(vehicle=vehicle).update(vehicle=None)
        DriverMaster.objects.filter(id__in=assigned_driver_ids).update(vehicle=vehicle)

        return JsonResponse({
            "status": "success",
            "id": vehicle.id
        })

    except Exception as e:

        return JsonResponse({
            "status": "error",
            "message": str(e)
        }, status=500)


@login_required
def vehicle_create(request):
    if request.method == 'POST':
        form = VehicleForm(request.POST, request.FILES)

        if form.is_valid():
            vehicle = form.save()

            DriverMaster.objects.filter(vehicle=vehicle).update(vehicle=None)

            DriverMaster.objects.filter(
                id__in=[
                    driver.id
                    for driver in form.cleaned_data.get("assigned_drivers", [])
                ]
            ).update(vehicle=vehicle)

            return JsonResponse({
                "status": "success",
                "id": vehicle.id,
                "text": f"{vehicle.registration_no} - {vehicle.model.name if vehicle.model else ''}"
            })

        return JsonResponse({
            "status": "error",
            "errors": form.errors
        })

    return render(request, "master/vehicle_new.html")

@login_required
def check_registration(request):
    reg = request.GET.get('registration_no')

    exists = Vehicle.objects.filter(
        registration_no__iexact=reg
    ).exists()

    return JsonResponse({
        'exists': exists
    })


@login_required
def add_model_ajax(request):
    if request.method == "POST":
        data = json
        name = data.get("name").strip()

        if VehicleModel.objects.filter(name__iexact=name).exists():
            return JsonResponse({
                "status": "error",
                "message": "Model already exists"
            })

        model = VehicleModel.objects.create(name=name)

        return JsonResponse({
            "status": "success",
            "id": model.id,
            "name": model.name
        })


@login_required
def add_variant_ajax(request):
    if request.method == "POST":
        data = json

        model_id = data.get('model_id')
        name = data.get('name').strip()

        if VehicleVariant.objects.filter(model_id=model_id, name__iexact=name).exists():
            return JsonResponse({
                "status": "error",
                "message": "Variant already exists for this model"
            })

        variant = VehicleVariant.objects.create(
            model_id=model_id,
            name=name
        )

        return JsonResponse({
            "status": "success",
            "id": variant.id,
            "name": variant.name
        })


@login_required
def check_customer(request):
    name = request.GET.get("name", "").strip()
    mobile = request.GET.get("mobile", "").strip()

    exists = Customer.objects.filter(
        name__iexact=name,
        mobile_no=mobile
    ).exists()

    return JsonResponse({"exists": exists})


@login_required
def customer_search(request):
    term = request.GET.get('term', '').strip()

    customers = Customer.objects.filter(
        Q(name__icontains=term)
        | Q(mobile_no__icontains=term)
        | Q(whatsapp_no__icontains=term)
        | Q(customer_code__icontains=term)
    )[:10]

    results = [
        {
            'id': c.id,
            'text': f"{c.customer_code or ''} - {c.name} ({c.mobile_no or c.whatsapp_no or ''})"
        }
        for c in customers
    ]

    return JsonResponse({'results': results})


@login_required
def add_customer(request):
    data = json

    name = data.get("name", "").strip()
    mobile = data.get("mobile", "").strip()
    whatsapp_no = data.get("whatsapp_no", "").strip()
    email = data.get("email", "").strip()

    if not name:
        return JsonResponse({"status": "error", "message": "Name required"})

    # 🔥 DUPLICATE CHECK
    existing = Customer.objects.filter(
        name__iexact=name,
        mobile_no=mobile
    ).first()

    if existing:
        return JsonResponse({
            "status": "exists",
            "id": existing.id,
            "text": f"{existing.name} - {existing.mobile_no or ''}",
            "message": "Customer already exists"
        })

    # ✅ CREATE
    customer = Customer.objects.create(
        name=name,
        mobile_no=mobile,
        whatsapp_no=whatsapp_no,
        email=email
    )

    return JsonResponse({
        "status": "success",
        "id": customer.id,
        "text": f"{customer.name} - {customer.mobile_no or ''}"
    })


@login_required
def get_customer_details(request):
    customer_id = request.GET.get("id")

    # ✅ FIX 1: empty check
    if not customer_id:
        return JsonResponse({
            "status": "error",
            "message": "No customer selected"
        })

    # ✅ FIX 2: numeric validation
    if not str(customer_id).isdigit():
        return JsonResponse({
            "status": "error",
            "message": "Invalid customer ID"
        })

    try:
        c = Customer.objects.get(id=customer_id)

        return JsonResponse({
            "status": "success",
            "data": {
                "customer_code": c.customer_code,
                "customer_type": c.customer_type,
                "salutation": c.salutation,
                "name": c.name,
                "gender": c.gender,
                "date_of_birth": c.date_of_birth,
                "anniversary_date": c.anniversary_date,
                "gst_registered": c.gst_registered,
                "mobile": c.mobile_no,
                "alternate_mobile_no": c.alternate_mobile_no,
                "whatsapp_no": c.whatsapp_no,
                "email": c.email,
                "preferred_contact_method": c.preferred_contact_method,
                "address_line_1": c.address_line_1,
                "address_line_2": c.address_line_2,
                "city": c.city,
                "state": c.state,
                "gst": c.gst_no,
                "pan_no": c.pan_no,
                "aadhaar_no": c.aadhaar_no,
                "address": c.address,
                "pin_code": c.pin_code,
                "country": c.country,
                "company_name": c.company_name,
                "contact_person": c.contact_person,
                "designation": c.designation,
                "company_gst_no": c.company_gst_no
            }
        })

    except Customer.DoesNotExist:
        return JsonResponse({
            "status": "error",
            "message": "Customer not found"
        })


@login_required
def customer_list(request):
    context = {

        "breadcrumbs": [

            {
                "title": "Master",
                "icon": "fa fa-database"
            },

            {
                "title": "Customer List",
                "icon": "fa fa-users"
            }
        ]
    }

    return render(
        request,
        "master/Customer_list.html",
        context
    )


@login_required
def customer_data(request):
    data = list(Customer.objects.values())
    return JsonResponse({'data': data})


@login_required
def customer_get(request, id):
    obj = Customer.objects.get(id=id)
    return JsonResponse({
        'id': obj.id,
        'customer_code': obj.customer_code,
        'customer_type': obj.customer_type,
        'name': obj.name,
        'salutation': obj.salutation,
        'gender': obj.gender,
        'date_of_birth': obj.date_of_birth,
        'anniversary_date': obj.anniversary_date,
        'gst_registered': obj.gst_registered,
        'gst_no': obj.gst_no,
        'pan_no': obj.pan_no,
        'aadhaar_no': obj.aadhaar_no,
        'mobile_no': obj.mobile_no,
        'alternate_mobile_no': obj.alternate_mobile_no,
        'whatsapp_no': obj.whatsapp_no,
        'email': obj.email,
        'preferred_contact_method': obj.preferred_contact_method,
        'address_line_1': obj.address_line_1,
        'address_line_2': obj.address_line_2,
        'address': obj.address,
        'city': obj.city,
        'state': obj.state,
        'pin_code': obj.pin_code,
        'country': obj.country,
        'company_name': obj.company_name,
        'contact_person': obj.contact_person,
        'designation': obj.designation,
        'company_gst_no': obj.company_gst_no,
    })


@login_required
def customer_save(request):
    if request.method == "POST":

        obj_id = request.POST.get("id")

        if obj_id:
            obj = Customer.objects.get(id=obj_id)
            form = CustomerForm(request.POST, instance=obj)
        else:
            form = CustomerForm(request.POST)

        if form.is_valid():
            form.save()
            return JsonResponse({"success": True})

        return JsonResponse({
            "success": False,
            "errors": form.errors
        })

    # core/views.py


@login_required
def customer_check_mobile(request):
    mobile = request.GET.get("mobile", "").strip()
    customer_id = request.GET.get("id", "").strip()

    qs = Customer.objects.filter(mobile_no=mobile)

    if customer_id.isdigit():
        qs = qs.exclude(id=customer_id)

    return JsonResponse({
        "exists": bool(mobile and qs.exists())
    })


@login_required
def save_column_pref(request):
    # ✅ Handle GET (for testing / safety)
    if request.method == "GET":
        return JsonResponse({
            "status": "error",
            "message": "Use POST request"
        })

    # ✅ Handle POST
    if request.method == "POST":

        if not request.user.is_authenticated:
            return JsonResponse({
                "status": "error",
                "message": "Login required"
            })

        try:
            data = json

            screen = data.get("screen")
            state = data.get("state")
            name = data.get("name", "default")

            if not screen or not state:
                return JsonResponse({
                    "status": "error",
                    "message": "Missing data"
                })

            ColumnPreference.objects.update_or_create(
                user=request.user,
                screen=screen,
                name=name,
                defaults={"state": state}
            )

            return JsonResponse({"status": "success"})

        except Exception as e:
            return JsonResponse({
                "status": "error",
                "message": str(e)
            })

    # ✅ Fallback (very important)
    return JsonResponse({
        "status": "error",
        "message": "Invalid request"
    })


@login_required
def load_column_pref(request):
    screen = request.GET.get("screen")
    name = request.GET.get("name", "default")

    try:
        pref = ColumnPreference.objects.get(
            user=request.user,
            screen=screen,
            name=name
        )
        return JsonResponse({"state": pref.state})

    except ColumnPreference.DoesNotExist:
        return JsonResponse({"state": []})


@login_required
def surveyor_page(request):
    form = SurveyorForm()
    return render(request, "master/surveyor.html", {"form": form})


@login_required
def surveyor_data(request):
    data = list(Surveyor.objects.values())
    return JsonResponse({"data": data})


@login_required
def surveyor_save(request):
    if request.method == "POST":
        try:
            surveyor_id = request.POST.get("id")

            if surveyor_id and surveyor_id.strip():
                obj = Surveyor.objects.get(id=int(surveyor_id))
                form = SurveyorForm(request.POST, instance=obj)
            else:
                form = SurveyorForm(request.POST)

            if form.is_valid():
                obj = form.save()
                return JsonResponse({"success": True, "id": obj.id})

            return JsonResponse({"success": False, "errors": form.errors})

        except IntegrityError:
            return JsonResponse({
                "success": False,
                "errors": {"mobile_no": ["Duplicate mobile or license"]}
            })

    return JsonResponse({"error": "Invalid request"}, status=400)


@login_required
def surveyor_get(request, id):
    data = Surveyor.objects.filter(id=id).values().first()
    return JsonResponse(data)


@login_required
def check_surveyor_mobile(request):
    mobile = request.GET.get("mobile")

    exists = Surveyor.objects.filter(mobile_no=mobile).exists()

    return JsonResponse({"exists": exists})


@login_required
def employee_page(request):
    form = EmployeeForm()
    branches = Branch.objects.filter(is_active=True).order_by("name")
    return render(request, "master/employee.html", {"form": form, "branches": branches})


@login_required
def driver_master(request):
    return render(request, "master/driver_master.html")


@login_required
def employee_data(request):
    data = [
        {
            **item,
            "branch_name": item["branch__name"] or "",
            "branch_code": item["branch__code"] or "",
        }
        for item in Employee.objects.select_related("branch").values(
            "id",
            "name",
            "employee_code",
            "mobile_no",
            "designation",
            "department",
            "employee_type",
            "branch_id",
            "branch__name",
            "branch__code",
        )
    ]
    return JsonResponse({"data": data})


@login_required
def employee_save(request):
    if request.method == "POST":

        emp_id = request.POST.get("id")

        if emp_id and emp_id.strip():
            obj = Employee.objects.get(id=int(emp_id))
            form = EmployeeForm(request.POST, instance=obj)
        else:
            form = EmployeeForm(request.POST)

        if form.is_valid():
            obj = form.save()
            return JsonResponse({"success": True, "id": obj.id})

        return JsonResponse({"success": False, "errors": form.errors})

    return JsonResponse({"error": "Invalid request"}, status=400)


@login_required
def employee_get(request, id):
    data = Employee.objects.filter(id=id).values().first()
    return JsonResponse(data)


from datetime import datetime


def generate_claim_no():
    return next_claim_no()


def generate_claim_no_for_user(user):
    return next_claim_no(branch_for_user(user))


from .models import Employee


@login_required
def add_vehicle(request):
    if request.method != "POST":
        return JsonResponse({
            "status": "error",
            "message": "Invalid request"
        })

    data = json

    registration_no = data.get("registration_no", "").strip().upper()
    chassis_no = data.get("chassis_no", "").strip()
    engine_no = data.get("engine_no", "").strip()

    # =========================
    # VALIDATION
    # =========================
    if not registration_no:
        return JsonResponse({
            "status": "error",
            "message": "Registration No required"
        })

    # =========================
    # DUPLICATE CHECK
    # =========================
    existing = Vehicle.objects.filter(
        registration_no__iexact=registration_no
    ).first()

    if existing:
        return JsonResponse({
            "status": "exists",
            "id": existing.id,
            "text": f"{existing.registration_no}",
            "message": "Vehicle already exists"
        })

    # =========================
    # CREATE VEHICLE
    # =========================
    vehicle = Vehicle.objects.create(
        registration_no=registration_no,
        chassis_no=chassis_no,
        engine_no=engine_no
    )

    return JsonResponse({
        "status": "success",
        "id": vehicle.id,
        "text": f"{vehicle.registration_no}"
    })


from django.db.models import Q


@login_required
def vehicle_search(request):
    term = request.GET.get('term', '').strip()
    gate_in_only = request.GET.get("gate_in") == "1"
    vehicle_filter = Q(
        Q(registration_no__icontains=term) |
        Q(chassis_no__icontains=term) |
        Q(engine_no__icontains=term) |
        Q(customer__name__icontains=term)
    )
    if gate_in_only:
        # A direct Job Card may only be started from an unused Gate In entry
        # belonging to the logged-in user's branch (admins can see all).
        gate_entries = GateInEntry.objects.filter(
            status="Pending", jobcard__isnull=True, vehicle__isnull=False
        )
        if not is_admin_user(request.user):
            user_branch = branch_for_user(request.user)
            gate_entries = gate_entries.filter(
                branch=user_branch if user_branch else None
            )
        vehicle_filter &= Q(gate_in_entries__in=gate_entries)

    vehicles = Vehicle.objects.filter(vehicle_filter).select_related(
        'model',
        'customer'
    ).distinct()[:10]

    results = [
        {
            'id': v.id,
            'text': (
                f"{v.registration_no} | "
                f"{v.customer.name if v.customer else ''} | "
                f"{v.model.name if v.model else ''}"
            )
        }
        for v in vehicles
    ]

    if gate_in_only:
        entry_map = {}
        for entry in GateInEntry.objects.filter(
            vehicle_id__in=[item["id"] for item in results],
            status="Pending", jobcard__isnull=True,
        ).order_by("-gate_in_datetime"):
            entry_map.setdefault(entry.vehicle_id, entry)
        for item in results:
            entry = entry_map.get(item["id"])
            if entry:
                item.update({
                    "gate_entry_id": entry.id,
                    "gate_in_display": timezone.localtime(entry.gate_in_datetime).strftime("%d/%m/%Y %I:%M %p"),
                    "current_km": entry.current_km,
                })

    return JsonResponse({
        'results': results
    })


@login_required
def get_vehicle_details(request):
    vehicle_id = request.GET.get("id")

    # =========================
    # EMPTY CHECK
    # =========================
    if not vehicle_id:
        return JsonResponse({
            "status": "error",
            "message": "No vehicle selected"
        })

    # =========================
    # NUMERIC VALIDATION
    # =========================
    if not str(vehicle_id).isdigit():
        return JsonResponse({
            "status": "error",
            "message": "Invalid vehicle ID"
        })

    try:

        v = Vehicle.objects.select_related(
            "model",
            "variant",
            "customer"
        ).get(id=vehicle_id)

        return JsonResponse({
            "status": "success",
            "data": {

                "registration_no": v.registration_no,
                "chassis_no": v.chassis_no,
                "engine_no": v.engine_no,

                "vehicle_type": v.vehicle_type,
                "color": v.color,

                "model": v.model.name if v.model else "",
                "variant": v.variant.name if v.variant else "",
                # CUSTOMER
                "customer_id": v.customer.id if v.customer else "",
                "customer_name": v.customer.name if v.customer else "",
                "mobile": v.customer.mobile_no if v.customer else "",
                "city": v.customer.city if v.customer else "",
                "gst": v.customer.gst_no if v.customer else "",

                "sale_date": (
                    v.sale_date.strftime("%Y-%m-%d")
                    if v.sale_date else ""
                )
            }
        })

    except Vehicle.DoesNotExist:

        return JsonResponse({
            "status": "error",
            "message": "Vehicle not found"
        })


# views.py

from .models import Claim
from .forms import ClaimForm


def job_save(self, *args, **kwargs):
    is_new = self.pk is None

    super().save(*args, **kwargs)

    # =====================================
    # CLAIM STAGE UPDATE
    # =====================================

    if self.claim:

        # JOB CARD CREATED
        if self.claim.employee_id and self.claim.claim_stage < ClaimStageCode.INTIMATION:
            self.claim.claim_stage = ClaimStageCode.INTIMATION
            self.claim.save(
                update_fields=["claim_stage"]
            )


def generate_job_no():
    return next_jobcard_no()


def generate_job_no_for_user(user):
    return next_jobcard_no(branch_for_user(user))


def generate_job_no_for_claim(claim):
    return next_jobcard_no(branch_for_claim(claim))


def get_inventory_context(job=None):
    inventory = None

    # SAFE INVENTORY LOAD
    if job:
        try:
            inventory = job.inventory
        except Exception:
            inventory = None

    # SAFE VALUES
    fuel_percent = 0
    cng_percent = 0

    if inventory:
        fuel_percent = inventory.fuel_percent or 0
        cng_percent = inventory.cng_percent or 0

    # GAUGE ANGLE
    fuel_angle = -90 + (fuel_percent * 1.8)
    cng_angle = -90 + (cng_percent * 1.8)

    # LABELS
    def get_fuel_label(value):

        value = int(value or 0)

        if value <= 0:
            return "Empty"

        elif value <= 25:
            return "1/4"

        elif value <= 50:
            return "Half"

        elif value <= 75:
            return "3/4"

        return "Full"

    # DAMAGE MARKS
    raw_marks = inventory.damage_marks if inventory else []

    damage_marks = []

    damage_image_ratio = 1117 / 736

    for raw_mark in raw_marks:
        m = dict(raw_mark)

        if m.get("type") == "scratch":
            x1 = float(m.get("x1", 0))
            y1 = float(m.get("y1", 0))
            x2 = float(m.get("x2", 0))
            y2 = float(m.get("y2", 0))

            dx = x2 - x1
            dy = y2 - y1

            m["length"] = round((dx * dx + (dy * damage_image_ratio) ** 2) ** 0.5, 2)
            m["angle"] = round(math.degrees(math.atan2(dy * damage_image_ratio, dx)), 2)

        damage_marks.append(m)
    tyre_map = {}

    if job:
        for t in job.tyres.all():
            tyre_map[t.position] = t

    tyre_positions = [
        ("front_left", "Front Left"),
        ("front_right", "Front Right"),
        ("rear_left", "Rear Left"),
        ("rear_right", "Rear Right"),
        ("stepney", "Stepney"),
    ]
    items = [
        ("lh_mirror", "LH Side Mirror"),
        ("jack", "Jack"),
        ("tool_kit", "Tool Kit"),
        ("floor_mat_count", "Floor Mat"),
        ("mud_flap_count", "Mud Flap"),
        ("stereo", "Stereo"),
        ("battery", "Battery"),
        ("rh_mirror", "RH Side Mirror"),
        ("number_plate", "Number Plate"),
        ("center_mirror", "Center Rear View Mirror"),
        ("frt_wiper", "Front Wiper"),
        ("rr_wiper", "Rear Wiper"),
        ("accessories", "Extra Accessories"),
    ]
    print("CTX INVENTORY:", inventory)
    fuel_percent = inventory.fuel_percent if inventory else 0
    cng_percent = inventory.cng_percent if inventory else 0

    fuel_label = get_fuel_label(fuel_percent)
    cng_label = get_fuel_label(cng_percent)
    return {
        "lh_mirror": inventory.lh_mirror if inventory else "",
        "jack": inventory.jack if inventory else "",
        "tool_kit": inventory.tool_kit if inventory else "",
        "floor_mat_count": inventory.floor_mat_count if inventory else "",
        "mud_flap_count": inventory.mud_flap_count if inventory else "",
        "stereo": inventory.stereo if inventory else "",
        "battery": inventory.battery if inventory else "",
        "rh_mirror": inventory.rh_mirror if inventory else "",
        "number_plate": inventory.number_plate if inventory else "",
        "center_mirror": inventory.center_mirror if inventory else "",
        "frt_wiper": inventory.frt_wiper if inventory else "",
        "rr_wiper": inventory.rr_wiper if inventory else "",
        "accessories": inventory.accessories if inventory else "",
        "inventory_remarks": inventory.remarks if inventory else "",
        "fuel_percent": inventory.fuel_percent if inventory else 0,
        "cng_percent": inventory.cng_percent if inventory else 0,
        "fuel_angle": fuel_angle,
        "cng_angle": cng_angle,
        "damage_marks_json": json.dumps(raw_marks),
        "damage_marks": damage_marks,
        "fuel_label": fuel_label,
        "cng_label": cng_label,
        "tyre_inventory": [
            {
                "position": key,
                "label": label,
                "make": tyre_map[key].make if key in tyre_map else "",
                "size": tyre_map[key].size if key in tyre_map else "",
                "depth": tyre_map[key].depth if key in tyre_map else "",
                "wheel_cap": tyre_map[key].wheel_cap if key in tyre_map else "",
            }
            for key, label in tyre_positions
        ],
    }


from .models import CompanySetup
from .forms import CompanySetupForm


@login_required
def company_setup(request):
    company = CompanySetup.objects.first()

    if request.method == 'POST':
        form = CompanySetupForm(
            request.POST,
            request.FILES,
            instance=company
        )

        if form.is_valid():
            form.save()
            return redirect('company_setup')

    else:
        form = CompanySetupForm(instance=company)

    return render(request, 'core/company_setup.html', {
        'form': form
    })


from django.db import transaction

from openpyxl import load_workbook

from .forms import ItemExcelUploadForm


@login_required
def upload_itemdata_excel(request):
    if request.method == "POST":
        action = request.POST.get("part_master_action") or ""

        if action == "toggle_status":
            item = get_object_or_404(ItemData, id=request.POST.get("item_id"))
            item.status = "Inactive" if item.status == "Active" else "Active"
            item.save(update_fields=["status"])
            messages.success(
                request,
                f"{item.item_code} marked {item.status}.",
            )
            return redirect("part")

        if action == "adjust_stock":
            item_id = request.POST.get("item_id")
            transaction_type = request.POST.get("transaction_type") or "Adjustment"
            quantity_value = (request.POST.get("quantity") or "").strip()
            direction = request.POST.get("direction") or "add"
            reference = (request.POST.get("reference") or "").strip()
            remarks = (request.POST.get("stock_remarks") or "").strip()

            try:
                quantity = Decimal(quantity_value)
                if quantity <= 0:
                    raise InvalidOperation
            except (InvalidOperation, ValueError):
                messages.error(request, "Stock quantity must be greater than zero.")
                return redirect("part")

            if transaction_type not in {
                value for value, _ in PartStockTransaction.TRANSACTION_CHOICES
            }:
                transaction_type = "Adjustment"
            quantity_change = -quantity if direction == "subtract" else quantity

            with transaction.atomic():
                item = get_object_or_404(
                    ItemData.objects.select_for_update(),
                    id=item_id,
                )
                new_balance = item.current_stock + quantity_change
                if new_balance < 0:
                    messages.error(
                        request,
                        f"Insufficient stock. Available balance is {item.current_stock} {item.unit}.",
                    )
                    return redirect("part")
                item.current_stock = new_balance
                item.save(update_fields=["current_stock", "updated_at"])
                PartStockTransaction.objects.create(
                    part=item,
                    transaction_type=transaction_type,
                    quantity_change=quantity_change,
                    balance_after=new_balance,
                    reference=reference,
                    remarks=remarks,
                    created_by=request.user,
                )

            messages.success(
                request,
                f"Stock updated for {item.item_code}. New balance: {new_balance} {item.unit}.",
            )
            return redirect("part")

        if action == "save":
            item_id = request.POST.get("item_id") or ""
            part_no = (request.POST.get("part_no") or "").strip().upper()
            part_description = (request.POST.get("part_description") or "").strip()
            model_name = (request.POST.get("model") or "").strip()
            rate_value = (request.POST.get("rate") or "0").strip()
            gst_value = (request.POST.get("gst_percent") or "0").strip()
            reorder_value = (request.POST.get("reorder_level") or "0").strip()
            opening_value = (request.POST.get("opening_stock") or "0").strip()
            item_status = (request.POST.get("status") or "Active").strip()

            if not part_no or not part_description:
                messages.error(request, "Part No and Part Description are required.")
                return redirect(f"{reverse('part')}?edit={item_id}" if item_id else "part")

            try:
                rate = Decimal(rate_value)
                gst_percent = Decimal(gst_value)
                reorder_level = Decimal(reorder_value)
                opening_stock = Decimal(opening_value)
                if (
                    rate < 0
                    or gst_percent < 0
                    or gst_percent > 100
                    or reorder_level < 0
                    or opening_stock < 0
                ):
                    raise InvalidOperation
            except (InvalidOperation, ValueError):
                messages.error(
                    request,
                    "Rate, stock and reorder values must be non-negative; GST must be between 0 and 100.",
                )
                return redirect(f"{reverse('part')}?edit={item_id}" if item_id else "part")

            if item_status not in {"Active", "Inactive"}:
                item_status = "Active"

            duplicate_qs = ItemData.objects.filter(item_code__iexact=part_no)

            if item_id:
                duplicate_qs = duplicate_qs.exclude(id=item_id)

            if duplicate_qs.exists():
                messages.error(request, f"Part No {part_no} already exists.")
                return redirect("part")

            if item_id:
                item = get_object_or_404(ItemData, id=item_id)
                message = "Part updated successfully."
                created = False
            else:
                item = ItemData()
                message = "Part saved successfully."
                created = True

            with transaction.atomic():
                item.item_code = part_no
                item.item_name = part_description
                item.category = model_name
                item.rate = rate
                item.status = item_status
                requested_unit = request.POST.get("unit") or "Nos"
                item.unit = (
                    requested_unit
                    if requested_unit in dict(ItemData.UNIT_CHOICES)
                    else "Nos"
                )
                item.manufacturer = (request.POST.get("manufacturer") or "").strip()
                item.hsn_code = (request.POST.get("hsn_code") or "").strip().upper()
                item.gst_percent = gst_percent
                item.preferred_supplier = (
                    request.POST.get("preferred_supplier") or ""
                ).strip()
                item.bin_location = (
                    request.POST.get("bin_location") or ""
                ).strip().upper()
                item.reorder_level = reorder_level
                if created:
                    item.current_stock = opening_stock
                item.save()
                if created and opening_stock:
                    PartStockTransaction.objects.create(
                        part=item,
                        transaction_type="Opening",
                        quantity_change=opening_stock,
                        balance_after=opening_stock,
                        remarks="Opening stock entered during part creation",
                        created_by=request.user,
                    )
            messages.success(request, message)
            return redirect("part")

        form = ItemExcelUploadForm(request.POST, request.FILES)

        if form.is_valid():
            excel_file = request.FILES["excel_file"]

            try:
                wb = load_workbook(excel_file, read_only=True, data_only=True)
                ws = wb.active
            except Exception:
                messages.error(request, "Unable to read the Excel file. Upload a valid .xlsx workbook.")
                return redirect("part")

            created_count = 0
            updated_count = 0
            skipped_count = 0
            error_rows = []

            with transaction.atomic():
                for row_no, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    item_code = str(row[0]).strip().upper() if row[0] else ""
                    item_name = str(row[1]).strip() if len(row) > 1 and row[1] else ""
                    category = str(row[2]).strip() if len(row) > 2 and row[2] else ""
                    raw_rate = row[3] if len(row) > 3 and row[3] not in (None, "") else 0
                    item_status = str(row[4]).strip().title() if len(row) > 4 and row[4] else "Active"
                    unit = str(row[5]).strip() if len(row) > 5 and row[5] else "Nos"
                    manufacturer = str(row[6]).strip() if len(row) > 6 and row[6] else ""
                    hsn_code = str(row[7]).strip() if len(row) > 7 and row[7] else ""
                    raw_gst = row[8] if len(row) > 8 and row[8] not in (None, "") else 0
                    supplier = str(row[9]).strip() if len(row) > 9 and row[9] else ""
                    bin_location = str(row[10]).strip() if len(row) > 10 and row[10] else ""
                    raw_stock = row[11] if len(row) > 11 and row[11] not in (None, "") else None
                    raw_reorder = row[12] if len(row) > 12 and row[12] not in (None, "") else 0

                    if not item_code or not item_name:
                        skipped_count += 1
                        error_rows.append(f"Row {row_no}: Part No and Description are required")
                        continue

                    try:
                        rate = Decimal(str(raw_rate))
                        gst_percent = Decimal(str(raw_gst))
                        reorder_level = Decimal(str(raw_reorder))
                        imported_stock = (
                            Decimal(str(raw_stock)) if raw_stock is not None else None
                        )
                        if (
                            rate < 0
                            or gst_percent < 0
                            or gst_percent > 100
                            or reorder_level < 0
                            or (imported_stock is not None and imported_stock < 0)
                        ):
                            raise InvalidOperation
                    except (InvalidOperation, ValueError):
                        skipped_count += 1
                        error_rows.append(f"Row {row_no}: Invalid rate")
                        continue

                    if item_status not in {"Active", "Inactive"}:
                        item_status = "Active"

                    existing = ItemData.objects.filter(
                        item_code__iexact=item_code
                    ).first()
                    previous_stock = existing.current_stock if existing else Decimal("0")
                    defaults = {
                        "item_code": item_code,
                        "item_name": item_name,
                        "category": category,
                        "rate": rate,
                        "status": item_status,
                        "unit": unit if unit in dict(ItemData.UNIT_CHOICES) else "Nos",
                        "manufacturer": manufacturer,
                        "hsn_code": hsn_code,
                        "gst_percent": gst_percent,
                        "preferred_supplier": supplier,
                        "bin_location": bin_location,
                        "reorder_level": reorder_level,
                    }
                    if imported_stock is not None:
                        defaults["current_stock"] = imported_stock
                    item, created = ItemData.objects.update_or_create(
                        item_code__iexact=item_code,
                        defaults=defaults,
                    )
                    if imported_stock is not None and imported_stock != previous_stock:
                        PartStockTransaction.objects.create(
                            part=item,
                            transaction_type="Opening" if created else "Adjustment",
                            quantity_change=imported_stock - previous_stock,
                            balance_after=imported_stock,
                            reference="Excel import",
                            created_by=request.user,
                        )
                    if created:
                        created_count += 1
                    else:
                        updated_count += 1

            messages.success(
                request,
                f"Import complete: {created_count} created, {updated_count} updated, "
                f"{skipped_count} skipped."
            )
            if error_rows:
                messages.warning(request, " | ".join(error_rows[:5]))

            return redirect("part")

        messages.error(request, "Invalid form or file not selected.")
        return redirect("part")

    else:
        form = ItemExcelUploadForm()

    search_text = (request.GET.get("q") or "").strip()
    selected_status = (request.GET.get("status") or "").strip()
    selected_category = (request.GET.get("category") or "").strip()
    selected_stock = (request.GET.get("stock") or "").strip()
    selected_supplier = (request.GET.get("supplier") or "").strip()
    edit_id = request.GET.get("edit") or ""
    items = ItemData.objects.all().order_by("item_code")

    if search_text:
        items = items.filter(
            Q(item_code__icontains=search_text)
            | Q(item_name__icontains=search_text)
            | Q(category__icontains=search_text)
        )
    if selected_status in {"Active", "Inactive"}:
        items = items.filter(status=selected_status)
    if selected_category:
        items = items.filter(category=selected_category)
    if selected_supplier:
        items = items.filter(preferred_supplier__icontains=selected_supplier)
    if selected_stock == "low":
        items = items.filter(
            current_stock__lte=F("reorder_level"),
            reorder_level__gt=0,
            status="Active",
        )
    elif selected_stock == "available":
        items = items.filter(current_stock__gt=0)
    elif selected_stock == "zero":
        items = items.filter(current_stock=0)

    edit_item = ItemData.objects.filter(id=edit_id).first() if edit_id else None
    model_options = (
        ItemData.objects.exclude(category__isnull=True)
        .exclude(category="")
        .values_list("category", flat=True)
        .distinct()
        .order_by("category")
    )
    paginator = Paginator(items, 25)
    page_obj = paginator.get_page(request.GET.get("page"))

    return render(request, "master/partmaster.html", {
        "form": form,
        "items": page_obj.object_list,
        "page_obj": page_obj,
        "edit_item": edit_item,
        "search_text": search_text,
        "selected_status": selected_status,
        "selected_category": selected_category,
        "selected_stock": selected_stock,
        "selected_supplier": selected_supplier,
        "model_options": model_options,
        "unit_choices": ItemData.UNIT_CHOICES,
        "stock_transaction_choices": PartStockTransaction.TRANSACTION_CHOICES,
        "recent_transactions": PartStockTransaction.objects.select_related(
            "part", "created_by"
        )[:10],
        "summary": {
            "total": ItemData.objects.count(),
            "active": ItemData.objects.filter(status="Active").count(),
            "inactive": ItemData.objects.filter(status="Inactive").count(),
            "categories": ItemData.objects.exclude(
                category__isnull=True
            ).exclude(category="").values("category").distinct().count(),
            "low_stock": ItemData.objects.filter(
                current_stock__lte=F("reorder_level"),
                reorder_level__gt=0,
                status="Active",
            ).count(),
            "stock_value": ItemData.objects.aggregate(
                total=Sum(F("current_stock") * F("rate"))
            )["total"] or Decimal("0"),
        },
    })


@login_required
def itemdata_list(request):
    return redirect("part")


@require_POST
@login_required
def part_master_quick_add(request):
    try:
        data = json.loads(request.body.decode("utf-8") or "{}")
    except (json.JSONDecodeError, UnicodeDecodeError):
        return JsonResponse(
            {"status": "error", "message": "Invalid request data."},
            status=400,
        )

    part_no = (data.get("part_no") or "").strip().upper()
    description = (data.get("description") or "").strip()
    category = (data.get("category") or "").strip()
    unit = (data.get("unit") or "Nos").strip()
    manufacturer = (data.get("manufacturer") or "").strip()
    supplier = (data.get("preferred_supplier") or "").strip()
    try:
        rate = Decimal(str(data.get("rate") or "0"))
        gst_percent = Decimal(str(data.get("gst_percent") or "0"))
        reorder_level = Decimal(str(data.get("reorder_level") or "0"))
        if (
            rate < 0
            or gst_percent < 0
            or gst_percent > 100
            or reorder_level < 0
        ):
            raise InvalidOperation
    except (InvalidOperation, ValueError):
        return JsonResponse(
            {
                "status": "error",
                "message": "Enter valid rate, GST and reorder values.",
            },
            status=400,
        )

    if not part_no or not description:
        return JsonResponse(
            {
                "status": "error",
                "message": "Part number and description are required.",
            },
            status=400,
        )
    if ItemData.objects.filter(item_code__iexact=part_no).exists():
        return JsonResponse(
            {
                "status": "error",
                "message": f"Part number {part_no} already exists.",
            },
            status=400,
        )
    if unit not in dict(ItemData.UNIT_CHOICES):
        unit = "Nos"

    item = ItemData.objects.create(
        item_code=part_no,
        item_name=description,
        category=category,
        rate=rate,
        status="Active",
        unit=unit,
        manufacturer=manufacturer,
        gst_percent=gst_percent,
        preferred_supplier=supplier,
        reorder_level=reorder_level,
    )
    return JsonResponse({
        "status": "success",
        "message": f"{part_no} added to Part Master.",
        "part": {
            "id": item.id,
            "part_no": item.item_code,
            "description": item.item_name,
            "unit": item.unit,
            "rate": str(item.rate),
            "current_stock": str(item.current_stock),
            "pending_requisition_qty": "0",
            "on_order_qty": "0",
            "available_after_pending": str(item.current_stock),
        },
    })


@login_required
def part_master_search(request):
    query = (request.GET.get("q") or "").strip()
    if len(query) < 3:
        return JsonResponse({"results": []})

    items = ItemData.objects.filter(status="Active").filter(
        Q(item_code__icontains=query)
        | Q(item_name__icontains=query)
    ).order_by("item_code")[:15]
    return JsonResponse({
        "results": [
            {
                "id": item.id,
                "part_no": item.item_code,
                "description": item.item_name,
                "unit": item.unit,
                "rate": str(item.rate),
                "current_stock": str(item.current_stock),
            }
            for item in items
        ]
    })


def sync_part_requisition_status(requisition):
    lines = list(requisition.lines.all())
    if not lines:
        status_value = "Submitted"
    elif all(line.fulfilled_qty >= line.requested_qty for line in lines):
        status_value = "Fulfilled"
    elif any(line.fulfilled_qty > 0 for line in lines):
        status_value = "Partially Fulfilled"
    else:
        status_value = "Submitted"
    if requisition.status != "Cancelled":
        requisition.status = status_value
        requisition.save(update_fields=["status", "updated_at"])
    return requisition.status


def is_advisor_employee(employee):
    if not employee:
        return False
    role_text = " ".join([
        employee.employee_type or "",
        employee.designation or "",
        employee.department or "",
    ]).upper()
    return "ADVISOR" in role_text


@require_POST
@login_required
def create_part_requisition_from_job(request, job_id):
    employee = Employee.objects.filter(user=request.user).first()
    if not request.user.is_superuser and not is_advisor_employee(employee):
        return JsonResponse(
            {
                "status": "error",
                "message": "Only an Advisor can create a job estimate requisition.",
            },
            status=403,
        )

    job = get_object_or_404(
        JobCard.objects.select_related("advisor").prefetch_related("parts"),
        id=job_id,
    )
    if (
        not request.user.is_superuser
        and job.advisor_id
        and job.advisor.user_id
        and job.advisor.user_id != request.user.id
    ):
        return JsonResponse(
            {
                "status": "error",
                "message": "This job card is assigned to another Advisor.",
            },
            status=403,
        )

    existing = job.part_requisitions.filter(
        status__in=["Submitted", "Partially Fulfilled"]
    ).first()
    if existing:
        return JsonResponse({
            "status": "success",
            "message": f"Open requisition {existing.requisition_no} already exists.",
            "redirect_url": reverse(
                "part_requisition_detail",
                args=[existing.id],
            ),
        })

    estimated_parts = list(job.parts.all().order_by("id"))
    if not estimated_parts:
        return JsonResponse(
            {
                "status": "error",
                "message": "Add estimated part lines before creating a requisition.",
            },
            status=400,
        )

    master_code_query = Q()
    for estimated_part in estimated_parts:
        master_code_query |= Q(item_code__iexact=estimated_part.part_no)
    master_by_code = {
        item.item_code.upper(): item
        for item in ItemData.objects.filter(status="Active").filter(
            master_code_query
        )
    }
    missing_parts = [
        part.part_no
        for part in estimated_parts
        if part.part_no.upper() not in master_by_code
    ]
    if missing_parts:
        return JsonResponse(
            {
                "status": "error",
                "message": (
                    "These estimated parts are missing or inactive in Part Master: "
                    + ", ".join(missing_parts[:8])
                ),
            },
            status=400,
        )

    with transaction.atomic():
        requisition = PartRequisition.objects.create(
            job=job,
            requested_by=request.user,
            priority="Normal",
            remarks="Created from approved job-card estimate.",
        )
        requisition.requisition_no = (
            f"PR-{timezone.localdate():%Y%m%d}-{requisition.id:04d}"
        )
        requisition.save(update_fields=["requisition_no", "updated_at"])
    PartRequisitionLine.objects.bulk_create([
            PartRequisitionLine(
                requisition=requisition,
                part=master_by_code[estimated_part.part_no.upper()],
                estimated_part=estimated_part,
                requested_qty=estimated_part.qty,
                remarks="Estimate line",
            )
            for estimated_part in estimated_parts
        ])

    notify_parts_team(
        "New Part Requisition",
        f"Advisor {request.user.get_full_name() or request.user.username} created {requisition.requisition_no} for job {job.job_no}.",
        reverse("part_requisition_detail", args=[requisition.id]),
    )

    return JsonResponse({
        "status": "success",
        "message": f"Requisition {requisition.requisition_no} created.",
        "redirect_url": reverse(
            "part_requisition_detail",
            args=[requisition.id],
        ),
    })


@never_cache
@login_required
def part_requisition_list(request):
    q = (request.GET.get("q") or "").strip()
    selected_status = (request.GET.get("status") or "").strip()
    requisitions = PartRequisition.objects.select_related(
        "job",
        "job__claim",
        "job__claim__vehicle",
        "requested_by",
    ).prefetch_related("lines")
    if q:
        requisitions = requisitions.filter(
            Q(requisition_no__icontains=q)
            | Q(job__job_no__icontains=q)
            | Q(job__claim__vehicle__registration_no__icontains=q)
        ).distinct()
    if selected_status:
        requisitions = requisitions.filter(status=selected_status)

    page_obj = Paginator(requisitions, 25).get_page(request.GET.get("page"))
    all_requisitions = PartRequisition.objects.all()
    return render(request, "parts/requisitionList.html", {
        "page_obj": page_obj,
        "requisitions": page_obj.object_list,
        "q": q,
        "selected_status": selected_status,
        "statuses": PartRequisition.STATUS_CHOICES,
        "summary": {
            "open": all_requisitions.filter(
                status__in=["Submitted", "Partially Fulfilled"]
            ).count(),
            "partial": all_requisitions.filter(
                status="Partially Fulfilled"
            ).count(),
            "fulfilled": all_requisitions.filter(status="Fulfilled").count(),
            "urgent": all_requisitions.filter(
                priority__in=["Urgent", "Vehicle Hold"],
                status__in=["Submitted", "Partially Fulfilled"],
            ).count(),
        },
    })


@never_cache
@login_required
def part_requisition_create(request):
    jobs = JobCard.objects.select_related(
        "claim",
        "claim__vehicle",
        "claim__vehicle__customer",
    ).order_by("-id")
    parts = ItemData.objects.filter(status="Active").order_by("item_code")

    if request.method == "POST":
        job = JobCard.objects.filter(id=request.POST.get("job_id")).first()
        part_ids = request.POST.getlist("part_id[]")
        quantities = request.POST.getlist("quantity[]")
        line_remarks = request.POST.getlist("line_remarks[]")
        valid_lines = []
        seen_parts = set()

        for index, part_id in enumerate(part_ids):
            if not part_id or part_id in seen_parts:
                continue
            try:
                quantity = Decimal(quantities[index])
                if quantity <= 0:
                    continue
            except (InvalidOperation, ValueError, IndexError):
                continue
            part = ItemData.objects.filter(id=part_id, status="Active").first()
            if part:
                seen_parts.add(part_id)
                valid_lines.append((
                    part,
                    quantity,
                    line_remarks[index].strip()
                    if index < len(line_remarks) else "",
                ))

        if not job:
            messages.error(request, "Select a valid job card.")
        elif not valid_lines:
            messages.error(request, "Add at least one valid part and quantity.")
        else:
            with transaction.atomic():
                requisition = PartRequisition.objects.create(
                    job=job,
                    requested_by=request.user,
                    needed_by=parse_date(request.POST.get("needed_by") or ""),
                    priority=request.POST.get("priority") or "Normal",
                    remarks=(request.POST.get("remarks") or "").strip(),
                )
                requisition.requisition_no = (
                    f"PR-{timezone.localdate():%Y%m%d}-{requisition.id:04d}"
                )
                requisition.save(update_fields=["requisition_no", "updated_at"])
                PartRequisitionLine.objects.bulk_create([
                    PartRequisitionLine(
                        requisition=requisition,
                        part=part,
                        requested_qty=quantity,
                        remarks=remarks_value,
                    )
                    for part, quantity, remarks_value in valid_lines
                ])
            notify_parts_team(
                "New Part Requisition",
                f"Advisor {request.user.get_full_name() or request.user.username} created {requisition.requisition_no} for job {job.job_no}.",
                reverse("part_requisition_detail", args=[requisition.id]),
            )
            messages.success(
                request,
                f"Requisition {requisition.requisition_no} submitted.",
            )
            return redirect("part_requisition_detail", requisition_id=requisition.id)

    return render(request, "parts/requisitionCreate.html", {
        "jobs": jobs,
        "parts": parts,
        "priorities": PartRequisition.PRIORITY_CHOICES,
        "unit_choices": ItemData.UNIT_CHOICES,
        "today": timezone.localdate(),
    })


@never_cache
@login_required
def part_requisition_detail(request, requisition_id):
    requisition = get_object_or_404(
        PartRequisition.objects.select_related(
            "job",
            "job__claim",
            "job__claim__vehicle",
            "job__claim__vehicle__customer",
            "requested_by",
        ).prefetch_related(
            "lines__part",
            "lines__fulfillments__issued_by",
        ),
        id=requisition_id,
    )

    if request.method == "POST":
        action = request.POST.get("action") or ""
        if action == "cancel":
            employee = Employee.objects.filter(user=request.user).first()
            if not (
                request.user.is_superuser
                or is_parts_manager(employee)
                or requisition.requested_by_id == request.user.id
            ):
                messages.error(request, "You cannot cancel this requisition.")
                return redirect(
                    "part_requisition_detail",
                    requisition_id=requisition.id,
                )
            if requisition.status == "Fulfilled":
                messages.error(request, "A fulfilled requisition cannot be cancelled.")
            elif requisition.lines.filter(fulfilled_qty__gt=0).exists():
                messages.error(request, "A partially fulfilled requisition cannot be cancelled.")
            else:
                requisition.status = "Cancelled"
                requisition.save(update_fields=["status", "updated_at"])
                messages.success(request, "Requisition cancelled.")
            return redirect("part_requisition_detail", requisition_id=requisition.id)

        if action == "return":
            employee = Employee.objects.filter(user=request.user).first()
            if not (request.user.is_superuser or is_parts_manager(employee)):
                messages.error(request, "Only a Parts Manager can return requisition parts.")
                return redirect("part_requisition_detail", requisition_id=requisition.id)
            if requisition.status == "Cancelled":
                messages.error(request, "Cancelled requisitions cannot be returned.")
                return redirect("part_requisition_detail", requisition_id=requisition.id)
            return_values = {}
            for line in requisition.lines.all():
                try:
                    return_qty = Decimal(request.POST.get(f"return_{line.id}") or "0")
                except (InvalidOperation, ValueError):
                    return_qty = Decimal("0")
                if return_qty > 0:
                    return_values[line.id] = return_qty
            if not return_values:
                messages.error(request, "Enter at least one quantity to return.")
                return redirect("part_requisition_detail", requisition_id=requisition.id)
            try:
                with transaction.atomic():
                    lines = PartRequisitionLine.objects.select_for_update().filter(
                        id__in=return_values, requisition=requisition
                    ).select_related("part")
                    for line in lines:
                        return_qty = return_values[line.id]
                        if return_qty > line.fulfilled_qty:
                            raise ValueError(f"{line.part.item_code}: return exceeds fulfilled quantity.")
                        part = ItemData.objects.select_for_update().get(id=line.part_id)
                        new_balance = part.current_stock + return_qty
                        part.current_stock = new_balance
                        part.save(update_fields=["current_stock", "updated_at"])
                        PartStockTransaction.objects.create(
                            part=part, transaction_type="Return", quantity_change=return_qty,
                            balance_after=new_balance, reference=requisition.requisition_no,
                            remarks=(request.POST.get("return_remarks") or request.POST.get("fulfillment_remarks") or "").strip()
                            or f"Returned against job {requisition.job.job_no}",
                            created_by=request.user,
                        )
                        line.fulfilled_qty -= return_qty
                        line.save(update_fields=["fulfilled_qty"])
                    sync_part_requisition_status(requisition)
            except ValueError as error:
                messages.error(request, str(error))
            else:
                messages.success(request, f"Parts returned to stock. Status: {requisition.status}.")
                create_user_notification(
                    requisition.requested_by,
                    "Parts Returned to Stock",
                    f"Parts were returned for {requisition.requisition_no}. Status: {requisition.status}.",
                    reverse("part_requisition_detail", args=[requisition.id]),
                )
            return redirect("part_requisition_detail", requisition_id=requisition.id)

        employee = Employee.objects.filter(user=request.user).first()
        if not request.user.is_superuser and not is_parts_manager(employee):
            messages.error(request, "Only a Parts Manager can fulfil requisitions.")
            return redirect("part_requisition_detail", requisition_id=requisition.id)
        if requisition.status in ["Cancelled", "Fulfilled"]:
            messages.error(request, "This requisition is not open for fulfilment.")
            return redirect("part_requisition_detail", requisition_id=requisition.id)

        issue_values = {}
        for line in requisition.lines.all():
            raw_value = request.POST.get(f"issue_{line.id}") or "0"
            try:
                issue_qty = Decimal(raw_value)
            except (InvalidOperation, ValueError):
                issue_qty = Decimal("0")
            if issue_qty > 0:
                issue_values[line.id] = issue_qty

        if not issue_values:
            messages.error(request, "Enter at least one quantity to issue.")
            return redirect("part_requisition_detail", requisition_id=requisition.id)

        try:
            with transaction.atomic():
                for line in PartRequisitionLine.objects.select_for_update().filter(
                    id__in=issue_values,
                    requisition=requisition,
                ).select_related("part"):
                    issue_qty = issue_values[line.id]
                    pending_qty = line.requested_qty - line.fulfilled_qty
                    if issue_qty > pending_qty:
                        raise ValueError(
                            f"{line.part.item_code}: issue quantity exceeds pending quantity."
                        )
                    part = ItemData.objects.select_for_update().get(id=line.part_id)
                    if issue_qty > part.current_stock:
                        raise ValueError(
                            f"{part.item_code}: only {part.current_stock} {part.unit} available."
                        )
                    new_balance = part.current_stock - issue_qty
                    part.current_stock = new_balance
                    part.save(update_fields=["current_stock", "updated_at"])
                    stock_tx = PartStockTransaction.objects.create(
                        part=part,
                        transaction_type="Issue",
                        quantity_change=-issue_qty,
                        balance_after=new_balance,
                        reference=requisition.requisition_no,
                        remarks=f"Issued against job {requisition.job.job_no}",
                        created_by=request.user,
                    )
                    line.fulfilled_qty += issue_qty
                    line.save(update_fields=["fulfilled_qty"])
                    PartRequisitionFulfillment.objects.create(
                        line=line,
                        quantity=issue_qty,
                        stock_transaction=stock_tx,
                        issued_by=request.user,
                        remarks=(request.POST.get("fulfillment_remarks") or "").strip(),
                    )
                sync_part_requisition_status(requisition)
        except ValueError as error:
            messages.error(request, str(error))
        else:
            messages.success(
                request,
                f"Parts issued successfully. Status: {requisition.status}.",
            )
            create_user_notification(
                requisition.requested_by,
                "Requisition Fulfilled",
                f"{requisition.requisition_no} was issued by Parts. Status: {requisition.status}.",
                reverse("part_requisition_detail", args=[requisition.id]),
            )
        return redirect("part_requisition_detail", requisition_id=requisition.id)

    logged_employee = Employee.objects.filter(user=request.user).first()
    advisor_view = is_advisor_employee(logged_employee)
    return render(request, "parts/requisitionDetail.html", {
        "requisition": requisition,
        "lines": requisition.lines.all(),
        "can_fulfill": (
            request.user.is_superuser
            or is_parts_manager(logged_employee)
        ),
        "back_url": (
            reverse("jobcard_edit", args=[requisition.job_id])
            + "#part-requisitions"
            if advisor_view
            else reverse("part_requisition_list")
        ),
        "back_label": "Back to Job Card" if advisor_view else "Back to Requisitions",
    })


@never_cache
@login_required
def part_requisition_print(request, requisition_id):
    requisition = get_object_or_404(
        PartRequisition.objects.select_related(
            "job",
            "job__claim",
            "job__claim__vehicle",
            "job__claim__vehicle__customer",
            "job__claim__vehicle__model",
            "requested_by",
        ).prefetch_related("lines__part"),
        id=requisition_id,
    )
    vehicle = (
        requisition.job.claim.vehicle
        if requisition.job.claim_id else None
    )
    return render(request, "parts/requisitionPrint.html", {
        "requisition": requisition,
        "job": requisition.job,
        "vehicle": vehicle,
        "lines": requisition.lines.all(),
    })


def sync_part_orders(job, header=None, part_ids=None):
    created_count = 0

    if header is None:
        header = PartOrderHeader.objects.create(
            job=job,
            vehicle=job.claim.vehicle if job.claim and job.claim.vehicle else None,
            order_no=job.part_order_no or "",
            order_date=job.part_order_date,
            status="Pending",
        )

    parts = job.parts.all()

    if part_ids:
        parts = parts.filter(id__in=part_ids)

    for part in parts:
        order, created = PartOrder.objects.get_or_create(
            order=header,
            job=job,
            part=part,
            defaults={
                "ordered_qty": part.qty,
                "status": "Pending",
            }
        )

        if order.order_id is None:
            order.order = header
            order.save(update_fields=["order"])

        if created:
            created_count += 1

    return created_count


def get_part_order_lines_status(orders, completed_status="Received"):
    orders = list(orders)

    if not orders:
        return "Pending"

    statuses = {order.status for order in orders}

    if statuses == {"Received"}:
        return completed_status

    if "Back Order" in statuses:
        return "Back Order"

    if "Received" in statuses:
        return "Partially Received"

    if "In Transit" in statuses:
        return "In Transit"

    if "Order Placed" in statuses:
        return "Order Placed"

    if "Cancelled" in statuses and len(statuses) == 1:
        return "Cancelled"

    return "Pending"


def sync_part_order_header(header):
    if not header:
        return None

    header.status = get_part_order_lines_status(
        header.lines.all(),
        completed_status="Received"
    )
    header.save(update_fields=["status", "updated_at"])
    return header


@never_cache
@login_required
def parts_manager_dashboard(request):
    employee = Employee.objects.filter(user=request.user).first()
    if not request.user.is_superuser and not is_parts_manager(employee):
        return redirect("dashboard")

    today = timezone.localdate()
    month_start = today.replace(day=1)
    headers = PartOrderHeader.objects.select_related(
        "job",
        "job__claim",
        "job__claim__vehicle",
        "job__claim__vehicle__customer",
        "vehicle",
        "vehicle__customer",
    ).prefetch_related("lines")

    if employee and employee.branch_id:
        headers = headers.filter(
            Q(job__branch_id=employee.branch_id)
            | Q(job__claim__branch_id=employee.branch_id)
            | Q(job__isnull=True)
        ).distinct()

    open_headers = headers.exclude(status__in=["Received", "Cancelled"])
    overdue_headers = open_headers.filter(expected_date__lt=today)
    search = (request.GET.get("search") or "").strip()
    recent_headers = headers
    if search:
        recent_headers = recent_headers.filter(
            Q(order_no__icontains=search)
            | Q(supplier__icontains=search)
            | Q(job__job_no__icontains=search)
            | Q(job__claim__vehicle__registration_no__icontains=search)
            | Q(vehicle__registration_no__icontains=search)
        ).distinct()

    def dashboard_row(header):
        job = header.job
        claim = job.claim if job else None
        vehicle = claim.vehicle if claim and claim.vehicle else header.vehicle
        lines = list(header.lines.all())
        ordered_qty = sum((line.ordered_qty or Decimal("0")) for line in lines)
        received_qty = sum((line.received_qty or Decimal("0")) for line in lines)
        progress = int((received_qty / ordered_qty) * 100) if ordered_qty else 0
        return {
            "header": header,
            "job": job,
            "vehicle": vehicle,
            "line_count": len(lines),
            "ordered_qty": ordered_qty,
            "received_qty": received_qty,
            "progress": min(progress, 100),
            "is_overdue": bool(
                header.expected_date
                and header.expected_date < today
                and header.status not in ["Received", "Cancelled"]
            ),
        }

    return render(request, "parts/partsManagerDashboard.html", {
        "employee": employee,
        "today": today,
        "search": search,
        "open_count": open_headers.count(),
        "overdue_count": overdue_headers.count(),
        "back_order_count": headers.filter(status="Back Order").count(),
        "received_month_count": headers.filter(
            status="Received",
            updated_at__date__gte=month_start,
            updated_at__date__lte=today,
        ).count(),
        "requisition_open_count": PartRequisition.objects.filter(
            status__in=["Submitted", "Partially Fulfilled"]
        ).count(),
        "overdue_rows": [
            dashboard_row(header)
            for header in overdue_headers.order_by("expected_date")[:6]
        ],
        "recent_rows": [
            dashboard_row(header)
            for header in recent_headers.order_by("-updated_at")[:10]
        ],
    })


@never_cache
@login_required
def part_order_list(request):
    q = request.GET.get("q", "").strip()
    status = request.GET.get("status", "").strip()
    supplier = request.GET.get("supplier", "").strip()
    date_from = parse_date(request.GET.get("date_from", ""))
    date_to = parse_date(request.GET.get("date_to", ""))

    headers = PartOrderHeader.objects.select_related(
        "job",
        "job__claim",
        "job__claim__vehicle",
        "job__claim__vehicle__customer",
        "job__claim__vehicle__model",
        "vehicle",
        "vehicle__customer",
        "vehicle__model",
    ).prefetch_related("lines")

    if q:
        headers = headers.filter(
            Q(order_no__icontains=q)
            | Q(supplier__icontains=q)
            | Q(job__job_no__icontains=q)
            | Q(job__claim__claim_no__icontains=q)
            | Q(job__claim__vehicle__registration_no__icontains=q)
            | Q(job__claim__vehicle__customer__name__icontains=q)
            | Q(vehicle__registration_no__icontains=q)
            | Q(vehicle__customer__name__icontains=q)
            | Q(lines__manual_part_no__icontains=q)
            | Q(lines__manual_description__icontains=q)
        )

    if status:
        headers = headers.filter(status=status)
    if supplier:
        headers = headers.filter(supplier__icontains=supplier)
    if date_from:
        headers = headers.filter(order_date__gte=date_from)
    if date_to:
        headers = headers.filter(order_date__lte=date_to)

    headers = headers.distinct().order_by("-order_date", "-updated_at")
    all_headers = PartOrderHeader.objects.all()
    today = timezone.localdate()
    summary = {
        "total": all_headers.count(),
        "open": all_headers.exclude(status__in=["Received", "Cancelled"]).count(),
        "transit": all_headers.filter(status="In Transit").count(),
        "overdue": all_headers.exclude(
            status__in=["Received", "Cancelled"]
        ).filter(expected_date__lt=today).count(),
        "received": all_headers.filter(status="Received").count(),
    }

    rows = []
    for header in headers:
        job = header.job
        claim = job.claim if job else None
        vehicle = claim.vehicle if claim and claim.vehicle else header.vehicle
        lines = list(header.lines.all())
        ordered_qty = sum((line.ordered_qty or Decimal("0")) for line in lines)
        received_qty = sum((line.received_qty or Decimal("0")) for line in lines)
        progress = int((received_qty / ordered_qty) * 100) if ordered_qty else 0
        rows.append({
            "header": header,
            "job": job,
            "claim": claim,
            "vehicle": vehicle,
            "line_count": len(lines),
            "ordered_qty": ordered_qty,
            "received_qty": received_qty,
            "progress": min(progress, 100),
            "is_overdue": bool(
                header.expected_date
                and header.expected_date < today
                and header.status not in ["Received", "Cancelled"]
            ),
        })

    if request.GET.get("export") == "csv":
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = 'attachment; filename="parts-order-tracking.csv"'
        writer = csv.writer(response)
        writer.writerow([
            "Order No", "Order Date", "Expected Date", "Status", "Supplier",
            "Job No", "Claim No", "Registration No", "Customer",
            "Lines", "Ordered Qty", "Received Qty", "Progress %", "Overdue",
        ])
        for row in rows:
            writer.writerow([
                row["header"].order_no or f"Order #{row['header'].id}",
                row["header"].order_date or "",
                row["header"].expected_date or "",
                row["header"].status,
                row["header"].supplier,
                row["job"].job_no if row["job"] else "",
                row["claim"].claim_no if row["claim"] else "",
                row["vehicle"].registration_no if row["vehicle"] else "",
                row["vehicle"].customer.name if row["vehicle"] and row["vehicle"].customer else "",
                row["line_count"],
                row["ordered_qty"],
                row["received_qty"],
                row["progress"],
                "Yes" if row["is_overdue"] else "No",
            ])
        return response

    return render(request, "parts/partOrderList.html", {
        "rows": rows,
        "statuses": PartOrder.STATUS_CHOICES,
        "selected_status": status,
        "q": q,
        "supplier": supplier,
        "date_from": date_from,
        "date_to": date_to,
        "summary": summary,
    })


@never_cache
@login_required
def part_order_create(request):
    jobs = JobCard.objects.select_related(
        "claim",
        "claim__vehicle",
        "claim__vehicle__customer",
    ).order_by("-id")
    vehicles = Vehicle.objects.select_related(
        "customer",
        "model",
    ).order_by("registration_no")

    if request.method == "POST":
        job_id = request.POST.get("job_id") or ""
        vehicle_id = request.POST.get("vehicle_id") or ""
        order_no = request.POST.get("order_no", "").strip()
        order_date = request.POST.get("order_date") or None
        expected_date = request.POST.get("expected_date") or None
        supplier = request.POST.get("supplier", "").strip()
        remarks = request.POST.get("remarks", "").strip()
        part_no_list = request.POST.getlist("part_no[]")
        part_desc_list = request.POST.getlist("part_desc[]")
        qty_list = request.POST.getlist("qty[]")

        job = JobCard.objects.filter(id=job_id).select_related(
            "claim",
            "claim__vehicle",
        ).first() if job_id else None
        vehicle = None

        if job and job.claim:
            vehicle = job.claim.vehicle
        elif vehicle_id:
            vehicle = Vehicle.objects.filter(id=vehicle_id).first()

        valid_lines = []

        for index, part_no in enumerate(part_no_list):
            part_no = part_no.strip()
            description = (
                part_desc_list[index].strip()
                if index < len(part_desc_list)
                else ""
            )
            qty = (
                qty_list[index]
                if index < len(qty_list)
                else "0"
            )

            if not part_no and not description:
                continue

            valid_lines.append({
                "part_no": part_no,
                "description": description,
                "qty": Decimal(qty or "0"),
            })

        if not job and not vehicle:
            messages.error(
                request,
                "Select either Job Card or Vehicle Registration No."
            )
        elif not valid_lines:
            messages.error(
                request,
                "Enter at least one part line."
            )
        else:
            with transaction.atomic():
                header = PartOrderHeader.objects.create(
                    job=job,
                    vehicle=vehicle,
                    order_no=order_no,
                    order_date=order_date,
                    expected_date=expected_date,
                    supplier=supplier,
                    status="Pending",
                    remarks=remarks,
                )
                if not header.order_no:
                    header.order_no = f"PO-{timezone.localdate():%Y%m%d}-{header.id:04d}"
                    header.save(update_fields=["order_no", "updated_at"])

                for line in valid_lines:
                    PartOrder.objects.create(
                        order=header,
                        job=job,
                        part=None,
                        manual_part_no=line["part_no"],
                        manual_description=line["description"],
                        order_no=header.order_no,
                        supplier=header.supplier,
                        order_date=header.order_date,
                        expected_date=header.expected_date,
                        ordered_qty=line["qty"],
                        status="Pending",
                    )

                sync_part_order_header(header)

            messages.success(request, "Part order saved successfully.")
            if request.POST.get("print_after_save") == "1":
                return redirect("part_order_print", header_id=header.id)
            return redirect("part_order_list")

    return render(request, "parts/partOrderCreate.html", {
        "jobs": jobs,
        "vehicles": vehicles,
        "today": timezone.localdate(),
        "prefill": {
            "job_id": request.GET.get("job_id") or "",
            "part_no": request.GET.get("part_no") or "",
            "part_desc": request.GET.get("part_desc") or "",
            "qty": request.GET.get("qty") or "1",
            "supplier": request.GET.get("supplier") or "",
        },
    })


@never_cache
@login_required
def part_order_print(request, header_id):
    header = get_object_or_404(
        PartOrderHeader.objects.select_related(
            "job",
            "job__claim",
            "job__claim__vehicle",
            "job__claim__vehicle__customer",
            "job__claim__vehicle__model",
            "vehicle",
            "vehicle__customer",
            "vehicle__model",
        ),
        id=header_id
    )
    job = header.job
    claim = job.claim if job else None
    vehicle = claim.vehicle if claim and claim.vehicle else header.vehicle
    lines = list(header.lines.select_related("part").order_by("id"))

    return render(request, "parts/partOrderPrint.html", {
        "header": header,
        "job": job,
        "job_created_date_value": datetime_local_value(job.job_date),
        "claim": claim,
        "vehicle": vehicle,
        "lines": lines,
    })


@never_cache
@login_required
def part_order_header_edit(request, header_id):
    header = get_object_or_404(
        PartOrderHeader.objects.select_related(
            "job",
            "job__claim",
            "job__claim__vehicle",
            "job__claim__vehicle__customer",
            "vehicle",
            "vehicle__customer",
        ).prefetch_related("lines__part"),
        id=header_id,
    )

    if request.method == "POST":
        order_no = (request.POST.get("order_no") or "").strip()
        supplier = (request.POST.get("supplier") or "").strip()
        order_date_value = request.POST.get("order_date") or ""
        expected_date_value = request.POST.get("expected_date") or ""
        remarks = (request.POST.get("remarks") or "").strip()
        order_date = parse_date(order_date_value) if order_date_value else None
        expected_date = (
            parse_date(expected_date_value) if expected_date_value else None
        )

        if order_date_value and order_date is None:
            messages.error(request, "Enter a valid order date.")
        elif expected_date_value and expected_date is None:
            messages.error(request, "Enter a valid expected delivery date.")
        else:
            if not order_no:
                order_no = f"PO-{timezone.localdate():%Y%m%d}-{header.id:04d}"

            with transaction.atomic():
                header.order_no = order_no
                header.supplier = supplier
                header.order_date = order_date
                header.expected_date = expected_date
                header.remarks = remarks
                header.save(update_fields=[
                    "order_no",
                    "supplier",
                    "order_date",
                    "expected_date",
                    "remarks",
                    "updated_at",
                ])
                header.lines.update(
                    order_no=order_no,
                    supplier=supplier,
                    order_date=order_date,
                    expected_date=expected_date,
                )

            messages.success(
                request,
                f"Order header {order_no} updated successfully.",
            )
            if request.POST.get("save_and_track") == "1" and header.job_id:
                return redirect(
                    reverse(
                        "part_order_job_detail",
                        args=[header.job_id],
                    ) + f"?order_id={header.id}"
                )
            return redirect("part_order_list")

    job = header.job
    claim = job.claim if job else None
    vehicle = claim.vehicle if claim and claim.vehicle else header.vehicle
    lines = list(header.lines.all().order_by("id"))

    return render(request, "parts/partOrderHeaderEdit.html", {
        "header": header,
        "job": job,
        "vehicle": vehicle,
        "lines": lines,
    })


@never_cache
@require_POST
@login_required
def create_part_order_from_job(request, job_id):
    job = get_object_or_404(JobCard, id=job_id)
    data = {}

    if request.body:
        data = json.loads(request.body.decode("utf-8"))

    header = PartOrderHeader.objects.create(
        job=job,
        vehicle=job.claim.vehicle if job.claim and job.claim.vehicle else None,
        order_no=data.get("order_no") or "",
        order_date=data.get("order_date") or None,
        expected_date=data.get("expected_date") or None,
        supplier=data.get("supplier") or "",
        status="Pending",
    )
    part_ids = data.get("part_ids") or None
    created_count = sync_part_orders(
        job,
        header=header,
        part_ids=part_ids
    )

    sync_part_order_header(header)

    return JsonResponse({
        "status": "success",
        "order_id": header.id,
        "created_count": created_count,
        "redirect_url": reverse(
            "part_order_job_detail",
            args=[job.id]
        ) + f"?order_id={header.id}"
    })

@never_cache
@login_required
def part_order_job_detail(request, job_id):
    job = get_object_or_404(
        JobCard.objects.select_related(
            "claim",
            "claim__vehicle",
            "claim__vehicle__customer",
            "claim__vehicle__model",
            "advisor",
        ),
        id=job_id
    )

    selected_order_id = request.GET.get("order_id")

    if request.method == "POST":
        order_ids = request.POST.getlist("order_id[]")
        touched_headers = set()

        for order_id in order_ids:
            order = PartOrder.objects.filter(
                id=order_id,
                job=job
            ).first()

            if not order:
                continue

            prefix = f"order_{order_id}_"
            order.order_no = request.POST.get(prefix + "order_no", "").strip()
            order.supplier = request.POST.get(prefix + "supplier", "").strip()
            order.order_date = request.POST.get(prefix + "order_date") or None
            order.expected_date = request.POST.get(prefix + "expected_date") or None
            order.received_date = request.POST.get(prefix + "received_date") or None
            order.ordered_qty = Decimal(request.POST.get(prefix + "ordered_qty") or "0")
            order.received_qty = Decimal(request.POST.get(prefix + "received_qty") or "0")
            order.status = request.POST.get(prefix + "status") or "Pending"
            order.tracking_ref = request.POST.get(prefix + "tracking_ref", "").strip()
            order.remarks = request.POST.get(prefix + "remarks", "").strip()

            if (
                order.ordered_qty
                and order.received_qty >= order.ordered_qty
                and order.status not in ["Cancelled", "Back Order"]
            ):
                order.status = "Received"

            order.save()
            if order.order_id:
                touched_headers.add(order.order_id)

        for header in PartOrderHeader.objects.filter(id__in=touched_headers):
            sync_part_order_header(header)
            notify_jobcard_advisor(
                job,
                "Part Order Updated",
                f"Part order {header.order_no or header.id} for job {job.job_no} is now {header.status}.",
            )

        if request.POST.get("print_after_save") == "1":
            print_header_id = selected_order_id or (
                next(iter(touched_headers)) if touched_headers else ""
            )

            if print_header_id:
                return redirect(
                    "part_order_print",
                    header_id=print_header_id
                )

        redirect_url = reverse("part_order_job_detail", args=[job.id])
        if selected_order_id:
            redirect_url += f"?order_id={selected_order_id}"
        return redirect(redirect_url)

    headers = job.part_order_headers.order_by("-id")
    selected_header = None

    if selected_order_id:
        selected_header = headers.filter(id=selected_order_id).first()

    if selected_header is None:
        selected_header = headers.first()

    orders = PartOrder.objects.filter(job=job).select_related("part")

    if selected_header:
        orders = orders.filter(order=selected_header)
    else:
        orders = orders.none()

    orders = orders.order_by("part__id")

    return render(request, "parts/partOrderDetail.html", {
        "job": job,
        "headers": headers,
        "selected_header": selected_header,
        "orders": orders,
        "job_part_order_status": get_job_part_order_status(job),
        "statuses": PartOrder.STATUS_CHOICES,
    })


def get_job_part_order_status(job):
    return get_part_order_lines_status(
        job.part_orders.all(),
        completed_status="Completed"
    )


@login_required
@never_cache
def part_order_jobs_api(request):
    jobs = JobCard.objects.select_related(
        "claim",
        "claim__vehicle",
        "claim__vehicle__customer",
        "claim__vehicle__model",
        "advisor",
    ).prefetch_related("parts", "part_orders").order_by("-id")

    data = []

    for job in jobs:
        orders = list(job.part_orders.all())
        header = job.part_order_headers.order_by("-id").first()
        total_parts = job.parts.count()
        ordered_parts = len(orders)
        received_parts = sum(1 for order in orders if order.status == "Received")
        back_order_parts = sum(1 for order in orders if order.status == "Back Order")

        data.append({
            "id": job.id,
            "job_no": job.job_no,
            "claim_no": job.claim.claim_no if job.claim else "",
            "reg_no": job.claim.vehicle.registration_no if job.claim and job.claim.vehicle else "",
            "model": job.claim.vehicle.model.name if job.claim and job.claim.vehicle and job.claim.vehicle.model else "",
            "customer": job.claim.vehicle.customer.name if job.claim and job.claim.vehicle and job.claim.vehicle.customer else "",
            "advisor": job.advisor.name if job.advisor else "",
            "job_date": job.job_date.strftime("%d-%m-%Y") if job.job_date else "",
            "order_status": get_job_part_order_status(job),
            "order_no": header.order_no if header else "",
            "order_date": header.order_date.strftime("%d-%m-%Y") if header and header.order_date else "",
            "order_count": job.part_order_headers.count(),
            "total_parts": total_parts,
            "ordered_parts": ordered_parts,
            "received_parts": received_parts,
            "back_order_parts": back_order_parts,
            "has_order": ordered_parts > 0,
        })

    return JsonResponse(data, safe=False)


@login_required
@never_cache
def part_order_headers_api(request):
    headers = PartOrderHeader.objects.select_related(
        "job",
        "job__claim",
        "job__claim__vehicle",
        "job__claim__vehicle__customer",
        "job__claim__vehicle__model",
        "vehicle",
        "vehicle__customer",
        "vehicle__model",
    ).prefetch_related("lines").order_by("-id")

    data = []

    for header in headers:
        job = header.job
        claim = job.claim if job else None
        vehicle = claim.vehicle if claim and claim.vehicle else header.vehicle
        line_count = header.lines.count()
        received_count = header.lines.filter(status="Received").count()
        back_order_count = header.lines.filter(status="Back Order").count()

        data.append({
            "id": header.id,
            "job_id": job.id if job else "",
            "order_no": header.order_no or f"Order #{header.id}",
            "order_date": header.order_date.strftime("%d-%m-%Y") if header.order_date else "",
            "expected_date": header.expected_date.strftime("%d-%m-%Y") if header.expected_date else "",
            "supplier": header.supplier,
            "status": header.status,
            "line_count": line_count,
            "received_count": received_count,
            "back_order_count": back_order_count,
            "job_no": job.job_no if job else "",
            "claim_no": claim.claim_no if claim else "",
            "reg_no": vehicle.registration_no if vehicle else "",
            "model": vehicle.model.name if vehicle and vehicle.model else "",
            "customer": vehicle.customer.name if vehicle and vehicle.customer else "",
        })

    return JsonResponse(data, safe=False)


@login_required
@never_cache
def part_order_job_headers_api(request, job_id):
    job = get_object_or_404(JobCard, id=job_id)
    headers = job.part_order_headers.order_by("-id")

    data = []

    for header in headers:
        line_count = header.lines.count()
        data.append({
            "id": header.id,
            "order_no": header.order_no or f"Order #{header.id}",
            "order_date": header.order_date.strftime("%Y-%m-%d") if header.order_date else "",
            "expected_date": header.expected_date.strftime("%Y-%m-%d") if header.expected_date else "",
            "supplier": header.supplier,
            "status": header.status,
            "line_count": line_count,
        })

    return JsonResponse(data, safe=False)


@login_required
@never_cache
def part_order_job_orders_api(request, job_id):
    job = get_object_or_404(JobCard, id=job_id)
    order_id = request.GET.get("order_id")
    orders = PartOrder.objects.filter(
        job=job
    ).select_related("part").order_by("part__id")

    if order_id:
        orders = orders.filter(order_id=order_id)
    else:
        header = job.part_order_headers.order_by("-id").first()
        orders = orders.filter(order=header) if header else orders.none()

    data = []

    for order in orders:
        part_no = order.part.part_no if order.part else order.manual_part_no
        description = order.part.description if order.part else order.manual_description
        estimated_qty = order.part.qty if order.part else order.ordered_qty

        data.append({
            "id": order.id,
            "part_no": part_no,
            "description": description,
            "estimated_qty": str(estimated_qty),
            "ordered_qty": str(order.ordered_qty),
            "received_qty": str(order.received_qty),
            "status": order.status,
            "order_no": order.order_no,
            "supplier": order.supplier,
            "order_date": order.order_date.strftime("%Y-%m-%d") if order.order_date else "",
            "expected_date": order.expected_date.strftime("%Y-%m-%d") if order.expected_date else "",
            "received_date": order.received_date.strftime("%Y-%m-%d") if order.received_date else "",
            "tracking_ref": order.tracking_ref,
            "remarks": order.remarks,
        })

    return JsonResponse(data, safe=False)


@login_required
@never_cache
def part_order_header_lines_api(request, header_id):
    header = get_object_or_404(PartOrderHeader, id=header_id)
    orders = header.lines.select_related("part").order_by("part__id")

    data = []

    for order in orders:
        part_no = order.part.part_no if order.part else order.manual_part_no
        description = order.part.description if order.part else order.manual_description
        estimated_qty = order.part.qty if order.part else order.ordered_qty

        data.append({
            "id": order.id,
            "part_no": part_no,
            "description": description,
            "estimated_qty": str(estimated_qty),
            "ordered_qty": str(order.ordered_qty),
            "received_qty": str(order.received_qty),
            "status": order.status,
            "order_no": order.order_no,
            "supplier": order.supplier,
            "order_date": order.order_date.strftime("%Y-%m-%d") if order.order_date else "",
            "expected_date": order.expected_date.strftime("%Y-%m-%d") if order.expected_date else "",
            "received_date": order.received_date.strftime("%Y-%m-%d") if order.received_date else "",
            "tracking_ref": order.tracking_ref,
            "remarks": order.remarks,
        })

    return JsonResponse(data, safe=False)


@require_POST
@login_required
@never_cache
def update_part_order_line_api(request, order_id):
    order = get_object_or_404(PartOrder, id=order_id)
    try:
        data = json.loads(request.body.decode("utf-8") or "{}")
    except (json.JSONDecodeError, UnicodeDecodeError):
        return JsonResponse({
            "status": "error",
            "message": "Invalid request data.",
        }, status=400)

    for field in [
        "order_no",
        "supplier",
        "status",
        "tracking_ref",
        "remarks",
    ]:
        if field in data:
            setattr(order, field, data.get(field) or "")

    for field in ["order_date", "expected_date", "received_date"]:
        if field in data:
            setattr(order, field, data.get(field) or None)

    if "ordered_qty" in data:
        order.ordered_qty = Decimal(str(data.get("ordered_qty") or "0"))

    if "received_qty" in data:
        order.received_qty = Decimal(str(data.get("received_qty") or "0"))

    if (
        order.ordered_qty
        and order.received_qty >= order.ordered_qty
        and order.status not in ["Cancelled", "Back Order"]
    ):
        order.status = "Received"

    order.save()
    sync_part_order_header(order.order)
    job_status = get_job_part_order_status(order.job) if order.job else ""

    return JsonResponse({
        "status": "success",
        "job_status": job_status,
        "line_status": order.status,
    })


@require_POST
@login_required
@never_cache
def delete_part_order_line_api(request, order_id):
    order = get_object_or_404(PartOrder, id=order_id)
    job = order.job
    header = order.order
    order.delete()

    if header:
        sync_part_order_header(header)

    return JsonResponse({
        "status": "success",
        "job_status": get_job_part_order_status(job) if job else "",
    })


from django.http import JsonResponse, HttpResponse
from .models import ItemData


@login_required
def part_lookup(request):
    part_no = request.GET.get("item_code", "").strip()

    item = ItemData.objects.filter(item_code__iexact=part_no).first()

    if not item:
        return JsonResponse({
            "status": "error",
            "message": "Part not found"
        })
    if item.status != "Active":
        return JsonResponse({
            "status": "error",
            "message": "Part exists but is inactive. Activate it in Part Master.",
            "inactive": True,
        })

    pending_lines = PartRequisitionLine.objects.filter(
        part=item,
        requisition__status__in=["Submitted", "Partially Fulfilled"],
    )
    pending_requisition_qty = sum(
        (
            max(line.requested_qty - line.fulfilled_qty, Decimal("0"))
            for line in pending_lines
        ),
        Decimal("0"),
    )
    open_order_lines = PartOrder.objects.filter(
        Q(manual_part_no__iexact=item.item_code)
        | Q(part__part_no__iexact=item.item_code)
    ).exclude(status__in=["Received", "Cancelled"])
    on_order_qty = sum(
        (
            max(order.ordered_qty - order.received_qty, Decimal("0"))
            for order in open_order_lines
        ),
        Decimal("0"),
    )
    available_after_pending = item.current_stock - pending_requisition_qty

    return JsonResponse({
        "status": "success",
        "id": item.id,
        "description": item.item_name,
        "rate": str(item.rate),
        "category": item.category or "",
        "unit": item.unit,
        "hsn_code": item.hsn_code,
        "gst_percent": str(item.gst_percent),
        "current_stock": str(item.current_stock),
        "reorder_level": str(item.reorder_level),
        "preferred_supplier": item.preferred_supplier,
        "bin_location": item.bin_location,
        "pending_requisition_qty": str(pending_requisition_qty),
        "on_order_qty": str(on_order_qty),
        "available_after_pending": str(available_after_pending),
        "is_out_of_stock": item.current_stock <= 0,
        "has_pending_demand": pending_requisition_qty > 0,
        "has_open_order": on_order_qty > 0,
    })


from django.http import HttpResponseForbidden
from django.shortcuts import render
from .models import JobCard, JobCardInventory


import json
from decimal import Decimal, InvalidOperation


def save_job_inventory(request, job):
    damage_marks_raw = request.POST.get("damage_marks", "[]")

    try:
        damage_marks = json.loads(damage_marks_raw)
    except json.JSONDecodeError:
        damage_marks = []

    JobCardInventory.objects.update_or_create(
        job=job,
        defaults={
            "lh_mirror": int(request.POST.get("lh_mirror") == "on"),
            "mud_flap_count": int(request.POST.get("mud_flap_count") or 0),
            "floor_mat_count": int(request.POST.get("floor_mat_count") or 0),
            "rh_mirror": int(request.POST.get("rh_mirror") == "on"),
            "center_mirror": int(request.POST.get("center_mirror") == "on"),
            "frt_wiper": int(request.POST.get("frt_wiper") == "on"),
            "rr_wiper": int(request.POST.get("rr_wiper") == "on"),
            "accessories": int(request.POST.get("accessories") == "on"),
            "spare_wheel": request.POST.get("spare_wheel") == "on",
            "jack": request.POST.get("jack") == "on",
            "tool_kit": request.POST.get("tool_kit") == "on",
            "stereo": request.POST.get("stereo") == "on",
            "battery": request.POST.get("battery") == "on",
            "number_plate": request.POST.get("number_plate") == "on",

            "fuel_percent": int(request.POST.get("fuel_percent") or 0),
            "cng_percent": int(request.POST.get("cng_percent") or 0),
            "damage_marks": damage_marks,

            "remarks": request.POST.get("inventory_remarks", ""),
        }
    )

    save_tyre_inventory(request, job)


def save_tyre_inventory(request, job):
    positions = request.POST.getlist("tyre_position[]")
    makes = request.POST.getlist("tyre_make[]")
    sizes = request.POST.getlist("tyre_size[]")
    depths = request.POST.getlist("tyre_depth[]")
    wheel_cap = request.POST.getlist("tyre_wheel_cap[]")

    for i in range(len(positions)):

        depth_value = None

        if depths[i]:
            try:
                depth_value = Decimal(depths[i])
            except InvalidOperation:
                depth_value = None
        print(wheel_cap[i])
        JobCardTyreInventory.objects.update_or_create(
            job=job,
            position=positions[i],
            defaults={
                "make": makes[i],
                "size": sizes[i],
                "depth": depth_value,
                "wheel_cap": wheel_cap[i],
            }
        )


def link_callback(uri, rel):
    # MEDIA files
    if uri.startswith(settings.MEDIA_URL):
        path = os.path.join(settings.MEDIA_ROOT, uri.replace(settings.MEDIA_URL, ""))
        if os.path.exists(path):
            return path

    # STATIC files
    if uri.startswith(settings.STATIC_URL):
        path = os.path.join(settings.STATIC_ROOT, uri.replace(settings.STATIC_URL, ""))
        if os.path.exists(path):
            return path

    return uri


def generate_jobcard_pdf(job):
    template = get_template("jobcard/jobcardPrint.html")

    html = template.render({
        "job": job,
        "claim": job.claim,
        "parts": job.parts.all(),
        "labours": job.labours.all(),
        **get_inventory_context(job),
    })
    result = BytesIO()

    pisa_status = pisa.CreatePDF(
        html,
        dest=result,
        link_callback=link_callback
    )

    if pisa_status.err:
        return None

    return result.getvalue()


import requests
from django.core.files.base import ContentFile
from django.utils import timezone


def legacy_send_whatsapp_template_message(mobile, template_name="hello_world", language_code="en_US"):
    token = getattr(settings, "WHATSAPP_ACCESS_TOKEN", "") or getattr(settings, "WHATSAPP_TOKEN", "")
    phone_number_id = getattr(settings, "WHATSAPP_PHONE_NUMBER_ID", "")
    graph_version = getattr(settings, "WHATSAPP_GRAPH_VERSION", "v23.0")

    if not token or not phone_number_id:
        return {
            "success": False,
            "status_code": 400,
            "response": "WHATSAPP_ACCESS_TOKEN and WHATSAPP_PHONE_NUMBER_ID are required in .env.",
        }

    mobile = "".join(ch for ch in str(mobile or "") if ch.isdigit())
    if len(mobile) == 10:
        mobile = "91" + mobile
    if len(mobile) < 11:
        return {
            "success": False,
            "status_code": 400,
            "response": "Enter valid WhatsApp mobile number with country code.",
        }

    url = f"https://graph.facebook.com/{graph_version}/{phone_number_id}/messages"
    payload = {
        "messaging_product": "whatsapp",
        "to": mobile,
        "type": "template",
        "template": {
            "name": template_name,
            "language": {"code": language_code},
        },
    }
    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json",
    }

    try:
        response = requests.post(url, json=payload, headers=headers, timeout=30)
        return {
            "success": response.status_code in [200, 201],
            "status_code": response.status_code,
            "response": response.text,
        }
    except requests.RequestException as error:
        return {
            "success": False,
            "status_code": 500,
            "response": str(error),
        }


def jobcard_tracking_url(job):
    token = TimestampSigner().sign(str(job.id))
    return settings.SITE_URL.rstrip("/") + reverse("customer_jobcard_tracking", args=[token])


def send_jobcard_whatsapp(job):
    vehicle = job.vehicle or (job.claim.vehicle if job.claim_id and job.claim and job.claim.vehicle_id else None)
    customer = vehicle.customer if vehicle and vehicle.customer_id else None
    if not customer:
        return None
    mobile = customer.mobile_no or customer.whatsapp_no

    if not mobile:
        return None

    mobile = "91" + mobile[-10:]

    pdf_bytes = generate_jobcard_pdf(job)

    if not pdf_bytes:
        return None

    filename = get_next_jobcard_pdf_filename(job)

    log = CommunicationLog.objects.create(
        job=job,
        channel="WhatsApp",
        mobile_no=mobile,
        message=f"Job Card {job.job_no}",
        status="Pending"
    )

    log.pdf_file.save(filename, ContentFile(pdf_bytes), save=True)

    pdf_url = settings.SITE_URL.rstrip("/") + log.pdf_file.url

    tracking_url = jobcard_tracking_url(job)
    print("PDF URL:", pdf_url)

    url = (
        f"https://graph.facebook.com/v23.0/"
        f"{settings.WHATSAPP_PHONE_NUMBER_ID}/messages"
    )

    payload = {
        "messaging_product": "whatsapp",
        "to": mobile,
        "type": "document",
        "document": {
            "link": pdf_url,
            "filename": filename,
            "caption": (
                f"Dear {customer.name},\n"
                f"Your Job Card {job.job_no} has been created.\n"
                f"Vehicle: {vehicle.registration_no if vehicle else '-'}\n"
                f"Track your repair progress: {tracking_url}\n"
                f"Thank you."
            )
        }
    }

    headers = {
        "Authorization": f"Bearer {settings.WHATSAPP_TOKEN}",
        "Content-Type": "application/json"
    }

    try:
        response = requests.post(
            url,
            json=payload,
            headers=headers,
            timeout=30
        )

        print("WHATSAPP STATUS:", response.status_code)
        print("WHATSAPP RESPONSE:", response.text)

        log.response = response.text

        if response.status_code in [200, 201]:
            log.status = "Sent"
            log.sent_at = timezone.now()
        else:
            log.status = "Failed"

        log.save()
        return log

    except Exception as e:
        print("WHATSAPP ERROR:", str(e))

        log.status = "Failed"
        log.response = str(e)
        log.save()
        return log


from django.shortcuts import get_object_or_404
from django.shortcuts import redirect


from urllib.parse import urlencode, quote
from django.urls import reverse

from playwright.sync_api import sync_playwright
import time


def generate_jobcard_pdf(job):
    url = (
            settings.SITE_URL
            + reverse("jobcard_print_preview", args=[job.id, settings.PDF_SECRET_TOKEN])

            + "?"
            + urlencode({"v": int(time.time())})
    )

    with sync_playwright() as p:
        browser = p.chromium.launch(
            headless=True
        )

        context = browser.new_context(
            ignore_https_errors=True
        )

        page = context.new_page()

        response = page.goto(
            url,
            wait_until="networkidle"
        )
        if not response or response.status >= 400:
            context.close()
            browser.close()
            return None

        pdf_bytes = page.pdf(
            format="A4",
            print_background=True,
            margin={
                "top": "8mm",
                "right": "8mm",
                "bottom": "8mm",
                "left": "8mm",
            }
        )

        context.close()
        browser.close()

    return pdf_bytes


import os
from django.conf import settings


def get_next_jobcard_pdf_filename(job):
    job_no = job.job_no

    folder = os.path.join(
        settings.MEDIA_ROOT,
        "jobcard_pdfs",
        job_no
    )

    os.makedirs(folder, exist_ok=True)

    existing = [
        f for f in os.listdir(folder)
        if f.endswith(".pdf")
    ]

    version = len(existing) + 1

    return f"jobcard_{job_no}_v{version}.pdf"


@login_required
def vehicle_detail_api(request, pk):
    vehicle = get_object_or_404(Vehicle, pk=pk)

    return JsonResponse({
        "id": vehicle.id,
        "registration_no": vehicle.registration_no,
        "id_chassis_no": vehicle.chassis_no,
        "id_engine_no": vehicle.engine_no,
        "id_vehicle_type": vehicle.vehicle_type,
        "model": vehicle.model_id,
        "variant": vehicle.variant_id,
        "id_color": vehicle.color,
        "id_sale_date": vehicle.sale_date,
        "insurance_company": vehicle.insurance_company_id,
        "insurance_company_name": vehicle.insurance_company.ins_co_name if vehicle.insurance_company else "",
        "policy_no": vehicle.policy_no,
        "policy_start_date": vehicle.policy_start_date,
        "policy_end_date": vehicle.policy_end_date,
        "rc_document": vehicle.rc_document.url if vehicle.rc_document else "",
        "insurance_policy_document": vehicle.insurance_policy_document.url if vehicle.insurance_policy_document else "",
        "primary_driver": vehicle.primary_driver_id,
        "assigned_drivers": list(vehicle.driver_master_records.values_list("id", flat=True)),
        "assigned_driver_details": [
            {"id": d.id, "name": d.name, "type": d.get_driver_type_display()}
            for d in vehicle.driver_master_records.all().order_by("name")
        ],
        "last_service_km": vehicle.last_service_km,
        "last_service_type": vehicle.last_service_type,
        "last_service_date": vehicle.last_service_date,
        "customer": vehicle.customer_id,
        "customer_name": vehicle.customer.name if vehicle.customer else "",
    })


@login_required
def unread_notifications(request):
    queryset = UserNotification.objects.filter(
        user=request.user,
        is_read=False
    ).order_by("-created_at")

    notifications = queryset[:10]

    data = [
        {
            "id": n.id,
            "title": n.title,
            "message": n.message,
            "url": n.url,
        }
        for n in notifications
    ]

    return JsonResponse({
        "count": queryset.count(),
        "notifications": data,
    })


@login_required
def mark_notification_read(request, pk):
    UserNotification.objects.filter(
        id=pk,
        user=request.user
    ).update(is_read=True)

    return JsonResponse({"status": "success"})


@login_required
def mark_all_notifications_read(request):
    updated = UserNotification.objects.filter(
        user=request.user,
        is_read=False
    ).update(is_read=True)

    return JsonResponse({
        "status": "success",
        "updated": updated,
    })


from django.contrib import messages


def validate_claim_stage_before_next(claim):
    missing = []
    # stage 3 → before going to stage 4
    if claim.claim_stage == ClaimStageCode.ADVISOR_ASSIGNED:

        has_jobcard = JobCard.objects.filter(
            claim=claim
        ).exists()

        if not has_jobcard:
            missing.append(
                "Job Card / Estimation not created"
            )

        if missing:
            return False, missing
    if claim.claim_stage == 3:

        if not claim.intimation_date:
            missing.append("Claim Intimation Date")

        if not claim.insurance_company:
            missing.append("Insurance Company")

        if not claim.policy_no:
            missing.append("Policy No")

        if not claim.ic_claim_no:
            missing.append("Insurance Claim No")

        if missing:
            return False, missing

    # stage 4 → before going to stage 5
    if claim.claim_stage == 4:

        missing = []

        if not claim.survey_date:
            missing.append("Survey Date")

        if not claim.surveyor:
            missing.append("Surveyor")

        if not claim.survey_status:
            missing.append("Survey Status")

        if missing:
            return False, missing

    # stage 5 → before going to stage 6
    if claim.claim_stage == 5:

        missing = []

        if not claim.insurance_approval_date:
            missing.append("Insurance Approval Date")

        if not claim.assessment_file:
            missing.append("Assessment File")

        if missing:
            return False, missing

    if claim.claim_stage == ClaimStageCode.WORK_ALLOCATION:
        has_progress_started = WorkProgress.objects.filter(
            allocation__job__claim=claim,
            start_time__isnull=False,
        ).exists()

        if not has_progress_started:
            return False, [
                "Start at least one progress stage from Work Allocation"
            ]

    if claim.claim_stage == ClaimStageCode.REPAIR_IN_PROGRESS:
        return False, [
            "Work Completed must be marked from Work Allocation"
        ]

    if claim.claim_stage == ClaimStageCode.INVOICED:

        missing = []

        if not claim.invoice_datetime:
            missing.append("Invoice Date & Time")

        if not claim.invoice_amount or claim.invoice_amount <= 0:
            missing.append("Invoice Amount")

        if not claim.payment_mode:
            missing.append("Payment Mode")

        if missing:
            return False, missing

    if claim.claim_stage == ClaimStageCode.DELIVERY:

        missing = []

        if not claim.delivery_datetime:
            missing.append("Delivery Date & Time")

        if not claim.delivered_by:
            missing.append("Delivered By")

        if not claim.delivered_to:
            missing.append("Delivered To")

        if claim.delivered_to == "Drop By Driver" and not claim.delivery_driver_name:
            missing.append("Driver Name")

        if missing:
            return False, missing

    return True, []


def get_work_progress_status(allocation):
    if not allocation:
        return "Work Not Started"

    progress_by_stage = {
        progress.stage: progress
        for progress in allocation.progress.all()
    }
    latest_started = None
    all_finished = True

    for stage_key, stage_label in WorkProgress.STAGES:
        progress = progress_by_stage.get(stage_key)

        if progress and progress.start_time:
            latest_started = (
                stage_label,
                "Completed" if progress.finish_time else "Started"
            )
            if not progress.finish_time:
                all_finished = False
            continue

        all_finished = False

    if not latest_started:
        return "Work Not Started"

    if all_finished:
        return "Work Completed"

    stage_label, status_text = latest_started
    return f"{stage_label} {status_text}"


def get_jobcard_main_status(job):
    if not job:
        return "Open"

    claim = job.claim

    if claim and claim.status == "Cancelled":
        return "Cancellation"

    if job.repair_status == "Closed":
        return "Closed"

    if claim and int(claim.claim_stage or 0) == ClaimStageCode.CLOSED:
        return "Closed"

    allocation = getattr(job, "allocation", None)

    if not allocation:
        return "Open"

    has_started = False
    all_finished = True

    for progress in allocation.progress.all():
        if progress.start_time:
            has_started = True

            if not progress.finish_time:
                all_finished = False
        else:
            all_finished = False

    if has_started and all_finished:
        return "Completed"

    return "Open"


def can_close_jobcard(job):
    return all(get_jobcard_close_ready_status(job).values())


def get_jobcard_close_ready_status(job):
    if not job:
        return {
            "work_completed": False,
            "qc_done": False,
            "ri_done": False,
            "part_entry_complete": False,
        }

    allocation = getattr(job, "allocation", None)

    work_completed = (
        job.repair_status in ["Completed", "Closed"]
        or (
            job.claim
            and int(job.claim.claim_stage or 0) >= ClaimStageCode.WORK_COMPLETED
        )
    )

    ri_done = bool(job.reinspection_done) or (
        job.claim
        and int(job.claim.claim_stage or 0) >= ClaimStageCode.LIABILITY
    )

    return {
        "work_completed": work_completed,
        "qc_done": bool(job.qc_done),
        "ri_done": ri_done,
        "part_entry_complete": bool(allocation and allocation.part_entry_complete),
    }


def get_jobcard_close_pending_items(job):
    status = get_jobcard_close_ready_status(job)
    labels = {
        "work_completed": "Work Completed",
        "qc_done": "QC Done",
        "ri_done": "RI Done",
        "part_entry_complete": "Part Entry Complete",
    }

    return [
        labels[key]
        for key, is_ready in status.items()
        if not is_ready
    ]


def sync_jobcard_main_status(job):
    status = get_jobcard_main_status(job)

    if job and job.repair_status != status:
        job.repair_status = status
        job.save(update_fields=["repair_status"])

    return status


from core.models import PartOrder


from core.models import PartOrder

def get_parts_not_available_status(job):
    allocation = getattr(job, "allocation", None)

    if allocation:
        pna_count = allocation.parts.filter(
            decision__in=["New", "KO"],
            pick_from_store=False
        ).count()

        if pna_count:
            return f"{pna_count} Parts Not Available Or Parts Not Issued"

    pending_count = PartOrder.objects.filter(
        job_id=job,
        status="Pending"
    ).count()

    return f"{pending_count} Parts Not Available Or Parts Not Issued" if pending_count else "No PNA"


def get_control_board_allocated_at(allocation):
    if not allocation:
        return None

    first_progress = (
        allocation.progress
        .filter(start_time__isnull=False)
        .order_by("start_time")
        .first()
    )

    if first_progress:
        return first_progress.start_time

    return allocation.allotment_date


def get_control_board_current_status(allocation):
    if not allocation:
        return "Work Allocation Pending", "bg-warning text-dark"

    job = allocation.job
    if (
        job.repair_status == "Completed"
        or int(job.claim.claim_stage or 0) >= ClaimStageCode.WORK_COMPLETED
    ):
        return "Work Completed", "bg-success"

    running_progress = (
        allocation.progress
        .filter(start_time__isnull=False, finish_time__isnull=True)
        .order_by("-start_time", "-id")
        .first()
    )
    if running_progress:
        return f"{running_progress.get_stage_display()} Running", "bg-primary"

    finished_progress = (
        allocation.progress
        .filter(finish_time__isnull=False)
        .order_by("-finish_time", "-id")
        .first()
    )
    if finished_progress:
        return f"{finished_progress.get_stage_display()} Completed", "bg-info text-dark"

    status = get_work_progress_status(allocation)
    return status, "bg-warning text-dark" if status == "Work Allocation Pending" else "bg-info text-dark"


def control_board_date_value(value):
    if not value:
        return None

    if hasattr(value, "date"):
        return timezone.localdate(value)

    return value


def get_control_board_in_date(job):
    return control_board_date_value(job.gate_in_datetime) or job.job_date


def get_control_board_promise_date(job):
    return (
        control_board_date_value(job.expected_delivery_datetime)
        or control_board_date_value(job.estimated_delivery)
    )


def get_control_board_tat_days(job):
    in_date = get_control_board_in_date(job)

    if not in_date:
        return ""

    claim = job.claim
    is_closed = (
        job.repair_status == "Closed"
        and int(claim.claim_stage or 0) == ClaimStageCode.CLOSED
    )

    end_date = timezone.localdate()
    if is_closed:
        end_date = (
            control_board_date_value(claim.delivery_datetime)
            or control_board_date_value(job.actual_delivery)
            or end_date
        )

    return max((end_date - in_date).days, 0)


def get_control_board_promise_class(job):
    promise_date = get_control_board_promise_date(job)
    if not promise_date:
        return ""

    today = timezone.localdate()
    days_left = (promise_date - today).days

    if days_left < 0:
        return "promise-overdue"

    if days_left <= 2:
        return "promise-warning"

    return ""


def control_board_datetime_value(value):
    if not value:
        return None

    if isinstance(value, datetime):
        return value if timezone.is_aware(value) else timezone.make_aware(value)

    value = datetime.combine(value, datetime_time.min)
    return timezone.make_aware(value)


def control_board_timeline_created_at(claim, keywords):
    if not claim:
        return None

    keywords = [keyword.upper() for keyword in keywords]
    timeline_rows = list(getattr(claim, "_prefetched_objects_cache", {}).get("timeline", []))
    if not timeline_rows:
        timeline_rows = list(claim.timeline.all())

    for row in timeline_rows:
        text = f"{row.stage or ''} {row.remarks or ''}".upper()
        if all(keyword in text for keyword in keywords):
            return row.created_at

    return None


def control_board_duration_text(start_at, end_at):
    if not start_at or not end_at:
        return "Not recorded"

    seconds = max(int((end_at - start_at).total_seconds()), 0)
    return control_board_duration_from_seconds(seconds)


def control_board_duration_from_seconds(seconds):
    seconds = max(int(seconds or 0), 0)
    days = seconds // 86400
    hours = (seconds % 86400) // 3600
    minutes = (seconds % 3600) // 60
    remaining_seconds = seconds % 60

    return f"{days}d {hours}h {minutes}m {remaining_seconds}s"


def control_board_timeline_display(value):
    if not value:
        return ""

    return timezone.localtime(value).strftime("%d-%b %I:%M:%S %p")


def get_control_board_tat_timeline(job):
    claim = job.claim
    try:
        allocation = job.allocation
    except WorkAllocation.DoesNotExist:
        allocation = None
    first_progress = None
    last_progress = None

    if allocation:
        first_progress = (
            allocation.progress
            .filter(start_time__isnull=False)
            .order_by("start_time")
            .first()
        )
        last_progress = (
            allocation.progress
            .filter(finish_time__isnull=False)
            .order_by("-finish_time")
            .first()
        )

    events = [
        ("Gate In", control_board_datetime_value(job.gate_in_datetime)),
        ("Claim Created", control_board_datetime_value(claim.created_at)),
        ("Jobcard Created", control_board_datetime_value(job.created_at)),
        ("Claim Intimation", control_board_datetime_value(claim.intimation_date)),
        ("Survey Done", control_board_datetime_value(claim.survey_date)),
        ("Insurance Approval", control_board_datetime_value(claim.insurance_approval_date)),
        ("Work Allocation", control_board_datetime_value(allocation.allotment_date) if allocation else None),
        ("Repair Started", first_progress.start_time if first_progress else None),
        ("Work Completed", last_progress.finish_time if last_progress else None),
        ("Re Inspection", control_board_datetime_value(job.reinspection_date)),
        ("Liability", control_board_datetime_value(claim.liability_received_at)),
        ("Invoiced", control_board_datetime_value(claim.invoice_datetime)),
        ("Delivery / Closed", control_board_datetime_value(claim.delivery_datetime)),
    ]

    rows = []
    now = timezone.now()
    for index in range(1, len(events)):
        start_label, start_at = events[index - 1]
        end_label, end_at = events[index]
        if not start_at:
            continue

        is_live = end_at is None
        effective_end_at = end_at or now
        rows.append({
            "label": f"{start_label} → {end_label}{' Pending' if is_live else ''}",
            "start_at": start_at,
            "end_at": end_at,
            "start_display": control_board_timeline_display(start_at),
            "end_display": control_board_timeline_display(effective_end_at),
            "duration": control_board_duration_text(start_at, effective_end_at),
            "duration_seconds": max(int((effective_end_at - start_at).total_seconds()), 0),
            "live": is_live,
            "pending": is_live,
        })
        if is_live:
            break

    return rows


def get_control_board_stage_and_status(allocation, pna_parts):
    job = allocation.job
    claim = job.claim

    if (
        job.repair_status == "Closed"
        and int(claim.claim_stage or 0) == ClaimStageCode.CLOSED
    ):
        return "Delivered", "🟢 Delivered", "status-ready"

    if (
        job.repair_status == "Completed"
        or int(claim.claim_stage or 0) >= ClaimStageCode.WORK_COMPLETED
    ):
        return "Work Completed", "🟢 Work Completed", "status-ready"

    if pna_parts:
        return "Parts Ordered", "🟠 Waiting Parts", "status-waiting"

    if job.ready_for_delivery or int(claim.claim_stage or 0) >= ClaimStageCode.DELIVERY:
        return "Ready For Delivery", "🟢 Ready For Delivery", "status-ready"

    if job.qc_done:
        return "QC Done", "🟢 Ready For Delivery", "status-ready"

    running_progress = (
        allocation.progress
        .filter(start_time__isnull=False, finish_time__isnull=True)
        .order_by("-start_time", "-id")
        .first()
    )
    if running_progress:
        return running_progress.get_stage_display(), "🔵 In Process", "status-process"

    return claim.get_claim_stage_display(), "🟡 Pending", "status-pending"


def get_control_board_progress_status(allocation):
    job = allocation.job

    if job.reinspection_done:
        return "RI Done"

    if job.qc_done:
        return "QC Done"

    if (
        job.repair_status == "Completed"
        or int(job.claim.claim_stage or 0) >= ClaimStageCode.WORK_COMPLETED
    ):
        return "Work Completed"

    running_progress = (
        allocation.progress
        .filter(start_time__isnull=False, finish_time__isnull=True)
        .order_by("-start_time", "-id")
        .first()
    )
    if running_progress:
        return f"{running_progress.get_stage_display()} Running"

    finished_progress = (
        allocation.progress
        .filter(finish_time__isnull=False)
        .order_by("-finish_time", "-id")
        .first()
    )
    if finished_progress:
        return f"{finished_progress.get_stage_display()} Completed"

    allocated_progress = (
        allocation.progress
        .filter(employee__isnull=False)
        .order_by("-id")
        .first()
    )
    if allocated_progress:
        return f"{allocated_progress.get_stage_display()} Allocated"

    return "Progress Not Started"


def get_control_board_pna_parts(allocation):
    if not allocation:
        return []

    return [
        part
        for part in allocation.parts.all()
        if part.decision in ["New", "KO"] and not part.pick_from_store
    ]


def progress_row_has_data(progress):
    return bool(
        progress
        and (
            progress.start_time
            or progress.finish_time
            or progress.employee_id
            or (progress.remarks or "").strip()
            or progress.photos.exists()
        )
    )


def claim_has_repair_progress_data(claim):
    if not claim:
        return False

    progress_qs = WorkProgress.objects.filter(
        allocation__job__claim=claim
    ).prefetch_related("photos")

    return any(progress_row_has_data(progress) for progress in progress_qs)


def workshop_resource_queryset():
    return Employee.objects.filter(
        is_active=True
    ).filter(
        Q(designation__icontains="Denter")
        | Q(designation__icontains="Painter")
        | Q(designation__icontains="Technician")
        | Q(designation__icontains="Paint")
        | Q(employee_type__icontains="Denter")
        | Q(employee_type__icontains="Painter")
        | Q(employee_type__icontains="Technician")
    ).order_by("designation", "name")


def clamp_hour(value, default):
    try:
        hour = int(value)
    except (TypeError, ValueError):
        return default

    return max(0, min(hour, 23))


def resource_slot_payload(progress):
    job = progress.allocation.job
    claim = job.claim
    vehicle = claim.vehicle if claim else None

    return {
        "job_id": job.id,
        "job_no": job.job_no,
        "claim_no": claim.claim_no if claim else "",
        "registration_no": vehicle.registration_no if vehicle else "",
        "model": (
            vehicle.model.name
            if vehicle and vehicle.model
            else ""
        ),
        "stage": progress.get_stage_display(),
        "is_running": not bool(progress.finish_time),
    }


@never_cache
@xframe_options_sameorigin
@login_required
def workshop_resource_view(request):
    from datetime import time, timedelta

    logged_emp = Employee.objects.filter(user=request.user).first()

    if not (
        request.user.is_superuser
        or is_admin_or_manager_user(request.user, logged_emp)
        or is_floor_supervisor(logged_emp)
    ):
        messages.error(request, "You are not allowed to access Workshop Resource.")
        return redirect("dashboard")

    today = timezone.localdate()
    from_date = parse_date(request.GET.get("from_date") or "") or today
    to_date = parse_date(request.GET.get("to_date") or "") or from_date

    if to_date < from_date:
        to_date = from_date

    max_days = 7
    if (to_date - from_date).days >= max_days:
        to_date = from_date + timedelta(days=max_days - 1)
        messages.info(request, "Workshop resource view is limited to 7 days at a time.")

    start_hour = clamp_hour(request.GET.get("start_hour"), 9)
    end_hour = clamp_hour(request.GET.get("end_hour"), 19)
    selected_employee_id = request.GET.get("employee_id") or ""
    resource_query = (request.GET.get("resource_q") or "").strip()
    role_filter = (request.GET.get("role") or "").strip()

    if end_hour <= start_hour:
        end_hour = min(start_hour + 1, 23)

    previous_from_date = from_date - timedelta(days=1)
    previous_to_date = to_date - timedelta(days=1)
    next_from_date = from_date + timedelta(days=1)
    next_to_date = to_date + timedelta(days=1)
    hours = list(range(start_hour, end_hour))
    base_resources = list(workshop_resource_queryset())
    resource_roles = sorted({
        resource.designation or resource.employee_type
        for resource in base_resources
        if resource.designation or resource.employee_type
    })
    filtered_resources = base_resources

    if resource_query:
        query = resource_query.lower()
        filtered_resources = [
            resource
            for resource in filtered_resources
            if query in (resource.name or "").lower()
            or query in (resource.employee_code or "").lower()
            or query in (resource.designation or "").lower()
        ]

    if role_filter:
        filtered_resources = [
            resource
            for resource in filtered_resources
            if (resource.designation or resource.employee_type) == role_filter
        ]

    resources = filtered_resources
    current_tz = timezone.get_current_timezone()
    range_start = timezone.make_aware(
        datetime.combine(from_date, time(start_hour, 0)),
        current_tz,
    )
    range_end = timezone.make_aware(
        datetime.combine(to_date, time(end_hour, 0)),
        current_tz,
    )

    progress_rows = list(
        WorkProgress.objects.select_related(
            "employee",
            "allocation",
            "allocation__job",
            "allocation__job__claim",
            "allocation__job__claim__vehicle",
            "allocation__job__claim__vehicle__model",
        )
        .filter(
            employee__in=resources,
            start_time__isnull=False,
            start_time__lt=range_end,
        )
        .filter(
            Q(finish_time__isnull=True)
            | Q(finish_time__gt=range_start)
        )
        .order_by("start_time", "id")
    )

    progress_by_employee = {}
    for progress in progress_rows:
        progress_by_employee.setdefault(progress.employee_id, []).append(progress)

    days = []
    current_day = from_date
    now = timezone.now()

    while current_day <= to_date:
        day_rows = []

        for employee in resources:
            cells = []
            busy_hours = 0

            for hour in hours:
                slot_start = timezone.make_aware(
                    datetime.combine(current_day, time(hour, 0)),
                    current_tz,
                )
                slot_end = slot_start + timedelta(hours=1)
                items = []

                for progress in progress_by_employee.get(employee.id, []):
                    progress_start = progress.start_time
                    progress_end = progress.finish_time or now

                    if progress_end < progress_start:
                        progress_end = progress_start

                    if progress_start < slot_end and progress_end > slot_start:
                        items.append(resource_slot_payload(progress))

                if items:
                    busy_hours += 1

                cells.append({
                    "label": f"{hour:02d}:00",
                    "items": items,
                })

            day_rows.append({
                "employee": employee,
                "role": employee.designation or employee.employee_type,
                "busy_hours": busy_hours,
                "available_hours": len(hours) - busy_hours,
                "cells": cells,
                "is_selected": str(employee.id) == selected_employee_id,
            })

        busy_resource_count = sum(1 for row in day_rows if row["busy_hours"] > 0)
        days.append({
            "date": current_day,
            "rows": day_rows,
            "busy_resource_count": busy_resource_count,
            "free_resource_count": len(day_rows) - busy_resource_count,
        })
        current_day += timedelta(days=1)

    return render(request, "floor/workshopResource.html", {
        "days": days,
        "hours": hours,
        "filter_from_date": from_date.strftime("%Y-%m-%d"),
        "filter_to_date": to_date.strftime("%Y-%m-%d"),
        "start_hour": start_hour,
        "end_hour": end_hour,
        "previous_from_date": previous_from_date.strftime("%Y-%m-%d"),
        "previous_to_date": previous_to_date.strftime("%Y-%m-%d"),
        "next_from_date": next_from_date.strftime("%Y-%m-%d"),
        "next_to_date": next_to_date.strftime("%Y-%m-%d"),
        "resource_count": len(resources),
        "total_resource_count": len(base_resources),
        "resource_query": resource_query,
        "role_filter": role_filter,
        "resource_roles": resource_roles,
        "selected_employee_id": selected_employee_id,
        "modal_mode": request.GET.get("modal") == "1",
        "breadcrumbs": [
            {
                "title": "WorkShop / Floor",
                "url": "",
                "icon": "fa fa-list",
            },
            {
                "title": "Workshop Resource",
                "icon": "fa fa-calendar",
            },
        ],
    })


@never_cache
@login_required
def work_allocation_list(request):
    logged_emp = Employee.objects.filter(
        user=request.user
    ).first()

    if not is_floor_supervisor(logged_emp):
        messages.error(request, "You are not allowed to access Work Allocation")
        return redirect("dashboard")

    jobs = JobCard.objects.select_related(
        "claim",
        "claim__vehicle",
        "claim__vehicle__customer",
        "claim__vehicle__model",

        "vehicle",
        "vehicle__customer",
        "vehicle__model",

        "advisor",
        "allocation",
        "jobcard_type",
    ).filter(
        Q(
            claim__isnull=False,
            claim__claim_stage__gte=ClaimStageCode.WORK_ALLOCATION
        )
        |
        Q(
            claim__isnull=True,
            jobcard_type__name="Paid"
        )
        |
        Q(
            claim__isnull=True,
            approval_evidence__status="Approved"
        )
        |
        Q(
            allocation__isnull=False
        )
    ).distinct().prefetch_related(
        "allocation__progress"
    ).order_by("-id")

    for job in jobs:

        allocation = getattr(job, "allocation", None)

        # Get vehicle from Claim OR Direct JobCard Vehicle
        if job.claim and job.claim.vehicle:
            vehicle = job.claim.vehicle
        else:
            vehicle = job.vehicle

        job.display_vehicle = vehicle

        job.display_vehicle_no = (
            vehicle.registration_no
            if vehicle else ""
        )

        job.display_customer_name = (
            vehicle.customer.name
            if vehicle and vehicle.customer
            else ""
        )

        job.display_model_name = (
            vehicle.model.name
            if vehicle and vehicle.model
            else ""
        )

        job.work_allocation_action = "Allocate"
        job.work_allocation_status = "Work Allocation Pending"
        job.work_allocation_status_class = "bg-warning text-dark"

        if not allocation:
            continue

        job.work_allocation_action = "Edit"

        if (
                job.repair_status == "Completed"
                or (
                job.claim
                and int(job.claim.claim_stage or 0) >= ClaimStageCode.WORK_COMPLETED
        )
        ):
            ...
            job.work_allocation_status = "Work Completed"
            job.work_allocation_status_class = "bg-success"
        elif (
            job.additional_approval_required
            and (job.second_approval_status or "Pending") == "Pending"
        ):
            job.work_allocation_status = (
                f"2nd Approval {job.second_approval_status or 'Pending'}"
            )
            job.work_allocation_status_class = "bg-danger"
        else:
            job.work_allocation_status = get_work_progress_status(allocation)
            job.work_allocation_status_class = (
                "bg-success"
                if job.work_allocation_status == "Work Completed"
                else "bg-info text-dark"
            )

    work_progress_count = sum(
        1
        for job in jobs
        if getattr(job, "work_allocation_status", "") not in [
            "Work Allocation Pending",
            "Work Completed",
        ]
    )
    pending_count = sum(
        1
        for job in jobs
        if getattr(job, "work_allocation_status", "") == "Work Allocation Pending"
    )
    completed_count = sum(
        1
        for job in jobs
        if getattr(job, "work_allocation_status", "") == "Work Completed"
    )

    return render(request, "floor/workAllocationList.html", {
        "jobs": jobs,
        "logged_emp": logged_emp,
        "work_progress_count": work_progress_count,
        "pending_count": pending_count,
        "completed_count": completed_count,
        "breadcrumbs": [
            {
                "title": "WorkShop/Floor",
                "url": "",
                "icon": "fa fa-list"
            },
            {
                "title": "Work Allocation",
                "icon": "fa fa-tools"
            }
        ]
    })


@never_cache
@login_required
def report_placeholder(request, report_key):
    report_title = str(report_key or "").strip("/").replace("-", " ").replace("/", " / ").title()
    return render(request, "reports/reportPlaceholder.html", {
        "report_title": report_title,
        "report_key": report_key,
        "breadcrumbs": [
            {"title": "Reports", "url": "", "icon": "fa fa-chart-bar"},
            {"title": report_title, "icon": "fa fa-file-alt"},
        ],
    })


@never_cache
@login_required
def daily_in_out_report(request):
    today = timezone.localdate()
    branch_context = report_branch_context(request)
    from_date = parse_date(request.GET.get("from_date") or "") or today
    to_date = parse_date(request.GET.get("to_date") or "") or from_date

    if to_date < from_date:
        from_date, to_date = to_date, from_date

    days = []
    cursor = from_date
    while cursor <= to_date:
        vehicle_in = apply_report_branch_scope(
            JobCard.objects.filter(
                gate_in_datetime__date=cursor
            ),
            branch_context,
            "claim__branch",
        ).count()
        vehicle_delivered = apply_report_branch_scope(
            JobCard.objects.filter(
                claim__delivery_datetime__date=cursor
            ),
            branch_context,
            "claim__branch",
        ).count()
        wip = apply_report_branch_scope(
            JobCard.objects.filter(
                gate_in_datetime__date__lte=cursor
            ),
            branch_context,
            "claim__branch",
        ).exclude(
            Q(claim__delivery_datetime__date__lte=cursor)
            | Q(claim__claim_stage=ClaimStageCode.CLOSED)
            | Q(repair_status="Closed")
        ).count()

        days.append({
            "date": cursor,
            "vehicle_in": vehicle_in,
            "vehicle_delivered": vehicle_delivered,
            "wip": wip,
        })
        cursor += timedelta(days=1)

    total_gate_in = sum(row["vehicle_in"] for row in days)
    total_deliveries = sum(row["vehicle_delivered"] for row in days)
    pending_vehicles = days[-1]["wip"] if days else 0

    return render(request, "reports/dailyInOutReport.html", {
        "from_date": from_date,
        "to_date": to_date,
        "rows": days,
        "total_gate_in": total_gate_in,
        "total_deliveries": total_deliveries,
        "pending_vehicles": pending_vehicles,
        **branch_context,
        "breadcrumbs": [
            {"title": "Reports", "url": "", "icon": "fa fa-chart-bar"},
            {"title": "Workshop Reports", "url": "", "icon": "fa fa-industry"},
            {"title": "Daily In/Out", "icon": "fa fa-exchange-alt"},
        ],
    })


def current_financial_year_start(today=None):
    today = today or timezone.localdate()
    return today.year if today.month >= 4 else today.year - 1


def month_on_month_tat_rows(jobs, fy_start, target_days):
    month_order = [4, 5, 6, 7, 8, 9, 10, 11, 12, 1, 2, 3]
    month_names = {
        1: "January",
        2: "February",
        3: "March",
        4: "April",
        5: "May",
        6: "June",
        7: "July",
        8: "August",
        9: "September",
        10: "October",
        11: "November",
        12: "December",
    }
    month_totals = {month: {"days": 0.0, "count": 0} for month in month_order}

    for job in jobs:
        claim = job.claim
        end_at = claim.delivery_datetime if claim else None
        start_at = job.gate_in_datetime or job.job_date or job.created_at
        if not start_at or not end_at:
            continue

        if timezone.is_naive(start_at):
            start_at = timezone.make_aware(start_at)
        if timezone.is_naive(end_at):
            end_at = timezone.make_aware(end_at)

        seconds = (end_at - start_at).total_seconds()
        if seconds < 0:
            continue

        delivered_month = timezone.localtime(end_at).month
        month_totals[delivered_month]["days"] += seconds / 86400
        month_totals[delivered_month]["count"] += 1

    avg_values = [
        month_totals[month]["days"] / month_totals[month]["count"]
        for month in month_order
        if month_totals[month]["count"]
    ]
    chart_max = max(14, math.ceil(max(avg_values + [target_days]) + 1))
    target_percent = min(100, max(0, (target_days / chart_max) * 100))

    rows = []
    for month in month_order:
        count = month_totals[month]["count"]
        avg_days = month_totals[month]["days"] / count if count else None
        rows.append({
            "month": month,
            "label": month_names[month],
            "year": fy_start if month >= 4 else fy_start + 1,
            "count": count,
            "avg_days": avg_days,
            "display": f"{avg_days:.2f}" if avg_days is not None else "#N/A",
            "height_percent": (avg_days / chart_max) * 100 if avg_days is not None else 0,
            "color_class": "tat-bar-good" if avg_days is not None and avg_days <= target_days else "tat-bar-delay",
            "status": "Below Target" if avg_days is not None and avg_days <= target_days else "Above Target",
            "is_na": avg_days is None,
        })

    total_count = sum(row["count"] for row in rows)
    overall_avg = (
        sum(month_totals[month]["days"] for month in month_order) / total_count
        if total_count
        else None
    )

    return rows, chart_max, target_percent, total_count, overall_avg


@never_cache
@login_required
def tat_summary_report(request):
    branch_context = report_branch_context(request)
    default_fy = current_financial_year_start()

    try:
        fy_start = int(request.GET.get("fy") or default_fy)
    except (TypeError, ValueError):
        fy_start = default_fy

    try:
        target_days = float(request.GET.get("target_days") or 7)
    except (TypeError, ValueError):
        target_days = 7

    target_days = max(1, min(target_days, 60))
    start_date = date(fy_start, 4, 1)
    end_date = date(fy_start + 1, 3, 31)

    jobs = apply_report_branch_scope(
        JobCard.objects.select_related("claim", "claim__branch").filter(
            claim__delivery_datetime__date__gte=start_date,
            claim__delivery_datetime__date__lte=end_date,
        ),
        branch_context,
        "claim__branch",
    )

    month_rows, chart_max, target_percent, total_count, overall_avg = month_on_month_tat_rows(
        jobs,
        fy_start,
        target_days,
    )
    y_axis_ticks = []
    for tick_index in range(8):
        value = chart_max - ((chart_max / 7) * tick_index)
        y_axis_ticks.append({
            "bottom": max(0, min(100, (value / chart_max) * 100)),
            "label": f"{value:.0f}",
        })
    months_with_data = [row for row in month_rows if not row["is_na"]]
    below_target_count = sum(1 for row in months_with_data if row["avg_days"] <= target_days)
    above_target_count = sum(1 for row in months_with_data if row["avg_days"] > target_days)
    fy_options = list(range(default_fy + 1, default_fy - 5, -1))

    return render(request, "reports/tatSummaryReport.html", {
        "fy_start": fy_start,
        "fy_end": fy_start + 1,
        "fy_options": fy_options,
        "target_days": target_days,
        "target_days_display": f"{target_days:g}",
        "target_percent": target_percent,
        "chart_max": chart_max,
        "y_axis_ticks": y_axis_ticks,
        "month_rows": month_rows,
        "total_count": total_count,
        "overall_avg": overall_avg,
        "overall_avg_display": f"{overall_avg:.2f}" if overall_avg is not None else "#N/A",
        "below_target_count": below_target_count,
        "above_target_count": above_target_count,
        **branch_context,
        "breadcrumbs": [
            {"title": "Reports", "url": "", "icon": "fa fa-chart-bar"},
            {"title": "TAT Reports", "url": "", "icon": "fa fa-clock"},
            {"title": "TAT Summary", "icon": "fa fa-chart-column"},
        ],
    })


@never_cache
@login_required
def kpi_cards_report(request):
    today = timezone.localdate()
    branch_context = report_branch_context(request)

    jobs = apply_report_branch_scope(
        JobCard.objects.select_related("claim"),
        branch_context,
        "claim__branch",
    )
    claims = apply_report_branch_scope(
        Claim.objects.all(),
        branch_context,
    )

    open_jobs = jobs.filter(
        claim__claim_stage__lt=ClaimStageCode.CLOSED
    ).exclude(repair_status="Closed")

    total_gate_in_today = jobs.filter(
        gate_in_datetime__date=today
    ).count()
    total_delivered_today = jobs.filter(
        claim__delivery_datetime__date=today
    ).count()
    current_wip = open_jobs.filter(
        gate_in_datetime__isnull=False
    ).count()
    overdue_vehicles = open_jobs.filter(
        expected_delivery_datetime__date__lt=today
    ).exclude(
        claim__delivery_datetime__isnull=False
    ).count()
    open_claims = claims.filter(
        claim_stage__lt=ClaimStageCode.CLOSED
    )
    survey_pending = open_claims.filter(
        claim_stage=ClaimStageCode.INTIMATION
    ).count()
    insurance_approval_pending = open_claims.filter(
        claim_stage__gte=ClaimStageCode.SURVEY,
        claim_stage__lt=ClaimStageCode.INSURANCE_APPROVAL,
    ).count()
    work_allocation_pending = open_claims.filter(
        claim_stage=ClaimStageCode.WORK_ALLOCATION
    ).count()
    repair_in_progress = open_claims.filter(
        claim_stage=ClaimStageCode.REPAIR_IN_PROGRESS
    ).count()
    ready_for_delivery = open_jobs.filter(
        Q(ready_for_delivery=True)
        | Q(claim__claim_stage=ClaimStageCode.DELIVERY)
    ).count()

    cards = [
        {
            "title": "Total Gate In Today",
            "value": total_gate_in_today,
            "icon": "fa fa-sign-in-alt",
            "tone": "primary",
            "caption": today.strftime("%d-%b-%Y"),
        },
        {
            "title": "Total Delivered Today",
            "value": total_delivered_today,
            "icon": "fa fa-car-side",
            "tone": "success",
            "caption": today.strftime("%d-%b-%Y"),
        },
        {
            "title": "Current WIP",
            "value": current_wip,
            "icon": "fa fa-tools",
            "tone": "info",
            "caption": "Open vehicles in workshop",
        },
        {
            "title": "Overdue Vehicles",
            "value": overdue_vehicles,
            "icon": "fa fa-exclamation-triangle",
            "tone": "danger",
            "caption": "Promise date crossed",
        },
        {
            "title": "Survey Pending",
            "value": survey_pending,
            "icon": "fa fa-search",
            "tone": "warning",
            "caption": "Claim intimation done",
        },
        {
            "title": "Insurance Approval Pending",
            "value": insurance_approval_pending,
            "icon": "fa fa-file-signature",
            "tone": "warning",
            "caption": "Survey done, approval pending",
        },
        {
            "title": "Work Allocation Pending",
            "value": work_allocation_pending,
            "icon": "fa fa-clipboard-list",
            "tone": "secondary",
            "caption": "Waiting for floor allocation",
        },
        {
            "title": "Repair In Progress",
            "value": repair_in_progress,
            "icon": "fa fa-tools",
            "tone": "info",
            "caption": "Active repair workflow",
        },
        {
            "title": "Ready for Delivery",
            "value": ready_for_delivery,
            "icon": "fa fa-check-circle",
            "tone": "success",
            "caption": "Delivery action pending",
        },
    ]

    return render(request, "reports/kpiCardsReport.html", {
        "cards": cards,
        "today": today,
        **branch_context,
        "breadcrumbs": [
            {"title": "Reports", "url": "", "icon": "fa fa-chart-bar"},
            {"title": "Graphical Dashboard", "url": "", "icon": "fa fa-chart-pie"},
            {"title": "KPI Cards", "icon": "fa fa-id-card"},
        ],
    })


@never_cache
@login_required
def surveyor_performance_report(request):
    today = timezone.localdate()
    branch_context = report_branch_context(request)
    month_start = today.replace(day=1)
    from_date = parse_date(request.GET.get("from_date") or "") or month_start
    to_date = parse_date(request.GET.get("to_date") or "") or today
    try:
        target_tat = Decimal(request.GET.get("target_tat") or "30")
    except Exception:
        target_tat = Decimal("30")
    if target_tat <= 0:
        target_tat = Decimal("30")

    if to_date < from_date:
        from_date, to_date = to_date, from_date

    claims = apply_report_branch_scope(
        Claim.objects
        .select_related("surveyor")
        .filter(
            surveyor__isnull=False,
            survey_date__date__gte=from_date,
            survey_date__date__lte=to_date,
        ),
        branch_context,
    )

    surveyor_stats = {}
    survey_tat_target = Decimal("1")
    survey_tat_buckets = {
        "on_time": 0,
        "delayed": 0,
        "critical": 0,
    }
    for claim in claims:
        survey_start = workflow_date_value(claim.survey_date)
        if not survey_start:
            continue

        approval_date = workflow_date_value(claim.insurance_approval_date)
        tat_end = approval_date or timezone.now()
        tat_days = max((tat_end - survey_start).total_seconds() / 86400, 0)
        surveyor_name = claim.surveyor.name if claim.surveyor else "Unknown"
        bucket = surveyor_stats.setdefault(
            surveyor_name,
            {
                "surveyor": surveyor_name,
                "total_days": 0,
                "survey_days": 0,
                "approval_days": 0,
                "liability_days": 0,
                "survey_on_time": 0,
                "survey_delayed": 0,
                "survey_critical": 0,
                "claim_count": 0,
                "approved_count": 0,
                "pending_count": 0,
            }
        )
        bucket["total_days"] += tat_days
        survey_base = workflow_date_value(claim.intimation_date)
        survey_tat = max((survey_start - survey_base).total_seconds() / 86400, 0) if survey_base else 0
        if Decimal(str(survey_tat)) > survey_tat_target + Decimal("2"):
            survey_tat_buckets["critical"] += 1
            bucket["survey_critical"] += 1
        elif Decimal(str(survey_tat)) > survey_tat_target:
            survey_tat_buckets["delayed"] += 1
            bucket["survey_delayed"] += 1
        else:
            survey_tat_buckets["on_time"] += 1
            bucket["survey_on_time"] += 1
        approval_tat = max((tat_end - survey_start).total_seconds() / 86400, 0)
        liability_start = workflow_date_value(claim.pre_invoice_sent_at)
        liability_end = workflow_date_value(claim.liability_received_at)
        liability_tat = max((liability_end - liability_start).total_seconds() / 86400, 0) if liability_start and liability_end else 0
        bucket["survey_days"] += survey_tat
        bucket["approval_days"] += approval_tat
        bucket["liability_days"] += liability_tat
        bucket["claim_count"] += 1
        if approval_date:
            bucket["approved_count"] += 1
        else:
            bucket["pending_count"] += 1

    rows = []
    for item in surveyor_stats.values():
        avg_tat = item["total_days"] / item["claim_count"] if item["claim_count"] else 0
        item["avg_tat"] = round(avg_tat, 1)
        item["avg_survey_tat"] = round(item["survey_days"] / item["claim_count"], 1) if item["claim_count"] else 0
        item["avg_approval_tat"] = round(item["approval_days"] / item["claim_count"], 1) if item["claim_count"] else 0
        item["avg_liability_tat"] = round(item["liability_days"] / item["claim_count"], 1) if item["claim_count"] else 0
        item["survey_on_time_percent"] = round((item["survey_on_time"] / item["claim_count"]) * 100, 1) if item["claim_count"] else 0
        item["survey_delayed_percent"] = round((item["survey_delayed"] / item["claim_count"]) * 100, 1) if item["claim_count"] else 0
        item["survey_critical_percent"] = round((item["survey_critical"] / item["claim_count"]) * 100, 1) if item["claim_count"] else 0
        item["survey_delayed_start"] = item["survey_on_time_percent"]
        item["survey_critical_start"] = item["survey_on_time_percent"] + item["survey_delayed_percent"]
        avg_total_tat = (
            Decimal(str(item["avg_survey_tat"]))
            + Decimal(str(item["avg_approval_tat"]))
            + Decimal(str(item["avg_liability_tat"]))
        )
        score = Decimal("100") - ((avg_total_tat / target_tat) * Decimal("100"))
        item["score"] = max(0, min(100, int(score.quantize(Decimal("1")))))
        rows.append(item)

    rows.sort(key=lambda row: row["avg_tat"])
    max_avg_tat = max((row["avg_tat"] for row in rows), default=0)
    for row in rows:
        row["bar_width"] = int((row["avg_tat"] / max_avg_tat) * 100) if max_avg_tat else 0
        row["donut_percent"] = round((row["avg_tat"] / max_avg_tat) * 100, 1) if max_avg_tat else 0
    axis_step = 0.5
    axis_max = math.ceil(max_avg_tat / axis_step) * axis_step if max_avg_tat else axis_step
    axis_ticks = []
    tick_count = int(axis_max / axis_step)
    for index in range(tick_count + 1):
        value = round(index * axis_step, 1)
        axis_ticks.append({
            "value": value,
            "label": f"{value:g} days",
            "left": int((value / axis_max) * 100) if axis_max else 0,
        })

    overall_avg_tat = round(
        sum(row["avg_tat"] * row["claim_count"] for row in rows)
        / sum(row["claim_count"] for row in rows),
        1
    ) if rows else 0
    survey_tat_total = sum(survey_tat_buckets.values())
    survey_tat_donut = {
        "on_time": survey_tat_buckets["on_time"],
        "delayed": survey_tat_buckets["delayed"],
        "critical": survey_tat_buckets["critical"],
        "total": survey_tat_total,
        "on_time_percent": round((survey_tat_buckets["on_time"] / survey_tat_total) * 100, 1) if survey_tat_total else 0,
        "delayed_percent": round((survey_tat_buckets["delayed"] / survey_tat_total) * 100, 1) if survey_tat_total else 0,
        "critical_percent": round((survey_tat_buckets["critical"] / survey_tat_total) * 100, 1) if survey_tat_total else 0,
    }
    survey_tat_donut["delayed_start"] = survey_tat_donut["on_time_percent"]
    survey_tat_donut["critical_start"] = survey_tat_donut["on_time_percent"] + survey_tat_donut["delayed_percent"]

    return render(request, "reports/surveyorPerformanceReport.html", {
        "from_date": from_date,
        "to_date": to_date,
        "rows": rows,
        "overall_avg_tat": overall_avg_tat,
        "total_claims": sum(row["claim_count"] for row in rows),
        "pending_claims": sum(row["pending_count"] for row in rows),
        "target_tat": target_tat,
        "axis_ticks": axis_ticks,
        "survey_tat_target": survey_tat_target,
        "survey_tat_donut": survey_tat_donut,
        **branch_context,
        "breadcrumbs": [
            {"title": "Reports", "url": "", "icon": "fa fa-chart-bar"},
            {"title": "Insurance Reports", "url": "", "icon": "fa fa-file-contract"},
            {"title": "Surveyor Performance", "icon": "fa fa-chart-bar"},
        ],
    })


@never_cache
@login_required
def bodyshop_control_menu(request):
    logged_emp = Employee.objects.filter(user=request.user).first()

    if not (
        request.user.is_superuser
        or is_admin_or_manager_user(request.user, logged_emp)
        or is_floor_supervisor(logged_emp)
    ):
        messages.error(request, "You are not allowed to access Bodyshop Control Board")
        return redirect("dashboard")

    today = timezone.localdate()
    allocations = (
        WorkAllocation.objects
        .select_related(
            "job",
            "job__claim",
            "job__claim__vehicle",
            "job__claim__vehicle__customer",
            "job__claim__vehicle__model",
            "job__claim__insurance_company",
            "job__claim__surveyor",
            "job__advisor",
        )
        .prefetch_related("progress", "parts", "parts__job_part")
        .filter(job__claim__claim_stage__lt=ClaimStageCode.CLOSED)
        .exclude(job__repair_status="Closed")
    )
    unallocated_jobs = (
        JobCard.objects
        .select_related(
            "claim",
            "claim__vehicle",
            "claim__vehicle__customer",
            "claim__vehicle__model",
            "claim__insurance_company",
            "claim__surveyor",
            "advisor",
        )
        .filter(claim__claim_stage__lt=ClaimStageCode.CLOSED)
        .exclude(repair_status="Closed")
        .filter(allocation__isnull=True)
    )

    status_counts = {
        "pending": 0,
        "waiting_parts": 0,
        "in_process": 0,
        "work_completed": 0,
        "ready": 0,
        "delivered": 0,
    }
    delay_count = 0
    today_delivery_count = 0
    parts_pending_count = 0
    technician_workload = {}
    tat_buckets = {
        "0-2 Days": 0,
        "3-5 Days": 0,
        "6-10 Days": 0,
        "10+ Days": 0,
    }

    for allocation in allocations:
        job = allocation.job
        pna_parts = get_control_board_pna_parts(allocation)
        _, board_status, _ = get_control_board_stage_and_status(allocation, pna_parts)
        status_key = "pending"

        if "Waiting Parts" in board_status:
            status_key = "waiting_parts"
        elif "In Process" in board_status:
            status_key = "in_process"
        elif "Work Completed" in board_status:
            status_key = "work_completed"
        elif "Ready For Delivery" in board_status:
            status_key = "ready"
        elif "Delivered" in board_status:
            status_key = "delivered"

        status_counts[status_key] += 1

        promise_date = get_control_board_promise_date(job)
        if promise_date and promise_date < today:
            delay_count += 1
        if promise_date == today or (
            job.claim.delivery_datetime
            and workflow_date_value(job.claim.delivery_datetime).date() == today
        ):
            today_delivery_count += 1

        if pna_parts:
            parts_pending_count += 1

        tat_days = get_control_board_tat_days(job)
        if tat_days != "":
            if tat_days <= 2:
                tat_buckets["0-2 Days"] += 1
            elif tat_days <= 5:
                tat_buckets["3-5 Days"] += 1
            elif tat_days <= 10:
                tat_buckets["6-10 Days"] += 1
            else:
                tat_buckets["10+ Days"] += 1

        running_rows = allocation.progress.filter(
            start_time__isnull=False,
            finish_time__isnull=True,
            employee__isnull=False,
        )
        for progress in running_rows:
            name = progress.employee.name if progress.employee else "Unassigned"
            technician_workload[name] = technician_workload.get(name, 0) + 1

    for job in unallocated_jobs:
        status_counts["pending"] += 1

        promise_date = get_control_board_promise_date(job)
        if promise_date and promise_date < today:
            delay_count += 1
        if promise_date == today or (
            job.claim.delivery_datetime
            and workflow_date_value(job.claim.delivery_datetime).date() == today
        ):
            today_delivery_count += 1

        tat_days = get_control_board_tat_days(job)
        if tat_days != "":
            if tat_days <= 2:
                tat_buckets["0-2 Days"] += 1
            elif tat_days <= 5:
                tat_buckets["3-5 Days"] += 1
            elif tat_days <= 10:
                tat_buckets["6-10 Days"] += 1
            else:
                tat_buckets["10+ Days"] += 1

    insurance_approval_pending = Claim.objects.filter(
        claim_stage__gte=ClaimStageCode.SURVEY,
        claim_stage__lt=ClaimStageCode.INSURANCE_APPROVAL,
    ).count()

    workload_rows = sorted(
        [
            {"name": name, "count": count}
            for name, count in technician_workload.items()
        ],
        key=lambda row: row["count"],
        reverse=True,
    )[:8]

    tat_rows = [
        {"label": label, "count": count}
        for label, count in tat_buckets.items()
    ]
    max_tat_count = max([row["count"] for row in tat_rows] + [1])
    for row in tat_rows:
        row["width"] = max(int((row["count"] / max_tat_count) * 100), 6) if row["count"] else 0

    module_cards = [
        {
            "title": "Stage-wise Kanban Board",
            "subtitle": "Live vehicle stage board",
            "count": allocations.count() + unallocated_jobs.count(),
            "class": "module-primary",
            "url": reverse("bodyshop_control_board"),
        },
        {
            "title": "Delay Vehicles",
            "subtitle": "Promise date overdue",
            "count": delay_count,
            "class": "module-danger",
            "url": f"{reverse('bodyshop_control_board')}?status=open",
        },
        {
            "title": "Today's Delivery List",
            "subtitle": "Due or delivered today",
            "count": today_delivery_count,
            "class": "module-success",
            "url": f"{reverse('bodyshop_control_board')}?status=open",
        },
        {
            "title": "Insurance Approval Pending",
            "subtitle": "Survey done, approval pending",
            "count": insurance_approval_pending,
            "class": "module-warning",
            "url": f"{reverse('bodyshop_control_board')}?status=open",
        },
        {
            "title": "Parts Pending",
            "subtitle": "PNA / waiting parts",
            "count": parts_pending_count,
            "class": "module-orange",
            "url": reverse("part_order_list"),
        },
        {
            "title": "Technician Workload",
            "subtitle": "Running jobs by resource",
            "count": sum(row["count"] for row in workload_rows),
            "class": "module-info",
            "url": reverse("workshop_resource"),
        },
        {
            "title": "TAT Graph",
            "subtitle": "Ageing by in-date",
            "count": sum(row["count"] for row in tat_rows),
            "class": "module-purple",
            "url": reverse("bodyshop_control_board"),
        },
        {
            "title": "Live Vehicle Count by Status",
            "subtitle": "Current board summary",
            "count": sum(status_counts.values()),
            "class": "module-dark",
            "url": reverse("bodyshop_control_board"),
        },
    ]

    status_rows = [
        {"label": "Pending", "count": status_counts["pending"], "class": "status-pending"},
        {"label": "Waiting Parts", "count": status_counts["waiting_parts"], "class": "status-waiting"},
        {"label": "In Process", "count": status_counts["in_process"], "class": "status-process"},
        {"label": "Work Completed", "count": status_counts["work_completed"], "class": "status-completed"},
        {"label": "Ready For Delivery", "count": status_counts["ready"], "class": "status-ready"},
        {"label": "Delivered", "count": status_counts["delivered"], "class": "status-delivered"},
    ]

    return render(request, "floor/bodyshopControlMenu.html", {
        "module_cards": module_cards,
        "status_rows": status_rows,
        "tat_rows": tat_rows,
        "workload_rows": workload_rows,
        "breadcrumbs": [
            {
                "title": "WorkShop/Floor",
                "url": "",
                "icon": "fa fa-list",
            },
            {
                "title": "Control Board",
                "icon": "fa fa-table",
            },
        ],
    })


@never_cache
@login_required
def bodyshop_control_board(request):
    logged_emp = Employee.objects.filter(user=request.user).first()

    if not (
        request.user.is_superuser
        or is_admin_or_manager_user(request.user, logged_emp)
        or is_floor_supervisor(logged_emp)
    ):
        messages.error(request, "You are not allowed to access Bodyshop Control Board")
        return redirect("dashboard")

    search = (request.GET.get("search") or "").strip()
    status_filter = request.GET.get("status") or "open"
    insurance_filter = request.GET.get("insurance") or ""
    advisor_filter = request.GET.get("advisor") or ""
    surveyor_filter = request.GET.get("surveyor") or ""

    allocations = (
        WorkAllocation.objects
        .select_related(
            "job",
            "job__claim",
            "job__claim__vehicle",
            "job__claim__vehicle__customer",
            "job__claim__vehicle__model",
            "job__claim__insurance_company",
            "job__claim__surveyor",
            "job__advisor",
        )
        .prefetch_related(
            "progress",
            "parts",
            "parts__job_part",
            "job__claim__timeline",
        )
        .order_by("-allotment_date", "-job__id")
    )
    unallocated_jobs = (
        JobCard.objects
        .select_related(
            "claim",
            "claim__vehicle",
            "claim__vehicle__customer",
            "claim__vehicle__model",
            "claim__insurance_company",
            "claim__surveyor",
            "advisor",
        )
        .filter(allocation__isnull=True)
        .order_by("-id")
    )

    if status_filter == "open":
        allocations = allocations.filter(
            job__claim__claim_stage__lt=ClaimStageCode.CLOSED
        ).exclude(job__repair_status="Closed")
        unallocated_jobs = unallocated_jobs.filter(
            claim__claim_stage__lt=ClaimStageCode.CLOSED
        ).exclude(repair_status="Closed")
    elif status_filter == "completed":
        allocations = allocations.filter(job__repair_status="Completed")
        unallocated_jobs = unallocated_jobs.filter(repair_status="Completed")
    elif status_filter == "closed":
        allocations = allocations.filter(
            Q(job__claim__claim_stage=ClaimStageCode.CLOSED)
            | Q(job__repair_status="Closed")
        )
        unallocated_jobs = unallocated_jobs.filter(
            Q(claim__claim_stage=ClaimStageCode.CLOSED)
            | Q(repair_status="Closed")
        )

    if search:
        allocations = allocations.filter(
            Q(job__job_no__icontains=search)
            | Q(job__claim__claim_no__icontains=search)
            | Q(job__claim__vehicle__registration_no__icontains=search)
            | Q(job__claim__vehicle__customer__name__icontains=search)
            | Q(job__claim__vehicle__model__name__icontains=search)
        )
        unallocated_jobs = unallocated_jobs.filter(
            Q(job_no__icontains=search)
            | Q(claim__claim_no__icontains=search)
            | Q(claim__vehicle__registration_no__icontains=search)
            | Q(claim__vehicle__customer__name__icontains=search)
            | Q(claim__vehicle__model__name__icontains=search)
        )

    if insurance_filter:
        allocations = allocations.filter(job__claim__insurance_company_id=insurance_filter)
        unallocated_jobs = unallocated_jobs.filter(claim__insurance_company_id=insurance_filter)

    if advisor_filter:
        allocations = allocations.filter(job__advisor_id=advisor_filter)
        unallocated_jobs = unallocated_jobs.filter(advisor_id=advisor_filter)

    if surveyor_filter:
        allocations = allocations.filter(job__claim__surveyor_id=surveyor_filter)
        unallocated_jobs = unallocated_jobs.filter(claim__surveyor_id=surveyor_filter)

    rows = []
    pna_count = 0
    work_completed_count = 0
    running_count = 0

    for allocation in allocations:
        status_text, status_class = get_control_board_current_status(allocation)
        pna_parts = get_control_board_pna_parts(allocation)
        current_stage, board_status, board_status_class = (
            get_control_board_stage_and_status(allocation, pna_parts)
        )

        if pna_parts:
            pna_count += 1
        if status_text == "Work Completed":
            work_completed_count += 1
        if "Running" in status_text:
            running_count += 1

        rows.append({
            "allocation": allocation,
            "job": allocation.job,
            "allocated_at": get_control_board_allocated_at(allocation),
            "in_date": get_control_board_in_date(allocation.job),
            "promise_date": get_control_board_promise_date(allocation.job),
            "promise_class": get_control_board_promise_class(allocation.job),
            "tat_days": get_control_board_tat_days(allocation.job),
            "tat_timeline": get_control_board_tat_timeline(allocation.job),
            "current_stage": current_stage,
            "claim_stage": allocation.job.claim.get_claim_stage_display(),
            "progress_status": get_control_board_progress_status(allocation),
            "board_status": board_status,
            "board_status_class": board_status_class,
            "status_text": status_text,
            "status_class": status_class,
            "pna_parts": pna_parts,
        })

    for job in unallocated_jobs:
        rows.append({
            "allocation": None,
            "job": job,
            "allocated_at": None,
            "in_date": get_control_board_in_date(job),
            "promise_date": get_control_board_promise_date(job),
            "promise_class": get_control_board_promise_class(job),
            "tat_days": get_control_board_tat_days(job),
            "tat_timeline": get_control_board_tat_timeline(job),
            "current_stage": job.claim.get_claim_stage_display(),
            "claim_stage": job.claim.get_claim_stage_display(),
            "progress_status": "Work Allocation Pending",
            "board_status": "🟡 Pending",
            "board_status_class": "status-pending",
            "status_text": "Work Allocation Pending",
            "status_class": "bg-warning text-dark",
            "pna_parts": [],
        })

    return render(request, "floor/bodyshopControlBoard.html", {
        "rows": rows,
        "search": search,
        "status_filter": status_filter,
        "insurance_filter": insurance_filter,
        "advisor_filter": advisor_filter,
        "surveyor_filter": surveyor_filter,
        "insurance_options": InsuranceCompany.objects.all().order_by("ins_co_name"),
        "advisor_options": Employee.objects.filter(
            is_active=True
        ).filter(
            Q(employee_type__iexact="Advisor")
            | Q(designation__iexact="Advisor")
        ).order_by("name"),
        "surveyor_options": Surveyor.objects.all().order_by("name"),
        "total_count": len(rows),
        "pna_count": pna_count,
        "work_completed_count": work_completed_count,
        "running_count": running_count,
        "breadcrumbs": [
            {
                "title": "WorkShop/Floor",
                "url": "",
                "icon": "fa fa-list",
            },
            {
                "title": "Bodyshop Control Board",
                "icon": "fa fa-table",
            },
        ],
    })


@never_cache
@login_required
def work_allocation_entry(request, job_id):
    from django.utils.dateparse import parse_date, parse_datetime

    job = get_object_or_404(
        JobCard,
        id=job_id
    )

    try:
        RepairWorkflowService.ensure_allocation_allowed(job)
    except RepairWorkflowBlocked as exc:
        messages.error(request, str(exc))
        return redirect("jobList")

    allocation, created = (
        WorkAllocation
        .objects
        .get_or_create(
            job=job
        )
    )
    progress_qs = allocation.progress.all().prefetch_related("photos")
    progress_map = {
        p.id: p
        for p in progress_qs
    }
    existing_reinspection_photo_count = job.reinspection_photos.count()
    existing_reinspection_photo_size = get_reinspection_photo_storage_size(job)

    if request.method == "POST":
        had_allocated_progress = any(
            progress.stage or progress.employee_id
            for progress in progress_qs
        )
        old_progress_state = {
            stage: {
                "start_time": progress.start_time,
                "finish_time": progress.finish_time,
                "employee_id": progress.employee_id,
                "has_data": progress_row_has_data(progress),
            }
            for stage, progress in progress_map.items()
        }
        uploaded_reinspection_images = request.FILES.getlist("reinspection_images")

        if uploaded_reinspection_images:
            total_photo_count = existing_reinspection_photo_count + len(uploaded_reinspection_images)

            if total_photo_count > REINSPECTION_MAX_PHOTOS_PER_JOBCARD:
                messages.error(
                    request,
                    "Re-inspection image limit exceeded. "
                    f"Maximum {REINSPECTION_MAX_PHOTOS_PER_JOBCARD} images are allowed per jobcard."
                )
                return redirect("work_allocation_entry", job_id=job.id)

            oversized_image = next(
                (
                    image for image in uploaded_reinspection_images
                    if image.size > REINSPECTION_MAX_IMAGE_SIZE_BYTES
                ),
                None
            )

            if oversized_image:
                messages.error(
                    request,
                    f"{oversized_image.name} is too large. "
                    f"Maximum {REINSPECTION_MAX_IMAGE_SIZE_MB} MB is allowed per image."
                )
                return redirect("work_allocation_entry", job_id=job.id)

            upload_total_size = sum(image.size for image in uploaded_reinspection_images)
            total_storage_size = existing_reinspection_photo_size + upload_total_size

            if total_storage_size > REINSPECTION_MAX_TOTAL_SIZE_BYTES:
                messages.error(
                    request,
                    "Re-inspection image storage limit exceeded. "
                    f"Maximum {REINSPECTION_MAX_TOTAL_SIZE_MB} MB is allowed per jobcard."
                )
                return redirect("work_allocation_entry", job_id=job.id)

        posted_work_completed = request.POST.get("mark_work_completed") == "1"
        posted_qc_done = request.POST.get("mark_qc_done") == "1"
        posted_reinspection_done = request.POST.get("reinspection_done") == "1"
        posted_part_entry_complete = request.POST.get("part_entry_complete") == "1"
        posted_reinspection_date = parse_workflow_datetime(
            request.POST.get("reinspection_date") or ""
        )
        posted_reinspection_done_by = request.POST.get(
            "reinspection_done_by",
            ""
        ).strip()
        posted_allotment_date = (
            parse_workflow_datetime(request.POST.get("allotment_date") or "")
            or allocation.allotment_date
        )
        validate_labels = set()
        if workflow_date_changed(allocation.allotment_date, posted_allotment_date):
            validate_labels.add("Work Allocation Date")
        if workflow_date_changed(job.reinspection_date, posted_reinspection_date):
            validate_labels.add("Re-Inspection Date")

        date_error = validate_workflow_dates([
            ("Gate In Date", job.gate_in_datetime),
            ("Claim Created Date", job.claim.created_at if job.claim else None),
            ("Jobcard Created Date", job.job_date or job.created_at),
            ("Claim Intimation Date", job.claim.intimation_date if job.claim else None),
            ("Survey Date", job.claim.survey_date if job.claim else None),
            ("Insurance Approval Date", job.claim.insurance_approval_date if job.claim else None),
            ("Work Allocation Date", posted_allotment_date),
            ("Re-Inspection Date", posted_reinspection_date),
            ("Liability Received Date", job.claim.liability_received_at if job.claim else None),
            ("Invoice Date", job.claim.invoice_datetime if job.claim else None),
            ("Delivery Date", job.claim.delivery_datetime if job.claim else None),
        ], validate_labels=validate_labels)
        if date_error:
            messages.error(request, date_error)
            return redirect("work_allocation_entry", job_id=job.id)

        future_error = validate_no_future_workflow_dates([
            ("Work Allocation Date", posted_allotment_date),
            ("Re-Inspection Date", posted_reinspection_date),
        ])
        if future_error:
            messages.error(request, future_error)
            return redirect("work_allocation_entry", job_id=job.id)

        if posted_work_completed and not (
            posted_qc_done
            and posted_reinspection_done
            and posted_part_entry_complete
        ):
            messages.error(
                request,
                "Tick QC Done, Re-Inspection and Part Entry before Work Completed."
            )
            return redirect("work_allocation_entry", job_id=job.id)

        if (
            posted_work_completed
            and job.additional_approval_required
            and (job.second_approval_status or "Pending") == "Pending"
        ):
            messages.error(
                request,
                "2nd Approval is pending. Advisor approval is required before Work Completed."
            )
            return redirect("work_allocation_entry", job_id=job.id)

        allocation.allotment_date = posted_allotment_date
        allocation.delivery_date = parse_date(
            request.POST.get("delivery_date") or ""
        )
        allocation.parts_slip_no = request.POST.get(
            "parts_slip_no",
            ""
        ).strip()
        allocation.remarks = request.POST.get(
            "remarks",
            ""
        ).strip()
        allocation.part_entry_complete = posted_part_entry_complete
        allocation.save()
        additional_approval_reasons = []

        progress_ids = request.POST.getlist("progress_id[]")
        progress_photo_keys = request.POST.getlist("progress_photo_key[]")
        stages = request.POST.getlist("stage[]")
        start_times = request.POST.getlist("start_time[]")
        finish_times = request.POST.getlist("finish_time[]")
        employee_ids = request.POST.getlist("employee[]")
        progress_remarks = request.POST.getlist("progress_remarks[]")
        clear_progress_photo_stages = set(
            request.POST.getlist("clear_progress_photo_stage[]")
        )
        clear_progress_photo_ids = set(
            request.POST.getlist("clear_progress_photo_id[]")
        )
        delete_progress_ids = {
            value
            for value in request.POST.getlist("delete_progress_id[]")
            if value.isdigit()
        }

        if delete_progress_ids:
            WorkProgress.objects.filter(
                allocation=allocation,
                id__in=delete_progress_ids,
            ).delete()
            progress_map = {
                key: value
                for key, value in progress_map.items()
                if str(key) not in delete_progress_ids
            }

        for index, start_time in enumerate(start_times):
            employee_id = (
                employee_ids[index]
                if index < len(employee_ids)
                else ""
            )

            if start_time and not employee_id:
                stage_key = stages[index] if index < len(stages) else ""
                stage_label = dict(WorkProgress.STAGES).get(
                    stage_key,
                    stage_key or "progress row"
                )
                messages.error(
                    request,
                    f"Select employee for {stage_label} before saving."
                )
                return redirect("work_allocation_entry", job_id=job.id)

        if posted_work_completed:
            unfinished_stages = []
            for index, start_time in enumerate(start_times):
                finish_time = (
                    finish_times[index]
                    if index < len(finish_times)
                    else ""
                )
                if start_time and not finish_time:
                    stage_key = (
                        stages[index]
                        if index < len(stages)
                        else ""
                    )
                    unfinished_stages.append(
                        dict(WorkProgress.STAGES).get(
                            stage_key,
                            stage_key or "progress row"
                        )
                    )

            if unfinished_stages:
                messages.error(
                    request,
                    "Finish started progress before Work Completed: "
                    + ", ".join(unfinished_stages)
                )
                return redirect("work_allocation_entry", job_id=job.id)

            posted_allocation_labour_ids = request.POST.getlist("allocation_labour_id[]")
            posted_labour_employee_ids = request.POST.getlist("labour_employee[]")
            missing_labour_employee_labels = []

            for index, labour_id in enumerate(posted_allocation_labour_ids):
                employee_id = (
                    posted_labour_employee_ids[index]
                    if index < len(posted_labour_employee_ids)
                    else ""
                )

                if employee_id:
                    continue

                labour = JobCardLabour.objects.filter(
                    id=labour_id,
                    job=job,
                ).first()
                missing_labour_employee_labels.append(
                    labour.job_code if labour else f"line {index + 1}"
                )

            if missing_labour_employee_labels:
                messages.error(
                    request,
                    "Select employee for labour line before Work Completed: "
                    + ", ".join(missing_labour_employee_labels)
                )
                return redirect("work_allocation_entry", job_id=job.id)

        for index, stage in enumerate(stages):
            progress_id = (
                progress_ids[index]
                if index < len(progress_ids)
                else ""
            )
            start_time = (
                start_times[index]
                if index < len(start_times)
                else ""
            )
            finish_time = (
                finish_times[index]
                if index < len(finish_times)
                else ""
            )
            employee_id = (
                employee_ids[index]
                if index < len(employee_ids)
                else ""
            )
            remarks = (
                progress_remarks[index].strip()
                if index < len(progress_remarks)
                else ""
            )
            has_uploaded_photos = bool(
                request.FILES.getlist(
                    progress_photo_input_name(
                        progress_photo_keys[index]
                        if index < len(progress_photo_keys)
                        else progress_id
                    )
                )
            )

            if not stage and not any([
                start_time,
                finish_time,
                employee_id,
                remarks,
                has_uploaded_photos,
            ]):
                continue

            if not stage:
                messages.error(
                    request,
                    "Select Progress Type before saving progress row."
                )
                return redirect("work_allocation_entry", job_id=job.id)

            progress = None
            if progress_id:
                progress = progress_map.get(int(progress_id)) if progress_id.isdigit() else None

                if progress is None:
                    messages.error(request, "Invalid progress row.")
                    return redirect("work_allocation_entry", job_id=job.id)

            if progress is None:
                progress = WorkProgress.objects.create(
                    allocation=allocation,
                    stage=stage
                )

            progress.stage = stage
            progress.start_time = parse_datetime(start_time) if start_time else None
            progress.finish_time = parse_datetime(finish_time) if finish_time else None
            progress.employee_id = employee_id or None
            progress.remarks = remarks
            progress.save()

            if (
                stage in clear_progress_photo_stages
                or str(progress.id) in clear_progress_photo_ids
            ):
                for photo in progress.photos.all():
                    if photo.image:
                        photo.image.delete(save=False)
                    photo.delete()

            for image in request.FILES.getlist(
                progress_photo_input_name(
                    progress_photo_keys[index]
                    if index < len(progress_photo_keys)
                    else progress_id
                )
            ):
                WorkProgressPhoto.objects.create(
                    progress=progress,
                    image=image
                )

            old_state = old_progress_state.get(progress.id, {})
            if old_state.get("has_data") and not progress_row_has_data(progress):
                notify_work_progress_change(progress, "cleared")
            elif not old_state.get("start_time") and progress.start_time:
                notify_work_progress_change(progress, "started")
            elif not old_state.get("finish_time") and progress.finish_time:
                notify_work_progress_change(progress, "finished")
            elif old_state.get("employee_id") != progress.employee_id and progress.employee_id:
                notify_work_progress_change(progress, "assigned")

            if old_state.get("employee_id") != progress.employee_id and progress.employee_id:
                notify_progress_employee_assigned(progress)

        sync_jobcard_main_status(job)

        has_allocated_progress = WorkProgress.objects.filter(
            allocation=allocation,
        ).filter(
            Q(stage__gt="")
            | Q(employee__isnull=False)
        ).exists()

        if not had_allocated_progress and has_allocated_progress:
            notify_floor_incharge_work_allocated(job)

        has_progress_started = WorkProgress.objects.filter(
            allocation=allocation,
            start_time__isnull=False,
        ).exists()
        has_any_progress_data = any(
            progress_row_has_data(row)
            for row in allocation.progress.all()
        )

        if not has_any_progress_data and job.claim:
            if int(job.claim.claim_stage or 0) in [
                ClaimStageCode.REPAIR_IN_PROGRESS,
                ClaimStageCode.WORK_COMPLETED,
                ClaimStageCode.RE_INSPECTION,
            ]:
                job.claim.claim_stage = ClaimStageCode.WORK_ALLOCATION
                job.claim.save(update_fields=["claim_stage"])

            job.qc_done = False
            job.reinspection_done = False
            job.reinspection_date = None
            job.reinspection_done_by = ""
            job.save(update_fields=[
                "qc_done",
                "reinspection_done",
                "reinspection_date",
                "reinspection_done_by",
            ])

            allocation.part_entry_complete = False
            allocation.save(update_fields=["part_entry_complete"])

        if (
            has_progress_started
            and job.claim
            and int(job.claim.claim_stage or 0) < ClaimStageCode.WORK_COMPLETED
        ):
            job.claim.claim_stage = ClaimStageCode.REPAIR_IN_PROGRESS
            job.claim.save(update_fields=["claim_stage"])

        allocation_part_ids = request.POST.getlist("allocation_part_id[]")
        decisions = request.POST.getlist("decision[]")
        pick_from_store = request.POST.getlist("pick_from_store[]")
        pick_dates = request.POST.getlist("pick_date[]")
        picker_names = request.POST.getlist("picker_name[]")
        ko_order_dates = request.POST.getlist("ko_order_date[]")
        ko_order_nos = request.POST.getlist("ko_order_no[]")
        etas = request.POST.getlist("eta[]")
        part_remarks = request.POST.getlist("part_remarks[]")

        for index, assessment_id in enumerate(allocation_part_ids):
            assessment = JobCardAssessmentPart.objects.filter(
                id=assessment_id,
                job=job,
            ).select_related("part").first()

            if not assessment:
                continue

            posted_decision = (
                decisions[index]
                if index < len(decisions)
                else assessment.decision
            )

            current_allocation_part = WorkAllocationPart.objects.filter(
                allocation=allocation,
                job_part=assessment.part,
            ).first()
            current_decision = (
                current_allocation_part.decision
                if current_allocation_part
                else assessment.decision
            )

            if current_decision == "KO" and posted_decision in ["New", "Repair"]:
                part_label = (
                    assessment.part.part_no
                    if assessment.part
                    else f"part line {index + 1}"
                )
                additional_approval_reasons.append(
                    f"KO part changed to {posted_decision}: {part_label}"
                )

            allocation_part, _ = WorkAllocationPart.objects.get_or_create(
                allocation=allocation,
                job_part=assessment.part,
                defaults={
                    "decision": assessment.decision,
                }
            )
            allocation_part.decision = posted_decision
            allocation_part.pick_from_store = (
                index < len(pick_from_store)
                and pick_from_store[index] == "Yes"
            )
            allocation_part.pick_date = parse_date(
                pick_dates[index]
            ) if index < len(pick_dates) and pick_dates[index] else None
            allocation_part.picker_name = (
                picker_names[index].strip()
                if index < len(picker_names)
                else ""
            )
            allocation_part.ko_order_date = parse_date(
                ko_order_dates[index]
            ) if index < len(ko_order_dates) and ko_order_dates[index] else None
            allocation_part.ko_order_no = (
                ko_order_nos[index].strip()
                if index < len(ko_order_nos)
                else ""
            )
            allocation_part.eta = parse_date(
                etas[index]
            ) if index < len(etas) and etas[index] else None
            allocation_part.remarks = (
                part_remarks[index].strip()
                if index < len(part_remarks)
                else ""
            )
            allocation_part.save()

        new_part_nos = request.POST.getlist("new_part_no[]")
        new_part_descriptions = request.POST.getlist("new_part_description[]")
        new_part_qtys = request.POST.getlist("new_part_qty[]")
        new_part_rates = request.POST.getlist("new_part_rate[]")
        new_part_decisions = request.POST.getlist("new_part_decision[]")
        new_part_pick_from_store = request.POST.getlist("new_part_pick_from_store[]")
        new_part_pick_dates = request.POST.getlist("new_part_pick_date[]")
        new_part_picker_names = request.POST.getlist("new_part_picker_name[]")
        new_part_ko_order_dates = request.POST.getlist("new_part_ko_order_date[]")
        new_part_ko_order_nos = request.POST.getlist("new_part_ko_order_no[]")
        new_part_etas = request.POST.getlist("new_part_eta[]")
        new_part_remarks = request.POST.getlist("new_part_remarks[]")
        new_part_photo_keys = request.POST.getlist("new_part_photo_key[]")
        has_new_part_lines = any(
            (part_no or "").strip()
            or (
                new_part_descriptions[index].strip()
                if index < len(new_part_descriptions)
                else ""
            )
            for index, part_no in enumerate(new_part_nos)
        )

        if has_new_part_lines:
            for index, part_no in enumerate(new_part_nos):
                description = (
                    new_part_descriptions[index].strip()
                    if index < len(new_part_descriptions)
                    else ""
                )

                if not (part_no or "").strip() and not description:
                    continue

                photo_key = (
                    new_part_photo_keys[index]
                    if index < len(new_part_photo_keys)
                    else str(index)
                )

                if not request.FILES.getlist(f"new_part_photo_{photo_key}"):
                    messages.error(
                        request,
                        "Upload photo(s) on each additional part line."
                    )
                    return redirect("work_allocation_entry", job_id=job.id)

        for index, part_no in enumerate(new_part_nos):
            part_no = part_no.strip()
            description = (
                new_part_descriptions[index].strip()
                if index < len(new_part_descriptions)
                else ""
            )

            if not part_no and not description:
                continue

            qty = int(
                Decimal(
                    new_part_qtys[index]
                    if index < len(new_part_qtys) and new_part_qtys[index]
                    else "1"
                )
            )
            rate = Decimal(
                new_part_rates[index]
                if index < len(new_part_rates) and new_part_rates[index]
                else "0"
            )
            decision = (
                new_part_decisions[index]
                if index < len(new_part_decisions) and new_part_decisions[index]
                else "New"
            )

            job_part = JobCardPart.objects.create(
                job=job,
                part_no=part_no or "Additional",
                description=description or part_no or "Additional part",
                qty=qty,
                rate=rate,
            )
            JobCardAssessmentPart.objects.create(
                job=job,
                part=job_part,
                decision=decision,
                revised_amount=job_part.amount,
            )
            allocation_part = WorkAllocationPart.objects.create(
                allocation=allocation,
                job_part=job_part,
                decision=decision,
                is_additional=True,
                advisor_approval_status="Pending",
                pick_from_store=(
                    index < len(new_part_pick_from_store)
                    and new_part_pick_from_store[index] == "Yes"
                ),
                pick_date=parse_date(
                    new_part_pick_dates[index]
                ) if index < len(new_part_pick_dates) and new_part_pick_dates[index] else None,
                picker_name=(
                    new_part_picker_names[index].strip()
                    if index < len(new_part_picker_names)
                    else ""
                ),
                ko_order_date=parse_date(
                    new_part_ko_order_dates[index]
                ) if index < len(new_part_ko_order_dates) and new_part_ko_order_dates[index] else None,
                ko_order_no=(
                    new_part_ko_order_nos[index].strip()
                    if index < len(new_part_ko_order_nos)
                    else ""
                ),
                eta=parse_date(
                    new_part_etas[index]
                ) if index < len(new_part_etas) and new_part_etas[index] else None,
                remarks=(
                    new_part_remarks[index].strip()
                    if index < len(new_part_remarks)
                    else ""
                ),
            )
            photo_key = (
                new_part_photo_keys[index]
                if index < len(new_part_photo_keys)
                else str(index)
            )
            for image in request.FILES.getlist(f"new_part_photo_{photo_key}"):
                JobCardAdditionalApprovalPhoto.objects.create(
                    job=job,
                    work_allocation_part=allocation_part,
                    image=image,
                )
            additional_approval_reasons.append(
                f"Additional part added: {allocation_part.job_part.part_no}"
            )

        allocation_labour_ids = request.POST.getlist("allocation_labour_id[]")
        labour_decisions = request.POST.getlist("labour_decision[]")
        labour_revised_amounts = request.POST.getlist("labour_revised_amount[]")
        labour_employee_ids = request.POST.getlist("labour_employee[]")
        labour_remarks = request.POST.getlist("labour_remarks[]")

        for index, labour_id in enumerate(allocation_labour_ids):
            labour = JobCardLabour.objects.filter(
                id=labour_id,
                job=job,
            ).first()

            if not labour:
                continue

            allocation_labour, _ = WorkAllocationLabour.objects.get_or_create(
                allocation=allocation,
                job_labour=labour,
                defaults={
                    "revised_amount": labour.amount,
                }
            )
            allocation_labour.decision = (
                labour_decisions[index]
                if index < len(labour_decisions)
                else "Approved"
            )
            allocation_labour.revised_amount = Decimal(
                labour_revised_amounts[index]
                if index < len(labour_revised_amounts)
                and labour_revised_amounts[index]
                else "0"
            )
            employee_id = (
                labour_employee_ids[index]
                if index < len(labour_employee_ids)
                else ""
            )
            allocation_labour.employee_id = employee_id or None
            allocation_labour.remarks = (
                labour_remarks[index].strip()
                if index < len(labour_remarks)
                else ""
            )
            allocation_labour.save()

        new_labour_codes = request.POST.getlist("new_labour_code[]")
        new_labour_descriptions = request.POST.getlist("new_labour_description[]")
        new_labour_amounts = request.POST.getlist("new_labour_amount[]")
        new_labour_hours = request.POST.getlist("new_labour_hrs[]")
        new_labour_rates = request.POST.getlist("new_labour_rate[]")
        new_labour_decisions = request.POST.getlist("new_labour_decision[]")
        new_labour_employees = request.POST.getlist("new_labour_employee[]")
        new_labour_remarks = request.POST.getlist("new_labour_remarks[]")
        new_labour_photo_keys = request.POST.getlist("new_labour_photo_key[]")
        has_new_labour_lines = any(
            (labour_code or "").strip()
            or (
                new_labour_descriptions[index].strip()
                if index < len(new_labour_descriptions)
                else ""
            )
            for index, labour_code in enumerate(new_labour_codes)
        )

        if has_new_labour_lines:
            for index, labour_code in enumerate(new_labour_codes):
                description = (
                    new_labour_descriptions[index].strip()
                    if index < len(new_labour_descriptions)
                    else ""
                )

                if not (labour_code or "").strip() and not description:
                    continue

                photo_key = (
                    new_labour_photo_keys[index]
                    if index < len(new_labour_photo_keys)
                    else str(index)
                )

                if not request.FILES.getlist(f"new_labour_photo_{photo_key}"):
                    messages.error(
                        request,
                        "Upload photo(s) on each additional labour line."
                    )
                    return redirect("work_allocation_entry", job_id=job.id)

        for index, labour_code in enumerate(new_labour_codes):
            labour_code = labour_code.strip()
            description = (
                new_labour_descriptions[index].strip()
                if index < len(new_labour_descriptions)
                else ""
            )

            if not labour_code and not description:
                continue

            if index < len(new_labour_amounts) and new_labour_amounts[index]:
                labour_hrs = Decimal("1")
                labour_rate = Decimal(new_labour_amounts[index])
            else:
                labour_hrs = Decimal(
                    new_labour_hours[index]
                    if index < len(new_labour_hours) and new_labour_hours[index]
                    else "1"
                )
                labour_rate = Decimal(
                    new_labour_rates[index]
                    if index < len(new_labour_rates) and new_labour_rates[index]
                    else "0"
                )

            if labour_hrs >= Decimal("1000"):
                messages.error(
                    request,
                    "Labour hours must be less than 1000."
                )
                return redirect("work_allocation_entry", job_id=job.id)
            decision = (
                new_labour_decisions[index]
                if index < len(new_labour_decisions) and new_labour_decisions[index]
                else "Approved"
            )
            employee_id = (
                new_labour_employees[index]
                if index < len(new_labour_employees)
                else ""
            )

            job_labour = JobCardLabour.objects.create(
                job=job,
                job_code=labour_code or "Additional",
                description=description or labour_code or "Additional labour",
                labour_hrs=labour_hrs,
                rate=labour_rate,
            )
            JobCardAssessmentLabour.objects.create(
                job=job,
                labour=job_labour,
                decision=decision,
                revised_amount=job_labour.amount,
            )
            allocation_labour = WorkAllocationLabour.objects.create(
                allocation=allocation,
                job_labour=job_labour,
                decision=decision,
                is_additional=True,
                advisor_approval_status="Pending",
                revised_amount=job_labour.amount,
                employee_id=employee_id or None,
                remarks=(
                    new_labour_remarks[index].strip()
                    if index < len(new_labour_remarks)
                    else ""
                ),
            )
            photo_key = (
                new_labour_photo_keys[index]
                if index < len(new_labour_photo_keys)
                else str(index)
            )
            for image in request.FILES.getlist(f"new_labour_photo_{photo_key}"):
                JobCardAdditionalApprovalPhoto.objects.create(
                    job=job,
                    work_allocation_labour=allocation_labour,
                    image=image,
                )
            additional_approval_reasons.append(
                f"Additional labour added: {job_labour.job_code}"
            )

        if additional_approval_reasons:
            existing_reason = (job.additional_approval_reason or "").strip()
            reason_text = "\n".join(additional_approval_reasons)
            job.additional_approval_required = True
            job.second_approval_status = "Pending"
            job.repair_status = "Open"
            job.additional_approval_reason = (
                f"{existing_reason}\n{reason_text}".strip()
                if existing_reason
                else reason_text
            )
            job.save(update_fields=[
                "additional_approval_required",
                "second_approval_status",
                "repair_status",
                "additional_approval_reason",
            ])
            posted_work_completed = False
            posted_qc_done = False
            posted_reinspection_done = False
            posted_part_entry_complete = False
            allocation.part_entry_complete = False
            allocation.save(update_fields=["part_entry_complete"])

            if job.claim:
                job.claim.claim_stage = ClaimStageCode.REPAIR_IN_PROGRESS
                job.claim.save(update_fields=["claim_stage"])

            notify_jobcard_advisor(
                job,
                "2nd Approval Required",
                f"Jobcard {job.job_no} needs Advisor approval for additional work.",
            )

        if posted_work_completed:
            job.repair_status = "Completed"
            job.save(update_fields=["repair_status"])

            if job.claim:
                job.claim.claim_stage = ClaimStageCode.WORK_COMPLETED
                job.claim.save(update_fields=["claim_stage"])

            notify_jobcard_advisor(
                job,
                "Jobcard Work Completed",
                f"Jobcard {job.job_no} repair work completed",
            )
        elif (
            job.claim
            and int(job.claim.claim_stage or 0) == ClaimStageCode.WORK_COMPLETED
        ):
            job.claim.claim_stage = (
                ClaimStageCode.REPAIR_IN_PROGRESS
                if has_progress_started
                else ClaimStageCode.WORK_ALLOCATION
            )
            job.claim.save(update_fields=["claim_stage"])
            sync_jobcard_main_status(job)

        job.qc_done = posted_qc_done
        job.save(update_fields=["qc_done"])

        job.reinspection_done = posted_reinspection_done
        job.reinspection_date = posted_reinspection_date
        job.reinspection_done_by = posted_reinspection_done_by
        job.save(update_fields=[
            "reinspection_done",
            "reinspection_date",
            "reinspection_done_by",
        ])

        for image in uploaded_reinspection_images:
            JobCardReInspectionPhoto.objects.create(
                job=job,
                image=image
            )

        if job.reinspection_done and job.claim:
            job.claim.claim_stage = ClaimStageCode.LIABILITY
            job.claim.save(update_fields=["claim_stage"])

            notify_jobcard_advisor(
                job,
                "Jobcard Re-Inspection Done",
                f"Jobcard {job.job_no} re-inspection completed",
            )

        messages.success(request, "Work allocation saved successfully")
        return redirect("work_allocation_entry", job_id=job.id)

    progress_rows = []

    for p in progress_qs:
        if not progress_row_has_data(p):
            continue

        progress_rows.append({
            "id": p.id,
            "stage": p.stage,
            "label": p.get_stage_display(),
            "start_time": p.start_time if p else None,
            "finish_time": p.finish_time if p else None,
            "employee_id": p.employee_id if p else None,
            "remarks": p.remarks if p else "",
            "progress_id": p.id if p else "",
            "photo_count": p.photos.count() if p else 0,
            "photo_input_name": progress_photo_input_name(p.id),
        })
    assessed_parts = list(JobCardAssessmentPart.objects.filter(
        job=job,
        decision__in=["New", "Repair", "KO"]
    ).select_related("part"))

    allocation_parts_by_part_id = {
        part.job_part_id: part
        for part in allocation.parts.all()
    }

    for assessment in assessed_parts:
        saved_part = allocation_parts_by_part_id.get(assessment.part_id)

        if saved_part:
            assessment.decision = saved_part.decision
            assessment.pick_from_store = saved_part.pick_from_store
            assessment.pick_date = saved_part.pick_date
            assessment.picker_name = saved_part.picker_name
            assessment.ko_order_date = saved_part.ko_order_date
            assessment.ko_order_no = saved_part.ko_order_no
            assessment.eta = saved_part.eta
            assessment.remarks = saved_part.remarks
            assessment.is_additional = saved_part.is_additional
            assessment.advisor_approval_status = saved_part.advisor_approval_status
            assessment.additional_approval_photos = list(
                saved_part.additional_approval_photos.all()
            )
        else:
            assessment.pick_from_store = False
            assessment.pick_date = None
            assessment.picker_name = ""
            assessment.ko_order_date = None
            assessment.ko_order_no = ""
            assessment.eta = None
            assessment.remarks = ""
            assessment.is_additional = False
            assessment.advisor_approval_status = ""
            assessment.additional_approval_photos = []

    assessed_labours = list(job.labours.filter(
        jobcardassessmentlabour__decision="Approved"
    ))
    allocation_labours_by_labour_id = {
        labour.job_labour_id: labour
        for labour in allocation.labours.all()
    }

    for labour in assessed_labours:
        saved_labour = allocation_labours_by_labour_id.get(labour.id)
        assessment = JobCardAssessmentLabour.objects.filter(
            job=job,
            labour=labour,
        ).first()

        labour.decision = (
            saved_labour.decision
            if saved_labour
            else "Approved"
        )
        labour.revised_amount = (
            saved_labour.revised_amount
            if saved_labour
            else (
                assessment.revised_amount
                if assessment
                else labour.amount
            )
        )
        labour.employee_id = (
            saved_labour.employee_id
            if saved_labour
            else None
        )
        labour.remarks = (
            saved_labour.remarks
            if saved_labour
            else ""
        )
        labour.is_additional = saved_labour.is_additional if saved_labour else False
        labour.advisor_approval_status = (
            saved_labour.advisor_approval_status
            if saved_labour
            else ""
        )
        labour.additional_approval_photos = (
            list(saved_labour.additional_approval_photos.all())
            if saved_labour
            else []
        )
    new_paint_panel_count = sum(
        1 for labour in assessed_labours
        if labour.paint_panel_type == "New"
    )
    repair_paint_panel_count = sum(
        1 for labour in assessed_labours
        if labour.paint_panel_type == "Repair"
    )
    total_paint_panel_count = new_paint_panel_count + repair_paint_panel_count
    technicians = Employee.objects.filter(
        designation__in=['Technician', 'Denter', 'Painter']
    )
    return render(
        request,

        "floor/workAllocationEntry.html",

        {

            "job": job,
            "technicians": technicians,
            "allocation": allocation,
            "can_print_work_report": (
                (
                    job.repair_status == "Completed"
                    or (
                        job.claim
                        and int(job.claim.claim_stage or 0) >= ClaimStageCode.WORK_COMPLETED
                    )
                )
                and job.qc_done
                and job.reinspection_done
            ),
            "is_work_completed": (
                job.repair_status == "Completed"
                or (
                    job.claim
                    and int(job.claim.claim_stage or 0) >= ClaimStageCode.WORK_COMPLETED
                )
            ),
            "progress_rows": progress_rows,
            "rows": rows,
            "allocation_parts": assessed_parts,
            "existing_reinspection_photo_count": existing_reinspection_photo_count,
            "existing_reinspection_photo_size_mb": round(
                existing_reinspection_photo_size / (1024 * 1024),
                2
            ),
            "reinspection_max_photos": REINSPECTION_MAX_PHOTOS_PER_JOBCARD,
            "reinspection_max_image_size_mb": REINSPECTION_MAX_IMAGE_SIZE_MB,
            "reinspection_max_total_size_mb": REINSPECTION_MAX_TOTAL_SIZE_MB,

            "allocation_labours": assessed_labours,
            "new_paint_panel_count": new_paint_panel_count,
            "repair_paint_panel_count": repair_paint_panel_count,
            "total_paint_panel_count": total_paint_panel_count,

            "stages":
                WorkProgress.STAGES,

        }
    )


@never_cache
@login_required
def work_completion_report(request, job_id):
    job = get_object_or_404(
        JobCard.objects.select_related(
            "claim",
            "claim__vehicle",
            "claim__vehicle__customer",
            "claim__vehicle__model",
            "claim__vehicle__variant",
            "advisor",
            "allocation",
        ),
        id=job_id
    )
    allocation = getattr(job, "allocation", None)

    if not allocation:
        messages.error(request, "Work allocation is not available for this jobcard.")
        return redirect("work_allocation_entry", job_id=job.id)

    can_print = (
        (
            job.repair_status == "Completed"
            or (
                job.claim
                and int(job.claim.claim_stage or 0) >= ClaimStageCode.WORK_COMPLETED
            )
        )
        and job.qc_done
        and job.reinspection_done
    )

    if not can_print:
        messages.error(
            request,
            "Complete Work Completed, QC, and Re-Inspection before printing the report."
        )
        return redirect("work_allocation_entry", job_id=job.id)

    progress_rows = (
        allocation.progress
        .select_related("employee")
        .filter(finish_time__isnull=False)
        .order_by("finish_time", "id")
    )
    parts = allocation.parts.select_related("job_part").order_by("id")
    labours = allocation.labours.select_related("job_labour", "employee").order_by("id")

    return render(request, "floor/workCompletionReport.html", {
        "job": job,
        "claim": job.claim,
        "allocation": allocation,
        "progress_rows": progress_rows,
        "parts": parts,
        "labours": labours,
        "work_status": get_work_progress_status(allocation),
    })


@login_required
def check_work_allocation_employee(request):
    employee_id = request.GET.get("employee_id")
    allocation_id = request.GET.get("allocation_id")

    if not employee_id:
        return JsonResponse({"assigned": False})

    progress = (
        WorkProgress.objects
        .select_related(
            "allocation",
            "allocation__job",
            "allocation__job__claim",
            "allocation__job__claim__vehicle",
            "employee"
        )
        .filter(
            employee_id=employee_id,
            start_time__isnull=False,
            finish_time__isnull=True,
        )
    )

    if allocation_id:
        progress = progress.exclude(allocation_id=allocation_id)

    progress = progress.first()

    if not progress:
        return JsonResponse({"assigned": False})

    return JsonResponse({
        "assigned": True,
        "employee": progress.employee.name if progress.employee else "",
        "job_no": progress.allocation.job.job_no if progress.allocation and progress.allocation.job else "",
        "registration_no": progress.allocation.job.claim.vehicle.registration_no if (
            progress.allocation
            and progress.allocation.job
            and progress.allocation.job.claim
            and progress.allocation.job.claim.vehicle
        ) else "",
        "stage": progress.get_stage_display(),
    })


@never_cache
@login_required
def reinspection_photo_view(request, job_id):
    job = get_object_or_404(
        JobCard.objects.select_related(
            "claim",
            "claim__vehicle",
        ),
        id=job_id
    )

    if request.method == "POST":
        photo_ids = request.POST.getlist("photo_ids")
        photos_to_delete = job.reinspection_photos.filter(id__in=photo_ids)

        if not photos_to_delete.exists():
            messages.error(request, "Select at least one image to delete.")
            return redirect("reinspection_photo_view", job_id=job.id)

        deleted_count = 0

        for photo in photos_to_delete:
            if photo.image:
                photo.image.delete(save=False)

            photo.delete()
            deleted_count += 1

        messages.success(request, f"{deleted_count} re-inspection image(s) deleted successfully.")

        return redirect("reinspection_photo_view", job_id=job.id)

    photos = job.reinspection_photos.order_by("uploaded_at")

    return render(request, "floor/reinspectionPhotos.html", {
        "job": job,
        "photos": photos,
    })


@login_required
def download_reinspection_photos(request, job_id):
    job = get_object_or_404(JobCard, id=job_id)
    photo_ids = request.POST.getlist("photo_ids")
    photos = job.reinspection_photos.filter(id__in=photo_ids).order_by("uploaded_at")

    if not photos.exists():
        messages.error(request, "Select at least one image to download.")
        return redirect("reinspection_photo_view", job_id=job.id)

    buffer = BytesIO()

    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
        for index, photo in enumerate(photos, start=1):
            if not photo.image:
                continue

            filename = os.path.basename(photo.image.name)
            _, ext = os.path.splitext(filename)
            zip_name = f"reinspection_{index}{ext or '.jpg'}"

            photo.image.open("rb")
            zip_file.writestr(zip_name, photo.image.read())
            photo.image.close()

    buffer.seek(0)
    claim_no = job.claim.claim_no if job.claim else job.job_no
    safe_claim_no = "".join(
        char if char.isalnum() or char in ["-", "_"] else "_"
        for char in claim_no
    )
    response = HttpResponse(buffer.getvalue(), content_type="application/zip")
    response["Content-Disposition"] = (
        f'attachment; filename="reinspection_{safe_claim_no}.zip"'
    )

    return response


@login_required
def delete_reinspection_photo(request, job_id, photo_id):
    if request.method != "POST":
        return redirect("reinspection_photo_view", job_id=job_id)

    job = get_object_or_404(JobCard, id=job_id)
    photo = get_object_or_404(JobCardReInspectionPhoto, id=photo_id, job=job)

    if photo.image:
        photo.image.delete(save=False)

    photo.delete()
    messages.success(request, "Re-inspection image deleted successfully.")

    return redirect("reinspection_photo_view", job_id=job.id)


@never_cache
@login_required
def work_progress_photo_view(request, progress_id):
    progress = get_object_or_404(
        WorkProgress.objects.select_related(
            "allocation",
            "allocation__job",
            "allocation__job__claim",
            "allocation__job__claim__vehicle",
        ),
        id=progress_id
    )
    job = progress.allocation.job

    if request.method == "POST":
        photo_ids = request.POST.getlist("photo_ids")
        photos_to_delete = progress.photos.filter(id__in=photo_ids)

        if not photos_to_delete.exists():
            messages.error(request, "Select at least one image to delete.")
            return redirect("work_progress_photo_view", progress_id=progress.id)

        deleted_count = 0
        for photo in photos_to_delete:
            if photo.image:
                photo.image.delete(save=False)
            photo.delete()
            deleted_count += 1

        messages.success(request, f"{deleted_count} progress image(s) deleted successfully.")
        return redirect("work_progress_photo_view", progress_id=progress.id)

    return render(request, "floor/workProgressPhotos.html", {
        "job": job,
        "progress": progress,
        "photos": progress.photos.order_by("uploaded_at"),
    })


@never_cache
@login_required
def my_work_list(request):
    logged_emp = Employee.objects.filter(user=request.user).first()

    if not is_repair_resource(logged_emp):
        messages.error(request, "Only Technician, Denter or Painter can access My Work.")
        return redirect("dashboard")

    today = timezone.localdate()
    month_start = today.replace(day=1)
    from_date = parse_date(request.GET.get("from_date") or "") or month_start
    to_date = parse_date(request.GET.get("to_date") or "") or today
    status_filter = request.GET.get("status") or "new"

    base_progress = my_work_base_queryset(logged_emp, from_date, to_date)
    rows = [
        my_work_row_payload(progress)
        for progress in apply_my_work_status_filter(base_progress, status_filter).order_by(
            "start_time",
            "allocation__job__job_no",
            "id",
        )
    ]

    counts = {
        "new": base_progress.filter(start_time__isnull=True).count(),
        "wip": base_progress.filter(start_time__isnull=False, finish_time__isnull=True).count(),
        "completed": base_progress.filter(finish_time__isnull=False).count(),
    }

    return render(request, "floor/myWorkList.html", {
        "logged_emp": logged_emp,
        "rows": rows,
        "counts": counts,
        "status_filter": status_filter,
        "filter_from_date": from_date.strftime("%Y-%m-%d"),
        "filter_to_date": to_date.strftime("%Y-%m-%d"),
        "breadcrumbs": [
            {
                "title": "WorkShop / Floor",
                "url": "",
                "icon": "fa fa-list"
            },
            {
                "title": "My Work",
                "icon": "fa fa-tools"
            }
        ]
    })


@require_POST
@never_cache
@login_required
def my_work_action(request, progress_id):
    logged_emp = Employee.objects.filter(user=request.user).first()

    if not is_repair_resource(logged_emp):
        messages.error(request, "Only Technician, Denter or Painter can update My Work.")
        return redirect("dashboard")

    progress = get_object_or_404(
        WorkProgress.objects.select_related("allocation", "allocation__job"),
        id=progress_id,
        employee=logged_emp,
    )
    action = request.POST.get("action") or ""

    if action == "start":
        try:
            RepairWorkflowService.ensure_start_allowed(
                progress.allocation.job
            )
        except RepairWorkflowBlocked as exc:
            messages.error(request, str(exc))
            return redirect(request.POST.get("next") or "my_work_list")
        if start_work_progress(progress):
            RepairWorkflowService.mark_repair_started(
                progress.allocation.job
            )
            notify_work_progress_change(progress, "started")
        messages.success(request, "Work progress started.")
    elif action == "finish":
        try:
            RepairWorkflowService.ensure_start_allowed(
                progress.allocation.job
            )
        except RepairWorkflowBlocked as exc:
            messages.error(request, str(exc))
            return redirect(request.POST.get("next") or "my_work_list")
        if finish_work_progress(progress):
            RepairWorkflowService.mark_repair_started(
                progress.allocation.job
            )
            notify_work_progress_change(progress, "finished")
        messages.success(request, "Work progress finished.")

    save_work_progress_uploaded_photos(request, progress)

    if request.FILES.getlist("progress_photos") and action not in ["start", "finish"]:
        messages.success(request, "Progress photo(s) uploaded.")

    return redirect(request.POST.get("next") or "my_work_list")


@login_required
def unread_announcements(request):
    read_ids = AnnouncementRead.objects.filter(
        user=request.user
    ).values_list(
        "announcement_id",
        flat=True
    )

    notices = Announcement.objects.filter(
        is_active=True
    ).exclude(
        id__in=read_ids
    ).order_by("-created_at")[:3]

    data = []

    for n in notices:
        data.append({
            "id": n.id,
            "title": n.title,
            "message": n.message,
            "type": n.notice_type,
            "image": "/static/images/independence-day.svg" if "independence" in n.title.lower() else "",
        })

    return JsonResponse(data, safe=False)


@login_required
def mark_announcement_read(request, pk):
    announcement = get_object_or_404(
        Announcement,
        pk=pk
    )

    AnnouncementRead.objects.get_or_create(
        announcement=announcement,
        user=request.user
    )

    return JsonResponse({
        "status": "success"
    })
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse

from .models import (
    Customer,
    VehicleModel,
    VehicleVariant,
    InsuranceCompany,
    DriverMaster,
)


@login_required
def vehicle_form_options_api(request):
    return JsonResponse({
        "customers": [
            {
                "id": c.id,
                "name": c.name,
            }
            for c in Customer.objects.all().order_by("name")
        ],

        "models": [
            {
                "id": m.id,
                "name": m.name,
            }
            for m in VehicleModel.objects.all().order_by("name")
        ],

        "variants": [
            {
                "id": v.id,
                "model_id": v.model_id,
                "name": v.name,
            }
            for v in VehicleVariant.objects.select_related("model").order_by("name")
        ],

        "insurance_companies": [
            {
                "id": i.id,
                "name": i.ins_co_name,
            }
            for i in InsuranceCompany.objects.all().order_by("ins_co_name")
        ],

        "drivers": [
            {
                "id": d.id,
                "name": d.name,
                "type": d.get_driver_type_display(),
                "license_no": d.driving_license_no,
            }
            for d in DriverMaster.objects.filter(
                is_active=True,
                vehicle__isnull=True,
            ).order_by("name")
        ],

        "vehicle_types": [
            {"value": "PV", "label": "PV"},
            {"value": "EV", "label": "EV"},
        ],
    })

@login_required
def vehicle_form_data_api(request):
    return JsonResponse({
        "customers": list(
            Customer.objects
            .values("id", "name")
            .order_by("name")
        ),

        "models": list(
            VehicleModel.objects
            .values("id", "name")
            .order_by("name")
        ),

        "variants": list(
            VehicleVariant.objects
            .values("id", "model_id", "name")
            .order_by("name")
        ),

        "insurance_companies": list(
            InsuranceCompany.objects
            .values("id", "ins_co_name")
            .order_by("ins_co_name")
        ),

        "drivers": list(
            DriverMaster.objects
            .filter(is_active=True)
            .values(
                "id",
                "name",
                "driver_type",
                "mobile_no",
                "driving_license_no",
                "license_valid_until",
                "face_photo",
                "license_document",
                "vehicle_id",
            )
            .order_by("name")
        ),

        "vehicles": list(
            Vehicle.objects
            .values(
                "id",
                "registration_no",
                "model__name",
                "variant__name",
            )
            .order_by("registration_no")
        ),

        "vehicle_types": [
            {"value": "PV", "label": "PV"},
            {"value": "EV", "label": "EV"},
        ],
    })


from django.contrib.auth.decorators import login_required
from django.shortcuts import render, get_object_or_404

@login_required
def vehicle_new_page(request):
    return render(request, "master/vehicle_list.html")


@login_required
def vehicle_edit_page(request, pk):
    vehicle = get_object_or_404(Vehicle, pk=pk)
    return render(request, "master/vehicle_list.html", {
        "vehicle": vehicle,
    })
@login_required
def driver_master_data_api(request):
    drivers = DriverMaster.objects.select_related("vehicle").all()

    data = []

    for driver in drivers:
        data.append({
            "id": driver.id,
            "name": driver.name,
            "driver_type": driver.driver_type,
            "type": driver.get_driver_type_display(),
            "mobile": driver.mobile_no or "",
            "driving_license_no": driver.driving_license_no or "",
            "valid_until": (
                driver.license_valid_until.strftime("%d-%m-%Y")
                if driver.license_valid_until
                else ""
            ),
            "photo": (
                driver.face_photo.url
                if driver.face_photo
                else ""
            ),
            "documents": (
                driver.license_document.url
                if driver.license_document
                else ""
            ),
            "vehicle_id": driver.vehicle_id,
            "vehicle_registration": (
                driver.vehicle.registration_no
                if driver.vehicle
                else ""
            ),
            "is_active": driver.is_active,
        })

    return JsonResponse({
        "drivers": data
    })
@login_required
def driver_master_get_api(request, pk):
    driver = get_object_or_404(
        DriverMaster.objects.select_related("vehicle"),
        pk=pk
    )

    return JsonResponse({
        "id": driver.id,
        "name": driver.name,
        "driver_type": driver.driver_type,
        "mobile": driver.mobile_no or "",
        "driving_license_no": driver.driving_license_no or "",
        "license_valid_until": (
            driver.license_valid_until.strftime("%Y-%m-%d")
            if driver.license_valid_until
            else ""
        ),
        "vehicle_id": driver.vehicle_id,
        "is_active": driver.is_active,
    })
@login_required
def driver_master_save_api(request):
    if request.method != "POST":
        return JsonResponse(
            {"success": False, "error": "POST required"},
            status=405
        )

    driver_id = request.POST.get("id", "").strip()

    if driver_id:
        driver = get_object_or_404(DriverMaster, pk=driver_id)
        form = DriverMasterForm(
            request.POST,
            request.FILES,
            instance=driver
        )
    else:
        form = DriverMasterForm(
            request.POST,
            request.FILES
        )

    if form.is_valid():
        driver = form.save(commit=False)
        driver.is_active = True
        driver.save()
    
        return JsonResponse({
            "success": True,
            "id": driver.id,
            "message": "Driver saved successfully.",
        })

    print("DRIVER FORM ERRORS:", form.errors)

    return JsonResponse({
        "success": False,
        "errors": form.errors.get_json_data(),
    }, status=400)
@login_required
def dashboard_data_api(request):
    return JsonResponse({
        "vehicles": Vehicle.objects.count(),
        "drivers": DriverMaster.objects.filter(is_active=True).count(),
        "customers": Customer.objects.count(),
        "employees": Employee.objects.count(),
    })
def user_can_view_menu(user, menu):

    # =========================
    # SUPERUSER
    # =========================

    if user.is_superuser:
        return True


    # =========================
    # USER SPECIFIC PERMISSION
    # =========================

    user_permission = (
        UserMenuPermission.objects
        .filter(
            user=user,
            menu=menu
        )
        .first()
    )


    if user_permission is not None:

        return user_permission.can_view


    # =========================
    # ROLE / GROUP PERMISSION
    # =========================

    user_groups = user.groups.all()


    role_permission_exists = (
        RoleMenuPermission.objects
        .filter(
            group__in=user_groups,
            menu=menu
        )
        .exists()
    )


    if role_permission_exists:

        return (
            RoleMenuPermission.objects
            .filter(
                group__in=user_groups,
                menu=menu,
                can_view=True
            )
            .exists()
        )


    # =========================
    # NO ACCESS
    # =========================

    return False
def build_react_menu_tree(user, parent=None):

    menus = (
        Menu.objects
        .filter(parent=parent)
        .order_by("order")
    )


    result = []


    for menu in menus:

        children = build_react_menu_tree(
            user,
            parent=menu
        )


        # Check whether THIS menu is allowed
        can_view = user_can_view_menu(
            user,
            menu
        )


        # =========================
        # SHOW PARENT IF:
        #
        # 1. User has permission
        # OR
        # 2. It contains accessible children
        # =========================

        if not can_view and not children:
            continue


        result.append({

            "id": menu.id,

            "name": menu.name,

            "icon": menu.icon or "",

            "url": menu.url or "",

            "children": children,

        })


    return result
@login_required
def dashboard_data(request):

    logged_emp = Employee.objects.filter(
            user=request.user
        ).select_related("branch").first()
    
    
        # =========================
        # EMPLOYEE
        # =========================
    
    employee_name = request.user.username
    employee_type = "User"
    
    if logged_emp:
    
            if logged_emp.name:
    
                employee_name = logged_emp.name
    
            if logged_emp.employee_type:
    
                employee_type = logged_emp.employee_type
    
    
        # Superuser override
    
    if request.user.is_superuser:
    
            employee_type = "Admin"
    print("employee_type:", employee_type)
    branch_id = request.GET.get("branch")

    from_date_value = request.GET.get("from_date")

    to_date_value = request.GET.get("to_date")


    # ==========================================
    # DEFAULT DATE RANGE
    # Current month start → Today
    # ==========================================

    today = timezone.localdate()


    # ==========================================
    # FROM DATE
    # ==========================================

    if from_date_value:

        try:

            from_date = datetime.strptime(
                from_date_value,
                "%Y-%m-%d"
            ).date()

        except ValueError:

            return JsonResponse(
                {
                    "success": False,
                    "error": "Invalid from_date format."
                },
                status=400
            )

    else:

        from_date = today.replace(day=1)


    # ==========================================
    # TO DATE
    # ==========================================

    if to_date_value:

        try:

            to_date = datetime.strptime(
                to_date_value,
                "%Y-%m-%d"
            ).date()

        except ValueError:

            return JsonResponse(
                {
                    "success": False,
                    "error": "Invalid to_date format."
                },
                status=400
            )

    else:

        to_date = today


    # ==========================================
    # VALIDATE DATE RANGE
    # ==========================================

    if from_date > to_date:

        return JsonResponse(
            {
                "success": False,
                "error": (
                    "From date cannot be greater "
                    "than To date."
                )
            },
            status=400
        )


    # ==========================================
    # BASE JOBCARD QUERYSET
    # ==========================================

    jobcards = JobCard.objects.all()


    # BRANCH FILTER

    if branch_id:

        jobcards = jobcards.filter(
            branch_id=branch_id
        )


    # DATE FILTER

    jobcards = jobcards.filter(

        job_date__date__range=(

            from_date,

            to_date

        )

    )


    # ==========================================
    # ACTIVE JOB CARDS
    #
    # Open job cards within selected filters
    # ==========================================

    active_job_cards = jobcards.filter(

        repair_status="Open"

    ).count()


    # ==========================================
    # ACTIVE CLAIMS
    # ==========================================

    claims = Claim.objects.all()


    # BRANCH FILTER

    if branch_id:

        claims = claims.filter(
            branch_id=branch_id
        )


    # DATE FILTER
    #
    # Using created_at because claims are created
    # within the selected date range.
    # ==========================================

    claims = claims.filter(

        created_at__date__range=(

            from_date,

            to_date

        )

    )


    # ACTIVE CLAIMS
    #
    # Open claims only
    # ==========================================

    active_claims = claims.exclude(

        status__in=[

            "Closed",

            "Cancelled"

        ]

    ).count()


    # ==========================================
    # VEHICLES IN WORKSHOP
    #
    # Jobcards that are currently Open
    # ==========================================

    vehicles_in_workshop = jobcards.filter(

        repair_status="Open"

    ).count()


    # ==========================================
    # PENDING DELIVERY
    #
    # Repair completed but vehicle not yet
    # actually delivered.
    # ==========================================

    pending_delivery = jobcards.filter(

        repair_status="Completed",

        actual_delivery__isnull=True

    ).count()

    # ==================================================
    # JOB CARD / CLAIM PIPELINE
    # ==================================================

    pipeline_claims = claims.exclude(

    status__in=[

        "Closed",

        "Cancelled"

    ]

    )


    # ==========================================
    # JOB CARD PIPELINE
    # ==========================================

        # ==================================================
    # JOB CARD PIPELINE
    # ==================================================


    # ------------------------------------------
    # TOTAL JOBCARDS IN SELECTED PERIOD
    # ------------------------------------------

    total_job_cards = jobcards.count()


    # ------------------------------------------
    # DELIVERED
    #
    # Claim stage 14 = Closed
    # OR JobCard repair status = Closed
    # ------------------------------------------

    delivered_jobs = jobcards.filter(

        claim__claim_stage=14

    ).distinct()


    delivered_ids = list(
        delivered_jobs.values_list("id", flat=True)
    )


    # ------------------------------------------
    # READY
    #
    # Claim stages:
    #
    # 9  = Work Completed
    # 10 = Re Inspection
    # 11 = Liability
    # 12 = Invoiced
    # 13 = Delivery
    #
    # Exclude already delivered jobs
    # ------------------------------------------

    ready_jobs = jobcards.filter(

        claim__claim_stage__in=[
            9,
            10,
            11,
            12,
            13
        ]

    ).exclude(

        id__in=delivered_ids

    ).distinct()


    ready_ids = list(
        ready_jobs.values_list("id", flat=True)
    )


    # ------------------------------------------
    # REPAIR
    #
    # Claim stage 8
    # Repair Work In Progress
    # ------------------------------------------

    repair_jobs = jobcards.filter(

        claim__claim_stage=8

    ).exclude(

        id__in=delivered_ids + ready_ids

    ).distinct()


    repair_ids = list(
        repair_jobs.values_list("id", flat=True)
    )


    # ------------------------------------------
    # PARTS
    #
    # Jobs having active parts orders
    # ------------------------------------------

    parts_jobs = jobcards.filter(

        part_order_headers__status__in=[

            "Pending",

            "Order Placed",

            "In Transit",

            "Partially Received",

            "Back Order",

        ]

    ).exclude(

        id__in=delivered_ids + ready_ids + repair_ids

    ).distinct()


    parts_ids = list(
        parts_jobs.values_list("id", flat=True)
    )


    # ------------------------------------------
    # APPROVAL
    #
    # Claim workflow stages:
    #
    # 4 = Claim Intimation
    # 5 = Survey Done
    # 6 = Insurance Approval
    # ------------------------------------------

    approval_jobs = jobcards.filter(

        claim__claim_stage__in=[
            4,
            5,
            6
        ]

    ).exclude(

        id__in=(
            delivered_ids +
            ready_ids +
            repair_ids +
            parts_ids
        )

    ).distinct()


    approval_ids = list(
        approval_jobs.values_list("id", flat=True)
    )


    # ------------------------------------------
    # ASSESSMENT
    #
    # Claim workflow stages:
    #
    # 1 = Claim Created
    # 2 = Advisor Assigned
    # 3 = Estimate Created
    # ------------------------------------------

    assessment_jobs = jobcards.filter(

        claim__claim_stage__in=[
            1,
            2,
            3
        ]

    ).exclude(

        id__in=(
            delivered_ids +
            ready_ids +
            repair_ids +
            parts_ids +
            approval_ids
        )

    ).distinct()


    assessment_ids = list(
        assessment_jobs.values_list("id", flat=True)
    )


    # ------------------------------------------
    # GATE IN
    #
    # Remaining Job Cards
    #
    # This ensures every Job Card belongs
    # to exactly ONE pipeline stage.
    # ------------------------------------------

    assigned_ids = (

        delivered_ids +

        ready_ids +

        repair_ids +

        parts_ids +

        approval_ids +

        assessment_ids

    )


    gate_in_jobs = jobcards.exclude(

        id__in=assigned_ids

    )


    # ==================================================
    # FINAL PIPELINE DATA
    # ==================================================

    job_card_pipeline = {

        "gate_in":
            gate_in_jobs.count(),

        "assessment":
            assessment_jobs.count(),

        "approval":
            approval_jobs.count(),

        "parts":
            parts_jobs.count(),

        "repair":
            repair_jobs.count(),

        "ready":
            ready_jobs.count(),

        "delivered":
            delivered_jobs.count(),

    }


    # ==================================================
    # SECTION 4 — REVENUE
    # ==================================================

   


    revenue_data = jobcards.aggregate(

        total_revenue=Coalesce(
            Sum("grand_total"),
            Decimal("0.00")
        ),

        parts_revenue=Coalesce(
            Sum("parts_total"),
            Decimal("0.00")
        ),

        labour_revenue=Coalesce(
            Sum("labour_total"),
            Decimal("0.00")
        )

    )


    total_revenue = revenue_data["total_revenue"]

    parts_revenue = revenue_data["parts_revenue"]

    labour_revenue = revenue_data["labour_revenue"]
    # ==================================================
    # SECTION 4A — REVENUE TREND
    # ==================================================

    revenue_trend_queryset = (

    jobcards

    .annotate(
        month=TruncMonth("job_date")
    )

    .values("month")

    .annotate(
        revenue=Coalesce(
            Sum("grand_total"),
            Decimal("0.00")
        )
    )

    .order_by("month"))


    revenue_trend = []


    for item in revenue_trend_queryset:

        revenue_trend.append({

        "month":
            item["month"].strftime("%b %Y"),

        "revenue":
            float(item["revenue"])

        })
    total_pipeline_jobs = jobcards.count()
        # ==========================================
        # AI INSPECTIONS
        #
        # Temporary until we connect the exact
        # AI inspection model.
        # ==========================================

    ai_inspections = 0


        # ==========================================
        # RESPONSE
        # ==========================================

    return JsonResponse({

        "success": True,

        "dashboard_type": employee_type,


        "filters": {

            "branch": branch_id,

            "from_date":
                from_date.isoformat(),

            "to_date":
                to_date.isoformat(),

        },


        # ==========================================
        # OVERVIEW
        # ==========================================

        "active_job_cards":
            active_job_cards,

        "active_claims":
            active_claims,

        "vehicles_in_workshop":
            vehicles_in_workshop,

        "pending_delivery":
            pending_delivery,


        # ==========================================
        # JOB CARD PIPELINE
        # ==========================================

       "job_card_pipeline": job_card_pipeline,

        "total_pipeline_jobs": total_pipeline_jobs,
        "revenue": {

        "total": float(total_revenue),

        "parts": float(parts_revenue),

        "labour": float(labour_revenue),},
        "revenue_trend": revenue_trend,

        })
@login_required
def advisor_dashboard_data(request):

    logged_emp = Employee.objects.filter(
        user=request.user
    ).first()

    if not logged_emp:

        return JsonResponse({
            "success": False,
            "message": "Employee record not found"
        })


    advisor_jobs = JobCard.objects.filter(
        advisor=logged_emp
    ).select_related(
        "vehicle",
        "claim"
    )


    assigned_jobs = advisor_jobs.count()

    pending_jobs = advisor_jobs.filter(
        repair_status="Open"
    ).count()

    completed_jobs = advisor_jobs.filter(
        repair_status="Completed"
    ).count()


    ready_jobs = advisor_jobs.filter(
        ready_for_delivery=True
    ).count()


    recent_jobs = advisor_jobs.order_by(
        "-created_at"
    )[:3]


    recent_work = []

    for job in recent_jobs:

        recent_work.append({

            "job_no": job.job_no,

            # Change vehicle field if needed
            "vehicle_no": str(job.vehicle) if job.vehicle else "-",

            "claim_no": (
                str(job.claim)
                if job.claim
                else "-"
            ),

            "status": job.repair_status,

            "progress": 100
            if job.repair_status == "Completed"
            else 50,

        })


    completion_percentage = 0

    if assigned_jobs > 0:

        completion_percentage = round(
            (completed_jobs / assigned_jobs) * 100
        )


    return JsonResponse({

        "success": True,

        "welcome_message":
            "Have a productive day!",


        "assigned_jobs": assigned_jobs,

        "pending_jobs": pending_jobs,

        "completed_jobs": completed_jobs,

        "ready_jobs": ready_jobs,


        "performance": {

            "completion_percentage":
                completion_percentage,

            "completed":
                completed_jobs,

            "running":
                pending_jobs,

            "pending":
                pending_jobs,

        },


        "recent_work":
            recent_work,

    })
@login_required
def header_data(request):

    logged_emp = Employee.objects.filter(
        user=request.user
    ).select_related("branch").first()


    # =========================
    # EMPLOYEE
    # =========================

    employee_name = request.user.username
    employee_type = "User"

    if logged_emp:

        if logged_emp.name:

            employee_name = logged_emp.name

        if logged_emp.employee_type:

            employee_type = logged_emp.employee_type


    # Superuser override

    if request.user.is_superuser:

        employee_type = "Admin"


    # =========================
    # COMPANY
    # =========================

    company_name = ""
    company_logo = ""

    company = CompanySetup.objects.first()

    if company:

        company_name = company.company_name or ""

        if company.logo:

            company_logo = company.logo.url


    # =========================
    # BRANCH
    # =========================

    branch_name = ""

    if logged_emp and logged_emp.branch_id:

        branch_name = logged_emp.branch.name


    # =========================
    # RESPONSE
    # =========================

    return JsonResponse({

        "success": True,

        "company": {

            "name": company_name,

            "logo": company_logo,

        },

        "branch": {

            "name": branch_name,

        },

        "employee": {

            "name": employee_name,

            "type": employee_type,

        },

    })
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse

from .models import Branch


@login_required
def dashboard_branches(request):

    branches = (
        Branch.objects
        .filter(is_active=True)
        .order_by("name")
    )

    data = [

        {
            "id": branch.id,
            "name": branch.name,
            "code": branch.code,
        }

        for branch in branches

    ]

    return JsonResponse({

        "success": True,

        "branches": data,

    })